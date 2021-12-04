import utils_IRdrone as IRd
import numpy as np
import glob
import matplotlib.pyplot as plt
import matplotlib.transforms as mtrans
import irdrone.process as pr
from datetime import timedelta
from irdrone.utils import Style
import os
import os.path as osp
import cv2
import irdrone.utils as ut
import datetime
import scipy.fft
import openpyxl


def readFlightSummary(dir_mission, mute=None):
    """
        Read the FlightSummary  in Excel file.

    :param dir_mission: chemin du dossier qui contient les données de la mission  type= string
    :param mute: affiche des informations si True.                                type= bool
    :return: listSummaryFlight  liste de données pour chaque couple d'images VIS-NIR (timeDeviation, altitude, etc)
    """
    summaryPath = osp.join(dir_mission, "FlightSummary.xlsx")
    txt = 'lecture des données de  ' + summaryPath
    print(Style.CYAN +  txt + Style.RESET)

    workbook = openpyxl.load_workbook(summaryPath, read_only=True, data_only=True)
    sheet = workbook['Summary']
    listSummaryFlight =[]
    nulg = 2                                                # première ligne de données
    while sheet.cell(nulg, 1).value:
        timeDeviation = float(sheet.cell(nulg, 4).value)    # ecart temporel entre image NIR et VIS (en s)
        date_string = str(sheet.cell(nulg, 8).value)        # date image VIS (rectifiee si time-lapse)  format string
        altiDrone2Sol = float(sheet.cell(nulg, 11).value)   # altitude du drone par rapport au sol (en m)
        distLastPt = float(sheet.cell(nulg, 14).value)      # distance au point suivant (en m)
        capLastPt = float(sheet.cell(nulg, 15).value)       # cap au point suivant (en °)
        gpsLat = float(sheet.cell(nulg, 9).value)           # latitude  dd.dddddd°
        gpsLong = float(sheet.cell(nulg, 10).value)         # longitude dd.dddddd°
        xUTM = float(sheet.cell(nulg, 17).value)            # coordonnee UTM x
        yUTM = float(sheet.cell(nulg, 18).value)            # coordonnee UTM y
        data = timeDeviation, altiDrone2Sol, distLastPt, capLastPt, gpsLat, gpsLong, xUTM, yUTM, date_string
        listSummaryFlight.append(data)
        nulg = nulg + 1
    workbook.close()

    if not mute:
        print(listSummaryFlight)

    return listSummaryFlight

def pitchDeviation(listSummaryFlight, motion_list_drone, motion_list_cameraDJI):
    speed = 0.5                              # vitesse du drone   en m/s
    txt = ' ...  Vitesse drone ' + str(speed) + ' m/s'
    print(Style.YELLOW + txt + Style.RESET)
    timelapse_Vis = 2.                       # période du timalapse de la caméra VIS

    pitch_Theorique=[]

    if len(listSummaryFlight) != len(motion_list_drone):
        print('BUG   Il faut relancer le process !  ', len(listSummaryFlight), len(motion_list_drone))

    for i in range(len(listSummaryFlight)):
        # interpolation linéaire du pitch du drone à l'instant où l'image NIR a ete prise
        dt = float(listSummaryFlight[i][0])
        if i == 1 or i == len(listSummaryFlight) - 1:
            alpha = motion_list_drone[i][2]
        elif dt < 0:
            alpha = (motion_list_drone[i][2]*(dt/timelapse_Vis +1)- motion_list_drone[i+1][2] * dt/timelapse_Vis)
        else:
            alpha = (motion_list_drone[i-1][2] * dt/timelapse_Vis - motion_list_drone[i][2]*(dt/timelapse_Vis -1))

        dateImgVis = listSummaryFlight[i][8]
        dateImgVis = IRd.dateExcelString2Py(dateImgVis)
        date = dateImgVis  #motion_list_drone[i][0]
        CnirCvis = speed * dt
        H = float(listSummaryFlight[i][1])
        thetaVis = motion_list_cameraDJI[i][2]+90.
        thetaNir = alpha
        #print("Theta NIR  ", thetaNir, "Theta VIS  ", thetaVis)
        anglePhi = np.rad2deg(np.arctan(CnirCvis/H + np.tan(np.deg2rad(thetaVis))))

        anglePsi = anglePhi - (alpha )
        #print("CnirCvis  ", CnirCvis, " H   ", H, "anglePhi ", anglePhi , "anglePsi   ",anglePsi)
        data = date, anglePsi
        pitch_Theorique.append(data)

    return pitch_Theorique

def decompositionHomography(mouvement, date, cal, motion_list_fin_a, translat_fin_a, normal_fin_a, motion_list_fin_b, translat_fin_b, normal_fin_b):
    solution = cv2.decomposeHomographyMat(mouvement["homography"], cal["mtx"])
    for idx in range(4):
        extra_pitch, extra_yaw, extra_roll = np.rad2deg(cv2.Rodrigues(solution[1][idx])[0].flatten())
        extra_tx, extra_ty, extra_tz = (solution[2][idx])
        extra_nx, extra_ny, extra_nz = (solution[3][idx])
        print('Solution N°', idx + 1, '\n',
              'pitch = %.4f°  yaw = %.4f°  roll = %.4f°  ' % (extra_pitch, extra_yaw, extra_roll))
        # print('  R', idx, '= ', solution[1][idx],'\n')
        print('  tx = %.4f  ty = %.4f  tz = %.4f  ' % (extra_tx, extra_ty, extra_tz))
        print('  nx = %.4f  ny = %.4f  nz = %.4f   ||n|| =%.4f'
              % (extra_nx, extra_ny, extra_nz, (extra_nx ** 2 + extra_ny ** 2 + extra_nz ** 2) ** 0.5))

        # break
    # WARNING! Choosing the first solution here, not sure it's the right one!!!
    extra_pitch, extra_yaw, extra_roll = np.rad2deg(cv2.Rodrigues(solution[1][0])[0].flatten())
    extra_tx, extra_ty, extra_tz = (solution[2][0])
    extra_nx, extra_ny, extra_nz = (solution[3][0])
    motion_list_fin_a.append([date, extra_yaw, extra_pitch, extra_roll])
    translat_fin_a.append([date, extra_tx, extra_ty, extra_tz])
    normal_fin_a.append([date, extra_nx, extra_ny, extra_nz])

    extra_pitch, extra_yaw, extra_roll = np.rad2deg(cv2.Rodrigues(solution[1][2])[0].flatten())
    extra_tx, extra_ty, extra_tz = (solution[2][2])
    extra_nx, extra_ny, extra_nz = (solution[3][2])
    motion_list_fin_b.append([date, extra_yaw, extra_pitch, extra_roll])
    translat_fin_b.append([date, extra_tx, extra_ty, extra_tz])
    normal_fin_b.append([date, extra_nx, extra_ny, extra_nz])

    return motion_list_fin_a, translat_fin_a, normal_fin_a, motion_list_fin_b, translat_fin_b, normal_fin_b


def plotYawPitchRollDroneAndCameraDJI(dir_mission, utilise_cache=False):
    #
    # Le 7 septembre 2021 la SJCam M20 a été équipée d'une cale de +3,1°  (+ <=> vision "vers l'avant")
    date_cale = datetime.date(2021,9,7)
    angle_wedge = 3.1  #angle of the wedge  in degre
    #

    # Estimation du défaut d'alignement de l'axe de visée caméra SJCam M20  après le 7 septembre 2021
    corrige_defaut_axe_visee = True
    if corrige_defaut_axe_visee:
        offsetPitch = - 2   # défaut d'alignement (pitch) de l'axe de visée de la caméra NIR  en °
        offsetYaw = 0.87       # défaut d'alignement (yaw) de l'axe de visée de la caméra NIR    en °
    else:
        offsetPitch = 0
        offsetYaw = 0

    motion_list, motion_list_drone, motion_list_cameraDJI = [], [], []
    motion_list_fin_a, translat_fin_a, normal_fin_a = [], [], []
    motion_list_fin_b, translat_fin_b, normal_fin_b = [], [], []
    motion_nul = []

    image_dir = osp.join(dir_mission, "AerialPhotography")
    result_dir = osp.join(dir_mission, "ImgIRdrone")
    desync = timedelta(seconds=-0)
    txt = str('Delay  ' + str(desync) + "s")
    print(Style.YELLOW + txt + Style.RESET)
    missionTitle= dir_mission.split("/")[-1]

    motion_cache = osp.join(result_dir, "motion_summary_vs_drone_dates.npy")
    count1 = 0
    count2 = 0
    # Chargement des paramètres intrinsèques de la caméra NIR. On récupère en particulier la matrice K
    cal = ut.cameracalibration("DJI_Raw")

    if not osp.exists(motion_cache) or not utilise_cache :

        listSummaryFlight = readFlightSummary(dirMission, mute=True)

        for ipath in glob.glob(osp.join(image_dir, "*.DNG")):
            try:
                count1 += 1
                mpath = glob.glob(osp.join(result_dir, osp.basename(ipath)[:-4]) + "*.npy")[0]
                img = pr.Image(ipath)
                print(Style.CYAN + img.name + Style.RESET)
                finfo = img.flight_info

                #  Todo securiser  cette partie   (si nb d'image DNG différent de celui dans flight summary)
                date = listSummaryFlight[count1-1][8]
                date = IRd.dateExcelString2Py(date)
                #date = img.date


                # roll, pitch & yaw   drone    (NIR camera)
                mouvement_drone = [date, finfo["Flight Roll"], finfo["Flight Pitch"], finfo["Flight Yaw"]]
                motion_list_drone.append(mouvement_drone)
                # roll, pitch & yaw   gimbal  (VIS camera)
                mouvement_cameraDJI = [date, finfo["Gimbal Roll"], finfo["Gimbal Pitch"], finfo["Gimbal Yaw"]]
                motion_list_cameraDJI.append(mouvement_cameraDJI)
                #  yaw & pitch  image NIR recalage grossier
                mouvement = np.load(mpath, allow_pickle=True).item()
                motion_list.append([date, mouvement["yaw"], mouvement["pitch"]])

                #decompositionHomography(mouvement, date, cal, motion_list_fin_a, translat_fin_a, normal_fin_a, motion_list_fin_b,
                #                        translat_fin_b, normal_fin_b)

            except:
                count2 += 1
                continue

        motion_list = np.array(motion_list)
        motion_list_fin_a = np.array(motion_list_fin_a)
        translat_fin_a = np.array(translat_fin_a)
        normal_fin_a = np.array(normal_fin_a)
        motion_list_fin_b = np.array(motion_list_fin_b)
        translat_fin_b = np.array(translat_fin_b)
        normal_fin_b = np.array(normal_fin_b)
        motion_list_drone = np.array(motion_list_drone)
        motion_list_cameraDJI = np.array(motion_list_cameraDJI)
        listSummaryFlight = np.array(listSummaryFlight)

        # calcul pitch théorique
        pitch_Theorique = pitchDeviation(listSummaryFlight, motion_list_drone, motion_list_cameraDJI)
        pitch_Theorique = np.array(pitch_Theorique)



        # sauvegarde des mouvements du drone et des angles de correction
        np.save(motion_cache[:-4], {"motion_list": motion_list,
                                    "motion_list_drone": motion_list_drone,
                                    "motion_list_cameraDJI": motion_list_cameraDJI,
                                    "listSummaryFlight": listSummaryFlight,
                                    "pitch_Theorique":pitch_Theorique,
                                    "motion_list_fin_a": motion_list_fin_a,
                                    "translat_fin_a": translat_fin_a,
                                    "normal_fin_a": normal_fin_a,
                                    "motion_list_fin_b": motion_list_fin_b,
                                    "translat_fin_b": translat_fin_b,
                                    "normal_fin_b": normal_fin_b
                                    })
        txt = "images retenues" + str(count1 - count2) + "   images rejettées  " + str(count2)
        print(Style.CYAN + txt + Style.RESET)

    else:
        # utilisation des des mouvements du drone et des angles de correction déjà sauvegardés
        print(Style.YELLOW + "Warning : utilisation du cache" + Style.RESET)
        m_cache = np.load(motion_cache, allow_pickle=True).item()
        motion_list = m_cache["motion_list"]
        motion_list_drone = m_cache["motion_list_drone"]
        motion_list_cameraDJI = m_cache["motion_list_cameraDJI"]
        listSummaryFlight = m_cache["listSummaryFlight"]
        pitch_Theorique = m_cache["pitch_Theorique"]
        motion_list_fin_a = m_cache["motion_list_fin_a"]
        translat_fin_a = m_cache["translat_fin_a"]
        normal_fin_a = m_cache["normal_fin_a"]
        motion_list_fin_b = m_cache["motion_list_fin_b"]
        translat_fin_b = m_cache["translat_fin_b"]
        normal_fin_b = m_cache["normal_fin_b"]

    dates = motion_list[:, 0]
    # détermination de l'offset théorique de l'image NIR
    date_mission=datetime.date(dates[0].year,dates[0].month,dates[0].day)

    if date_mission > date_cale:
        print(Style.YELLOW + "présence d'une cale de 3,1° sous la SJCam M20" + Style.RESET)
        calePitchM20 = angle_wedge
    else :
        calePitchM20 = 0
    print('average  flight_Roll %.2f°  Gimbal_Roll %.2f°' %
              (np.average(motion_list_drone[:, 1]), np.average(motion_list_cameraDJI[:, 1])))
    print('average  flight_Yaw %.2f°  Gimbal_Yaw %.2f°' %
          (np.average(motion_list_drone[:, 3]), np.average(motion_list_cameraDJI[:, 3])))
    print('average  flight_Pitch %.2f°  Gimbal_Pitch %.2f°'%
          (np.average(motion_list_drone[:,2]),np.average( motion_list_cameraDJI[:,2])))
    print(' average delta Roll Drone/Gimbal  %.2f'%
          np.average((-motion_list_drone[:,1]+motion_list_cameraDJI[:,1])))
    print(' average delta Yaw Drone/Gimbal  %.2f' %
          np.average((-motion_list_drone[:, 3] + motion_list_cameraDJI[:, 3])))

    #  Tracé des courbes
    #
    #    Yaw Pitch Roll

    if True:
        YawPitchTeoriticalAndCoarse_plot(motion_list,
                                     motion_list_cameraDJI,
                                     motion_list_drone,
                                     pitch_Theorique,
                                     offsetPitch,
                                     offsetYaw,
                                     missionTitle)


    Pitch_plot(motion_list, motion_list_cameraDJI, motion_list_drone, pitch_Theorique, offsetPitch, missionTitle)



    DeltaYawPitchTeoriticalAndCoarse_plot(motion_list,
                                     motion_list_cameraDJI,
                                     motion_list_drone,
                                     pitch_Theorique,
                                     offsetPitch,
                                     offsetYaw,
                                     missionTitle)

    # Analyse de Fourier du Pitch
    '''
    signal = motion_list[:, 2]
    Fourier_plot(signal, title1='Pitch Image NIR  ', title2='', color='orange')
    signal = (- (motion_list_cameraDJI[:, 2] + 90.))
    Fourier_plot(signal, title1='Pitch_Camera DJI  ', title2='', color='r')
    signal = motion_list_drone[:, 2]
    Fourier_plot(signal, title1='Pitch_Camera NIR  ', title2='', color='g')
    '''
    signal = motion_list[:, 2]
    Fourier_plot(signal, title1='NIR Pitch   (coarse process)', title2='', color='black')

    signal = (motion_list_cameraDJI[:, 2] + 90. - motion_list_drone[:, 2])
    Fourier_plot(signal, title1='Theoritical Pitch  synchro', title2='', color='orange')

    signal = pitch_Theorique[:, 1] - offsetPitch
    Fourier_plot(signal, title1='Theoritical Pitch  un-synchro', title2='', color='m')

    """
    signal = motion_list[:, 1]
    Fourier_plot(signal, title1='NIR Yaw    (coarse process)', title2='', color='c')


    signal = motion_list_drone[:, 1]
    Fourier_plot(signal, title1='Theoritical Yaw  synchro', title2='', color='b')
    """

    # Etude de la dispersion  Pitch et Yaw des caméras (donc des images)
    for i in range(len(motion_list)):
        motion_nul.append([0,0,0,0])

    comparaisonPitchYaw_plot(motion_list,
                                    motion_list_drone,
                                    motion_list_cameraDJI,
                                    missionTitle,
                                    pitch_Theorique,
                                    offsetPitch,
                                    offsetYaw,
                                    process='Coarse ',
                                    motion_refined= np.array(motion_nul))
    """
    comparaisonPitchYaw_plot(motion_list,
                                    motion_list_drone,
                                    motion_list_cameraDJI,
                                    missionTitle,
                                    pitch_Theorique,
                                    0*calePitchM20,
                                    process='Refined ',
                                    motion_refined =motion_list_fin_b)
    """


def Pitch_plot(motion_list, motion_list_cameraDJI, motion_list_drone, pitch_Theorique, offsetPitch, missionTitle):
    plt.title(missionTitle )

    plt.plot(motion_list[:, 0], motion_list[:, 2],
             color='black', linestyle='-', linewidth=1,
             label="pitch_imgNIR       (coarse process)    average = {:.2f}°"
             .format(np.average(motion_list[:, 2], axis=0)))
    plt.plot(motion_list[:, 0], motion_list_cameraDJI[:, 2] + 90.- motion_list_drone[:, 2],
             color='orange', linestyle=':', linewidth=1,
             label="Theoretical pitch_imgNIR synchro   average = {:.2f}°"
             .format(np.abs(-np.average(motion_list_cameraDJI[:, 2] + 90. - motion_list_drone[:, 2], axis=0))))

    plt.plot(motion_list[:, 0], pitch_Theorique[:, 1] - offsetPitch,
             color='magenta', linestyle='-', linewidth=1,
             label="Theoretical pitch_imgNIR  unsynchro  average = {:.2f}°"
             .format(np.abs(np.average(pitch_Theorique[:, 1] - offsetPitch, axis=0))))
    plt.grid()
    plt.legend()
    plt.show()



def YawPitchTeoriticalAndCoarse_plot(motion_list,
                                     motion_list_cameraDJI,
                                     motion_list_drone,
                                     pitch_Theorique,
                                     offsetPitch,
                                     offsetYaw,
                                     missionTitle):
    plt.title(missionTitle )
    plt.plot(motion_list[:, 0], motion_list[:, 1], "b--",
             label="yaw_imgNIR       (coarse process)    average = {:.2f}°".format(
                 np.average(motion_list[:, 1], axis=0)))
    plt.plot(motion_list[:, 0] , offsetYaw - motion_list_drone[:, 1] , "c-",
             label="Theoretical yaw_imgNIR (roll_drone)  average = {:.2f}°"
             .format(np.abs(np.average( offsetYaw - motion_list_drone[:, 1], axis=0))))
    plt.plot(motion_list[:, 0], motion_list[:, 2], color='black', linestyle='-', linewidth=0.8,
             label="pitch_imgNIR       (coarse process)    average = {:.2f}°"
             .format(np.average(motion_list[:, 2], axis=0)))
    """
    plt.plot(motion_list[:, 0] + desync,
             motion_list_cameraDJI[:, 2] + 90. - calePitchM20 - motion_list_drone[:, 2], "m-",
             label="Theoretical pitch_imgNIR  (90+pitch gimbal-pitch drone)  average = {:.2f}°"
             .format(np.abs(-np.average(motion_list_cameraDJI[:, 2] + 90.
                                        - calePitchM20 - motion_list_drone[:, 2], axis=0))))
    """
    plt.plot(motion_list[:, 0], pitch_Theorique[:, 1] - offsetPitch, "m-",
             label="Theoretical pitch_imgNIR  new  average = {:.2f}°"
             .format(np.abs(np.average(pitch_Theorique[:, 1] - offsetPitch, axis=0))))
    plt.plot(motion_list[:, 0], motion_list_drone[:, 3] - motion_list_cameraDJI[:, 3], "g-",
             label="Yaw Drone - Yaw Gimball           average = {:.2f}°"
             .format(np.abs(np.average( motion_list_drone[:, 3]-motion_list_cameraDJI[:, 3], axis=0))))
    plt.grid()
    plt.legend()
    plt.show()


def DeltaYawPitchTeoriticalAndCoarse_plot(motion_list,
                                     motion_list_cameraDJI,
                                     motion_list_drone,
                                     pitch_Theorique,
                                     offsetPitch,
                                     offsetYaw,
                                     missionTitle):
    plt.title(missionTitle )
    plt.plot(motion_list[:, 0], motion_list[:, 1] - (offsetYaw - motion_list_drone[:, 1]), "b--",
             label="delta yaw_imgNIR      average = {:.2f}°".format(
                 np.average(motion_list[:, 1] - (offsetYaw - motion_list_drone[:, 1]), axis=0)))
    """
    plt.plot(motion_list[:, 0], motion_list[:, 2] -
             (motion_list_cameraDJI[:, 2] + 90. - calePitchM20 - motion_list_drone[:, 2]), "r--",
             label="delta pitch_imgNIR    average = {:.2f}°"
             .format(np.average(motion_list[:, 2] -
             (motion_list_cameraDJI[:, 2] + 90. - calePitchM20 - motion_list_drone[:, 2]), axis=0)))
    """
    plt.plot(motion_list[:, 0], motion_list[:, 2] - (pitch_Theorique[:, 1] - offsetPitch), "r--",
             label="delta pitch_imgNIR    average = {:.2f}°"
             .format(np.average(motion_list[:, 2] - (pitch_Theorique[:, 1] - offsetPitch), axis=0)))


    plt.plot(motion_list[:, 0],
              motion_list_drone[:, 3] - motion_list_cameraDJI[:, 3], "g-",
             label="delta Roll_imgNIR     average = {:.2f}°"
             .format(np.abs(np.average( motion_list_drone[:, 3]-motion_list_cameraDJI[:, 3], axis=0))))
    plt.grid()
    plt.legend()
    plt.show()

def decomposeHomographyRefined_Plot(motion_list_fin_a, motion_list_fin_b,
                                    translat_fin_a, translat_fin_b,
                                    normal_fin_a, normal_fin_b,
                                    missionTitle):
    # Courbe d'evolution des angles fin solution a et b)
        print('Average refined yaw solution a = ', np.average(motion_list_fin_a[:, 1]),
              ' solution b = ',np.average(motion_list_fin_b[:, 1]))
        print('Average refined pitch solution a = ', np.average(motion_list_fin_a[:, 2]),
              ' solution b = ', np.average(motion_list_fin_b[:, 2]))
        print('Average refined roll solution a = ', np.average(motion_list_fin_a[:, 3]),
              ' solution b = ', np.average(motion_list_fin_b[:, 3]))
        plt.title(missionTitle)
        plt.plot(motion_list_fin_a[:, 1], "m-", label="pitch_a")
        plt.plot(motion_list_fin_b[:, 1], "m-.", label="pitch_b")
        plt.plot(motion_list_fin_a[:, 2], "r-", label="yaw_a")
        plt.plot(motion_list_fin_b[:, 2], "r-.",label="yaw_b")
        plt.plot(motion_list_fin_a[:, 3], "b-", label="roll_a")
        plt.plot(motion_list_fin_b[:, 3], "b-.",label="roll_b")
        plt.legend()
        plt.grid()
        plt.show()

    # Courbe d'evolution des translation fine solution a et b)
        plt.title(missionTitle)
        plt.plot(translat_fin_a[:, 1], "m-", label="tx_a")
        #plt.plot(translat_fin_b[:, 1], "m-.", label="tx_b")
        plt.plot(translat_fin_a[:, 2], "r-", label="ty_a")
        #plt.plot(translat_fin_b[:, 2], "r-.",label="ty_b")
        plt.plot(translat_fin_a[:, 3], "b-", label="tz_a")
        #plt.plot(translat_fin_b[:, 3], "b-.",label="tz_b")
        plt.legend()
        plt.grid()
        plt.show()

        # Courbe d'evolution des normale fine solution a et b)
        normal_fin_a = np.array(normal_fin_a)
        normal_fin_b = np.array(normal_fin_b)
        plt.title(missionTitle)
        plt.plot(normal_fin_a[:, 1], "m-", label="nx_a")
        #plt.plot(normal_fin_b[:, 1], "m-.", label="nx_b")
        plt.plot(normal_fin_a[:, 2], "r-", label="ny_a")
        #plt.plot(normal_fin_b[:, 2], "r-.", label="ny_b")
        plt.plot(abs(normal_fin_a[:, 3]), "b-", label="| nz_a |")
        #plt.plot(normal_fin_b[:, 3], "b-.", label="nz_b")
        plt.legend()
        plt.grid()
        plt.show()

def comparaisonPitchYaw_plot(motion_list,
                                    motion_list_drone,
                                    motion_list_cameraDJI,
                                    missionTitle,
                                    pitch_Theorique,
                                    offsetPitch,
                                    offsetYaw,
                                    process='',
                                    motion_refined=None):
    maxiIRrefined = (np.max(((motion_list[:, 1] + offsetYaw - motion_refined[:, 1]), -motion_list_drone[:, 1])),
                     np.max(((motion_list[:, 2] - motion_refined[:, 2]), (pitch_Theorique[:, 1] - offsetPitch))))
    miniIRrefined = (np.min(((motion_list[:, 1] + offsetYaw - motion_refined[:, 1]), -motion_list_drone[:, 1])),
                     np.min(((motion_list[:, 2] - motion_refined[:, 2]), (pitch_Theorique[:, 1] - offsetPitch))))

    _, (ax1, ax2) = plt.subplots(nrows=1, ncols=2)
    # Courbe de dispersion des angles   pitch(yaw)
    myTitle='Theoritical un-synchronized.   Offset : Pitch =' + str(offsetPitch) + '°  | Yaw = '+ str(offsetYaw) + '°'
    disperPitchYaw_plot(ax1, offsetYaw - motion_list_drone[:, 1],
                        (pitch_Theorique[:, 1] - offsetPitch),
                        miniIRrefined, maxiIRrefined,
                        missionTitle, myTitle, color='green')
    # Courbe de dispersion des angles de reclage de l'mage NIR sur l'image VIS  (pitch et yaw "refined")
    # ATTENTION: pour redresser l'image NIR on applique une rotation avec un pitch_NIR (resp yaw_NIR)
    # de signe opposé à celui du pitch (resp yaw) de la M20.
    disperPitchYaw_plot(ax2, motion_list[:, 1] - motion_refined[:, 1], motion_list[:, 2] - motion_refined[:, 2],
                        miniIRrefined, maxiIRrefined,
                        missionTitle, process, color='orange')

    plt.tight_layout()
    plt.show()


def disperPitchYaw_plot(ax, fx, fy,  mini, maxi, missionTitle, spectralType, color='yellow'):
    avg = (np.average(fx, axis=0), np.average(fy, axis=0))
    std = (np.std(fx, axis=0), np.std(fy, axis=0))
    ax.plot(fx, fy,
             marker="o", markersize=6, alpha=0.4,  mec=color, mfc=color, linestyle='')
    ax.set_xlim(mini[0]-2., maxi[0]+2.)
    ax.set_ylim(mini[1]-2., maxi[1]+2.)
    ax.plot(0., 0., "k+")
    ax.plot(avg[0], avg[1], "r+")
    myTitle = str(missionTitle + "\n" + spectralType +\
              "\n Average yaw {:.2f}° +/- {:.2f}°    | pitch {:.2f}°  +/-{:.2f}°"\
                  .format(avg[0], std[0], avg[1], std[1]))
    ax.set_title(myTitle, fontsize=8)
    ax.set_xlabel("%s yaw"% spectralType, fontsize=8)
    ax.set_ylabel("%s pitch"%spectralType, fontsize=8)
    ax.grid(color='gray', linestyle='-', linewidth=0.5)


def Fourier_plot(signal, title1='', title2='', color='p'):
    #signal_fin = signal[5]
    #signal = np.append(signal, signal_fin)
    signal = signal - np.average(signal)           # soustration de la moyenne  (elimination du terme constant)
    signal = signal / np.max(signal)               # normalisation du signal entre -1 et +1
    nombre_image = len(signal)  #
    print('nombre d\'images ', nombre_image)
    N = nombre_image // 2             # echantillonnage du signal pour la fft
    timeLapse_Img = 2                 # Timelapse de la caméra VIS du drone DJI.   en s
    t = np.linspace(0.0, timeLapse_Img * nombre_image, nombre_image, endpoint=False)  # axe du temps

    # Représentation graphique du signal
    fig, ax = plt.subplots()
    myTitle = 'Evolution du ' + title1 + ' en fonction du temps'
    ax.set_title(myTitle, fontsize=8)
    ax.set_xlabel('Temps  [s]', fontsize=8)
    myYlabel = title1 + '[°] '
    ax.set_ylabel(myYlabel, fontsize=8)
    ax.grid(color='gray', linestyle='-', linewidth=0.5)
    ax.plot(t, signal, color=color,  linewidth=0.2)
    plt.plot(t, signal, marker="o", markersize=3, alpha=0.6, mec='black', mfc=color, linestyle='')
    plt.show()

    # Calcul de la transformée de Fourier du signal
    Fourier = scipy.fft.fft(signal, norm='ortho')
    fondamentale = (1 / timeLapse_Img) * np.argmax(np.abs(Fourier[:])) / (nombre_image)
    print('%s  Féquence fondamentale f1 =  %.3f Hz    Période fondamentale  T1 = %.2f s'
          % (title1, fondamentale, 1 / fondamentale))
    x_f = np.linspace(0.0, 1 / 2 * 1 / timeLapse_Img, N, endpoint=False)        # axe des fréquences

    # Représentation graphique de la transformée (discrète) du signal
    fig, ax = plt.subplots()
    myTitle = 'Analyse spectrale du ' + title1
    ax.set_title(myTitle, fontsize=8)
    ax.set_xlabel('Fréquence  [Hz]', fontsize=8)
    myYlabel = 'FFT(' + title1 + ')  '
    ax.set_ylabel(myYlabel, fontsize=8)
    ax.grid(color='gray', linestyle='-', linewidth=0.5)
    plt.plot(x_f, 2 / N * np.abs(Fourier[: N]), color=color, linestyle='-',  linewidth=0.5)
    plt.fill_between(x_f, [0.]*len(x_f) ,2 / N * np.abs(Fourier[: N]), alpha=0.2, color='gray')
    plt.plot((1 / timeLapse_Img) * np.argmax(np.abs(Fourier[:])) / (nombre_image),
             timeLapse_Img / N * np.max(np.abs(Fourier[:])),
             marker="o", markersize=6, alpha=0.5, mec=color, mfc=color, linestyle='', linewidth=0.6)
    trans_offset = mtrans.offset_copy(ax.transData, fig=fig,
                                      x=0.08, y=-0.10, units='inches')
    plt.text((1 / timeLapse_Img) * np.argmax(np.abs(Fourier[:])) / (nombre_image),
             timeLapse_Img / N * np.max(np.abs(Fourier[:])),
             '%.3f Hz' % (fondamentale), transform=trans_offset)
    plt.show()



if __name__ == "__main__":
    versionIRdrone = '1.05'  # 02 december 2021
    # ----------------------------------------------------
    # 0 > Choix interactif de la mission
    #
    print(Style.CYAN + "File browser")
    dirMission = os.path.dirname(IRd.loadFileGUI(mute=True))
    print(Style.CYAN + dirMission + Style.RESET)

    print(Style.MAGENTA + 'Voulez vous utiliser les données en cache  ?' + Style.RESET)
    utilise_cache = IRd.answerYesNo('Oui (1) |  Non (0):')

    # ----------------------------------------------------
    # 1 > Trace les angles Roll, Pitch et Yaw  (roulis, tangage & lacet)
    #     pour le drone, le gimbal et l'image NIR
    #

    plotYawPitchRollDroneAndCameraDJI(dirMission, utilise_cache=utilise_cache)
