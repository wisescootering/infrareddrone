# -*- coding: utf-8 -*-

import PIL.Image
import PIL.ExifTags
import datetime
import time
from operator import itemgetter
import openpyxl
import urllib.request
import json
import numpy as np
import math
import irdrone.utils as ut
from irdrone.utils import Style
from irdrone.process import Image, show
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
import irdrone.process as pr
import irdrone.imagepipe as ipipe


# -------------     Convertisseurs      ---------------------------------

def dateExif2Py(exifTag):
    """
    :param exifTag: données Exif d'une image
    :return: datePhotoPython     (datetime python)

       Extraction de la date  de la prise de photo à partir de ses données Exif
       La date est fournie par  l'horloge de la caméra.
       Format Exif  YYYY:MM:DD  hh:mm:ss  type string
       Puis conversion en datetime python en analysant la chaine de caractère Exif
       On récupère Year Month Day Hour Min et Second  de type string
       Il faut convertir les string en integer (via la fonction int()  )

       Attention la clé Exif à utiliser est 'DateTimeOriginal'
       La clé Exif 'DateTime' indique la date de création du fichier sur l'ordinateur!
       La valeur de la clé 'DateTimeOriginal' est une chaine de caractères (string) 'YYYY:MM:DD  hh:mm:ss'
       Il faut extraire de cette chaine les valeurs numériques (integer) DD,MM,YYYY,hh,mm,ss
       puis construire une date "python" datetime.datetime(YYYY,MM,DD,hh,mm,ss,micross). Sous cette
       forme on peut manipuler les dates. En particulier on peut connaître  la durée entre deux
       dates  (en seconde). Il est aussi possible de connaître le nom du jour et du mois.

    """

    dateTimeOriginal = exifTag.get('DateTimeOriginal')  # date exacte de la prise de vue
    imgYear = int(dateTimeOriginal[0:4])
    imgMonth = int(dateTimeOriginal[5:7])
    imgDay = int(dateTimeOriginal[8:11])
    imgHour = int(dateTimeOriginal[11:13])
    imgMin = int(dateTimeOriginal[14:16])
    imgSecond = int(dateTimeOriginal[17:19])
    imgMicroSecond = 0
    # construction de la date python
    #  Rem: Pour accéder aux éléments individuels de la date python
    #       datePhotoPython.year(.month | .day | .hour | .minute | .second | .microseconde)
    #
    datePy = datetime.datetime(imgYear, imgMonth, imgDay, imgHour, imgMin, imgSecond, imgMicroSecond)

    return datePy


def datePy2Exif(datePy):
    """

    :param datePy:     (Year, Month, Day, Hour, Minute, Second, Microsecond, timezone)
    :return: dateExif     format  YYYY:MM:DD hh:mm:ss
    """
    dateExif = str(datePy.year) + ':' + str(datePy.month) + ':' + str(datePy.day) + ' ' + \
              str(datePy.hour) + ':' + str(datePy.minute) + ':' + str(datePy.second)
    return dateExif


def dateExif2Excel(dateExif):
    """
    :param dateExif:      format  YYYY:MM:DD hh:mm:ss
    :return: dateExcel    format  YYYY-MM-DD hh:mm:ss
    """
    date = dateExif.split(' ')[0].split(':')
    date.extend(dateExif.split(' ')[1].split(':'))
    dateExcel = str(date[0]) + '-' + str(date[1]) + '-' + str(date[2]) + ' ' + \
              str(date[3]) + ':' + str(date[4]) + ':' + str(date[5])

    return dateExcel


def geo2UTM(lat, lon):
    """
    :param lat: latitude  point P  dd.ddddddd   (<0 si S  >0 si N )
    :param lon: longitude point P  dd.ddddddd   (<0 si W  >0 si E )
    :return: xUTM, yUTM  UTM coordinates in m

    Conversion of geocentric coordinates to UTM coordinates.
    https://fr.wikipedia.org/wiki/Transverse_universelle_de_Mercator
    """
    a = 6378137.000  # equitorial radius in m
    f = 1. / 298.257223563
    e2 = 2. * f - f ** 2  # eccentricity
    # N° fuseau UTM
    fuseau = math.floor((lon + 180.) / 6.) + 1
    # convert Degrees to Radians
    lat = np.deg2rad(lat)
    lon = np.deg2rad(lon)
    lon_0 = np.deg2rad((fuseau - 30) * 6. - 3.)  # longitude of the center of the UTM zone
    #
    BB = 1. / (1 - e2 * (np.sin(lat)) ** 2) ** 0.5
    AA = (lon - lon_0) * np.cos(lat)
    S1 = (1. - e2 / 4. - 3. * (e2 ** 2) / 64. - 5. * (e2 ** 3) / 256.) * lat
    S2 = -(3. * e2 / 8. + 3. * (e2 ** 2) / 32. + 45. * (e2 ** 3) / 1024.) * np.sin(2. * lat)
    S3 = (15. * (e2 ** 2) / 256. + 45. * (e2 ** 3) / 1024.) * np.sin(4. * lat)
    S4 = -(35. * (e2 ** 3) / 3072.) * np.sin(6. * lat)
    S = S1 + S2 + S3 + S4
    T = (np.tan(lat)) ** 2
    C = (e2 / (1. - e2)) * (np.cos(lat)) ** 2
    K0 = 0.9996

    if lat > 0.:
        N0 = 0
    else:
        N0 = 10000000.

    xUTM = 500000. + K0 * a * BB * (AA + (1 - T + C) * (AA ** 3) / 6. + (5. - 18. * T + T ** 2) * (AA ** 5) / 120)
    yUTM = N0 + K0 * a * (S + BB * np.tan(lat) * (
                (AA ** 2) / 2. + (5 - T + 9 * C + 4 * C ** 2) * (AA ** 4) / 24. + (61. - 58. * T + T ** 2) * (
                    AA ** 6) / 720.))

    return xUTM, yUTM


def convertGPSExif_dms2GPSdd(repSign, coord):
    """
    :param repSign: {'E','W','S','N'}          str
    :param coord:   (dd,dd,ss.ssss)
    :return: coordDD    dd.dddddddd           float

         Conversion from dd° mm' ss.sssss"   en dd.dddddddd
    """
    if repSign == "W" or repSign == "S":
        signe = -1.
    else:
        signe = 1.
    coordDD = signe * (coord[0] + (coord[1] / 60) + (coord[2] / 3600))
    return coordDD

# -------------------     GPS     ----------------------------------------

def getGPS_Exif(imgName, mute=True):
    """
        :param imgName:
        :param mute:
        :return: gpsLatitude, gpsLongitude, gpsAltitude

       Extraction des coordonnées GPS des données Exif de l'image img
       La clé est GPSInfo  qui est une liste d'item
       pour la latitude   item(1) in {'N','S'} , item(2) =[(dd,1),(mm,1),(ssssss,10000)]
       pour la longitude item(3) in {'W','E'} , item(4) =[(dd,1),(mm,1),(ssssss,10000)]
       Retourne longitude [hémisphère, dd°, mm', ss.ssss", dd.ddddd]
              & latitude [méridien, dd°, mm', ss.ssss", dd.ddddd]
              & altitude
       Rem :   hémisphère in {'N','S'}    if hémisphère == Sud  then dd <0
               méridien in {'E','W'}      if méridien == West   then dd <0
    """
    img = PIL.Image.open(imgName)
    exifTag = {
        PIL.ExifTags.TAGS[k]: v
        for k, v in img._getexif().items()
        if k in PIL.ExifTags.TAGS
    }
    # Recherche de la latitude du lieu
    gpsHemisphere = exifTag.get('GPSInfo').__getitem__(1)
    if gpsHemisphere == 'S':
        signeLat = -1.
    else:
        signeLat = 1.
    gpsLat = [chiffre / (1. * precision) for chiffre, precision in exifTag.get('GPSInfo').__getitem__(2)]

    gpsLatDD = convertGPSExif_dms2GPSdd(gpsHemisphere, gpsLat)
    gpsLatitude = [gpsHemisphere, signeLat * gpsLat[0], gpsLat[1], gpsLat[2], gpsLatDD]

    # Recherche de la longitude du lieu
    gpsMeridien = exifTag.get('GPSInfo').__getitem__(3)
    if gpsMeridien == 'W':
        signeLong = -1.
    else:
        signeLong = 1.
    gpsLong = [chiffre / (1. * precision) for chiffre, precision in exifTag.get('GPSInfo').__getitem__(4)]
    gpsLongDD = convertGPSExif_dms2GPSdd(gpsMeridien, gpsLong)
    gpsLongitude = [gpsMeridien, signeLong * gpsLong[0], gpsLong[1], gpsLong[2], gpsLongDD]

    # Recherche de l'altitude du lieu
    #   Attention le DJI indique l'altitude par rapport au point de décollage.
    gpsAltitude = exifTag.get('GPSInfo').__getitem__(6)[0] / (1. * exifTag.get('GPSInfo').__getitem__(6)[1])

    if not mute:  printGPS(gpsLatitude, gpsLongitude, gpsAltitude)

    return gpsLatitude, gpsLongitude, gpsAltitude


def formatCoordGPSforGpx(listImgMatch):
    """

    :param listImgMatch: list of images pairs
    :return: pointTrk  list of GPS coordinates (str)    (format gpx Garmin)
             sample      '<trkpt 45.05022 3.89567 >
                          <ele>110.5</ele>
                          </trkpt>'
            maxLat, minLat, maxLon, minLon
    """
    coordGPSgpxLat = []
    coordGPSgpxLon = []
    coordGPSgpxAlt = []
    pointTrk = []
    for k in range(len(listImgMatch)):
        gpsLatitude, gpsLongitude, gpsAltitude = getGPS_Exif(listImgMatch[k][0], mute=True)
        lat = convertGPSExif_dms2GPSdd(gpsLatitude[0], [gpsLatitude[1], gpsLatitude[2], gpsLatitude[3]])
        lon = convertGPSExif_dms2GPSdd(gpsLongitude[0], [gpsLongitude[1], gpsLongitude[2], gpsLongitude[3]])
        alt = gpsAltitude
        coordGPSgpxLat.append(lat)
        coordGPSgpxLon.append(lon)
        coordGPSgpxAlt.append(alt)
        pointTrk.append("\n<trkpt lat=\"%s\" lon=\"%s\">\n<ele>%s</ele>\n</trkpt>" % (lat, lon, alt))

    maxLat = max(coordGPSgpxLat)
    minLat = min(coordGPSgpxLat)
    maxLon = max(coordGPSgpxLon)
    minLon = min(coordGPSgpxLon)

    return pointTrk, maxLat, minLat, maxLon, minLon


def writeGPX(listImgMatch, dirNameVol, dateEtude, mute=True):
    """
    :param listImgMatch:
    :param dirNameVol:
    :param dateEtude:
    :param mute:
    :return:
        Construction d'un fichier gpx contenant le tracé du plan de vol
        Il y a au début une tres grosse étiquette !!
    """

    #  mise en forme de la date pour le format gpx Garmin
    if dateEtude.month < 10:
        monthGpx = str('0' + str(dateEtude.month))
    else:
        monthGpx = str(dateEtude.month)
    if dateEtude.day < 10:
        dayGpx = str('0' + str(dateEtude.day))
    else:
        dayGpx = str(dateEtude.day)

    dateGpx = '%i-%s-%sT00:00:00Z' % (dateEtude.year, monthGpx, dayGpx)

    #     mise en forme des coordonnées GPS pour le format gpx
    #     Il faut aussi calculer la zone GPS  définie par le domaine [maxLat,minLat]x[maxLon,minLon]
    pointTrk, maxLat, minLat, maxLon, minLon = formatCoordGPSforGpx(listImgMatch)

    # affectation du nom du fichier  et d'une description
    nameTrkGPS = "IRdrone-%s-%s-%i" % (dayGpx, monthGpx, dateEtude.year)
    descriptionTrkGPS = str('IRdrone v01.1   Trace GPS du vol %s-%s-%i' % (dayGpx, monthGpx, dateEtude.year))

    fichierGpx = "<?xml version=\"1.0\" encoding=\"utf-8\"?>" \
                 "<gpx creator=\"IRdrone v01.1\" " \
                 "version=\"1.1\" " \
                 "xsi:schemaLocation=\"http://www.topografix.com/GPX/1/1 \" "

    fichierGpx = "{0}xmlns=\"http://www.topografix.com/GPX/1/1\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:wptx1=\"http://www.garmin.com/xmlschemas/WaypointExtension/v1\" xmlns:gpxtrx=\"http://www.garmin.com/xmlschemas/GpxExtensions/v3\" xmlns:gpxtpx=\"http://www.garmin.com/xmlschemas/TrackPointExtension/v1\" xmlns:gpxx=\"http://www.garmin.com/xmlschemas/GpxExtensions/v3\" xmlns:trp=\"http://www.garmin.com/xmlschemas/TripExtensions/v1\" xmlns:adv=\"http://www.garmin.com/xmlschemas/AdventuresExtensions/v1\" xmlns:prs=\"http://www.garmin.com/xmlschemas/PressureExtension/v1\" xmlns:tmd=\"http://www.garmin.com/xmlschemas/TripMetaDataExtensions/v1\" xmlns:vptm=\"http://www.garmin.com/xmlschemas/ViaPointTransportationModeExtensions/v1\" xmlns:ctx=\"http://www.garmin.com/xmlschemas/CreationTimeExtension/v1\" xmlns:gpxacc=\"http://www.garmin.com/xmlschemas/AccelerationExtension/v1\" xmlns:gpxpx=\"http://www.garmin.com/xmlschemas/PowerExtension/v1\" xmlns:vidx1=\"http://www.garmin.com/xmlschemas/VideoExtension/v1\">".format(
        fichierGpx)

    """
        Couleur du trait :   Red,Green,Blue,Yellow,Gray, et DarkRed, DarkGreen etc

        Attention format des dates   2021-02-11T19:00:00Z
    """

    fichierGpx = fichierGpx + "\n<metadata>\n" \
                              "<link href=\"http://www.garmin.com\">\n" \
                              "<text>Garmin International</text>\n" \
                              "</link>\n" \
                              "<time>" + dateGpx + "</time>\n" \
                                                   "<bounds maxlat=\"" + str(maxLat) + "\" maxlon=\"" + str(
        minLat) + "\" minlat=\"" + str(maxLon) + "\" minlon=\"" + str(minLon) + "\" />\n" \
                                                                                "</metadata>\n" \
                                                                                "<trk>\n" \
                                                                                "<name>" + nameTrkGPS + "</name>\n" \
                                                                                                        "<desc>" + descriptionTrkGPS + "</desc>\n" \
                                                                                                                                       "<extensions>\n" \
                                                                                                                                       "<gpxx:TrackExtension>\n" \
                                                                                                                                       "<gpxx:DisplayColor>Red</gpxx:DisplayColor>\n" \
                                                                                                                                       "</gpxx:TrackExtension>\n" \
                                                                                                                                       "</extensions>\n" \
                                                                                                                                       "<trkseg>"

    for k in range(len(listImgMatch)):
        fichierGpx = fichierGpx + pointTrk[k]

    fichierGpx = fichierGpx + "\n</trkseg>\n" \
                              "</trk>\n" \
                              "</gpx>"

    if not mute: print(fichierGpx)

    dirpath = '%s\\TrkGpx-%s-%s-%i.gpx' % (dirNameVol, dayGpx, monthGpx, dateEtude.year)

    if not mute: print('Ecriture du fichier gpx %s' % dirpath)

    with open(dirpath, "w") as fichier:
        fichier.write(fichierGpx)

    return


def getAltitudeEncode(listPtdGPS=None):
    """
        :param listPtdGPS: the table of coordinates to be processed.
        :return: Altitude (or elevation) of GPS points relative to sea level.    array

        Script for returning elevation from lat, long, based on open elevation data
        which in turn is based on SRTM  (NASA)

        CAUTION
        The coverage area of the Elevation API is limited to latitudes between 56°S and 60°N.
    """
    if listPtdGPS==None: return None

    tries=3   # nombre de tentatives de connexion au serveur
    dimListPtdGPS = len(listPtdGPS)
    elev_list = [0.] * dimListPtdGPS

    # CONSTRUCT JSON
    # Transform the table of coordinates into a encoded dictionary.
    d_ar = [{}] * dimListPtdGPS
    for i in range(dimListPtdGPS ):
        d_ar[i] = {"latitude": listPtdGPS[i][0], "longitude": listPtdGPS[i][1]}
    location = {"locations": d_ar}
    json_data = json.dumps(location, skipkeys=int).encode('utf8')

    # SEND REQUEST
    for i in range(tries):
            try:
                url = "https://api.open-elevation.com/api/v1/lookup"
                response = urllib.request.Request(url, json_data, headers={'Content-Type': 'application/json'})
                fp = urllib.request.urlopen(response)

                # RESPONSE PROCESSING
                res_byte = fp.read()
                res_str = res_byte.decode("utf8")
                js_str = json.loads(res_str)
                # print (js_mystr)
                fp.close()

                # GETTING ELEVATION
                response_len = len(js_str['results'])
                elev_list = []
                for j in range(response_len):
                    elev_list.append(js_str['results'][j]['elevation'])

                print(Style.BLUE+ 'https://api.open-elevation.com/api/v1/lookup'+Style.GREEN+ ' : OK'+Style.RESET)
                break

            except:
                if i< tries-1:
                    continue
                else:
                    print(Style.RED + 'No response from web server https://api.open-elevation.com/api/v1/lookup', '\n '
                        'Ground level set to zero.' + Style.RESET)

    return elev_list


def segmentUTM(lat1, lon1, lat2, lon2):
    """
           :param lat1: latitude  point P1  dd.ddddddd   (<0 si S  >0 si N )
           :param lon1: longitude point P1  dd.ddddddd   (<0 si W  >0 si E )
           :param lat2: latitude  point P2  dd.ddddddd   (<0 si S  >0 si N )
           :param lon2: longitude point P2  dd.ddddddd   (<0 si W  >0 si E )
           :return:  distUTM  Euclidean distance between P1 and P2  in m
                     capUTM   heading between P1 and P2 from the geographic north

           Calculation of the distance between two points P1 and P2 on the earth surface.
           We assume that the earth is an ellipsoid  (WGS84).
           A compliant transverse Mercator projection is used. (UTM means Universal Transverse Mercator projection)
           Points are identified by their geocentric coordinates latitude,longitude (input).
           Geocentric coordinates are converted to UTM coordinates x,y.
           The transverse projection of Mercator is a conformal map.
           Conformal map is a function that locally preserves angles, but not necessarily lengths.
           In other words, it's therefore easy to determine the direction between two points
            relative to the geographic north.
           In the (x,y) plan we can use the Euclidean distance.
           The approximation is excellent in each UTM zone.
       """
    x1, y1 = geo2UTM(lat1, lon1)
    x2, y2 = geo2UTM(lat2, lon2)
    distUtm = ((x1 - x2) ** 2 + (y1 - y2) ** 2) ** 0.5
    capUtm = capUTM(x1, y1, x2, y2)
    return distUtm, capUtm


def capUTM(x1, y1, x2, y2):
    """
    :param x1: Point P1  UTM coordinate W-E
    :param y1: Point P1  UTM coordinate S-N
    :param x2: Point P2  UTM coordinate W-E
    :param y2: Point P2  UTM coordinate S-N
    :return:   cap  Direction from the geographical north between P1 and P2.
    """
    cap = 0.
    if y1 == y2:
        if x1 > x2:
            cap = 270.
        else:
            cap = 90.
    else:
        if y2 > y1 and x2 >= x1:
            cap = np.rad2deg(np.arctan(np.abs((x2 - x1) / (y2 - y1))))
        elif y2 < y1 and x2 > x1:
            cap = 180 - np.rad2deg(np.arctan(np.abs((x2 - x1) / (y2 - y1))))
        elif y2 > y1 and x2 < x1:
            cap = 360. - np.rad2deg(np.arctan(np.abs((x2 - x1) / (y2 - y1))))
        elif y2 < y1 and x2 <= x1:
            cap = 180. + np.rad2deg(np.arctan(np.abs((x2 - x1) / (y2 - y1))))
        else:
            pass
    return cap

# ----------------------     Exif   ------------------------------------

def extractExif(imgName):
    """
    :param imgName:   image path (str)
    :return: exifTag
    """
    try:
        img = PIL.Image.open(imgName)
        exifTag = {
            PIL.ExifTags.TAGS[k]: v
            for k, v in img._getexif().items()
            if k in PIL.ExifTags.TAGS
        }
    except:
        print(Style.RED + "No Exif tags in %s" % imgName + Style.RESET)
        exifTag = None

    return exifTag

# ------------------     Plan de Vol      -------------------------------


def readFlightPlan(pathPlanVolExcel):
    """
    :param pathPlanVolExcel: chemin du fichier Excel qui contient le plan de vol  (string)
    :return: planVol  Dictionnaire contenant les données du plan de vol  (dic)

        Read the Flight Plan  in Excel file
        Lecture du fichier Excel qui contient les informations sur le plan de vol

    """
    workbook = openpyxl.load_workbook(pathPlanVolExcel, read_only=True)
    sheet = workbook['Plan_de_Vol']
    nuetude = 2  # 2
    nudrone = (nuetude + 8) + 1  # 2+8+1 =11
    nucameraIR = (nudrone + 7) + 1  # 11+7+1=19
    nuimages = (nucameraIR + 5) + 1  # 19+5+1=25
    numeteo = (nuimages + 11) + 1  # 25+11+1=37
    planVol = {sheet.cell(nuetude, 1).value:
                   {sheet.cell(nuetude + 1, 1).value: sheet.cell(nuetude + 1, 2).value,  # 'client'
                    sheet.cell(nuetude + 2, 1).value: sheet.cell(nuetude + 2, 2).value,  # 'lieu'
                    sheet.cell(nuetude + 3, 1).value: sheet.cell(nuetude + 3, 2).value,
                    # 'GPS'    N dd.ddddd E dd.ddddd
                    sheet.cell(nuetude + 4, 1).value: sheet.cell(nuetude + 4, 2).value,  # 'altitude'
                    sheet.cell(nuetude + 5, 1).value: sheet.cell(nuetude + 5, 2).value,  # 'date'   DD/MM/YYYY  hh:mm:ss
                    sheet.cell(nuetude + 6, 1).value: sheet.cell(nuetude + 6, 2).value,  # 'heure_solaire'
                    sheet.cell(nuetude + 7, 1).value: sheet.cell(nuetude + 7, 2).value,  # 'numero_du_vol'
                    sheet.cell(nuetude + 8, 1).value: sheet.cell(nuetude + 8, 2).value  # 'pilote'
                    },
               sheet.cell(nudrone, 1).value:  # 'drone'
                   {sheet.cell(nudrone + 1, 1).value: sheet.cell(nudrone + 1, 2).value,  # 'marque'
                    sheet.cell(nudrone + 2, 1).value: sheet.cell(nudrone + 2, 2).value,  # 'type'
                    sheet.cell(nudrone + 3, 1).value: sheet.cell(nudrone + 3, 2).value,  # 'timelapse'(Drone)
                    sheet.cell(nudrone + 4, 1).value: sheet.cell(nudrone + 4, 2).value,  # 'deltatime'
                    sheet.cell(nudrone + 5, 1).value: sheet.cell(nudrone + 5, 2).value,  # 'imatriculation'
                    sheet.cell(nudrone + 6, 1).value: sheet.cell(nudrone + 6, 2).value,  # 'altitude  de vol'
                    sheet.cell(nudrone + 7, 1).value: sheet.cell(nudrone + 7, 2).value  # 'altitude  de vol'
                    },
               sheet.cell(nucameraIR, 1).value:  # 'cameraIR'
                   {sheet.cell(nucameraIR + 1, 1).value: sheet.cell(nucameraIR + 1, 2).value,  # 'marque'
                    sheet.cell(nucameraIR + 2, 1).value: sheet.cell(nucameraIR + 2, 2).value,  # 'type'
                    sheet.cell(nucameraIR + 3, 1).value: sheet.cell(nucameraIR + 3, 2).value,  # 'timelapse'(IR)
                    sheet.cell(nucameraIR + 4, 1).value: sheet.cell(nucameraIR + 4, 2).value,  # 'deltatime'
                    sheet.cell(nucameraIR + 5, 1).value: sheet.cell(nucameraIR + 5, 2).value  # libre
                    },
               sheet.cell(nuimages, 1).value:  # 'images'
                   {sheet.cell(nuimages + 1, 1).value: sheet.cell(nuimages + 1, 2).value,
                    # 'repertoire' répertoire général des images du vol
                    sheet.cell(nuimages + 2, 1).value: sheet.cell(nuimages + 2, 2).value,
                    # 'repertoireDrone'  répertoire des images drone
                    sheet.cell(nuimages + 3, 1).value: sheet.cell(nuimages + 3, 2).value,
                    # 'extDrone' extention des images (JPG, DNG)
                    sheet.cell(nuimages + 4, 1).value: sheet.cell(nuimages + 4, 2).value,
                    # 'filtreDrone'   densité du filtre DNxxx
                    sheet.cell(nuimages + 5, 1).value: sheet.cell(nuimages + 5, 2).value,  # libre
                    sheet.cell(nuimages + 6, 1).value: sheet.cell(nuimages + 6, 2).value,  # libre
                    sheet.cell(nuimages + 7, 1).value: sheet.cell(nuimages + 7, 2).value,
                    # 'repertoireIR'  repertoire des images IR
                    sheet.cell(nuimages + 8, 1).value: sheet.cell(nuimages + 8, 2).value,
                    # 'extIR'       extention des images infrarouge
                    sheet.cell(nuimages + 9, 1).value: sheet.cell(nuimages + 9, 2).value,
                    # 'filtreIR'     longueur d'onde du filtre IR
                    sheet.cell(nuimages + 10, 1).value: sheet.cell(nuimages + 10, 2).value,  # libre
                    sheet.cell(nuimages + 11, 1).value: sheet.cell(nuimages + 11, 2).value  # libre
                    },
               sheet.cell(numeteo, 1).value:  # 'meteo'
                   {sheet.cell(numeteo + 1, 1).value: sheet.cell(numeteo + 1, 2).value,  # 'ensoleillement'
                    sheet.cell(numeteo + 2, 1).value: sheet.cell(numeteo + 2, 2).value,  # 'vent'
                    sheet.cell(numeteo + 3, 1).value: sheet.cell(numeteo + 3, 2).value  # 'temperature'
                    }
               }
    workbook.close()

    return planVol


def extractFlightPlan(dirPlanVol, mute=True):
    """
        Lecture du plan de vol puis extrcation des données
        > Chemin des images Visible et IR, type de drone et de caméra, synchroniation horloges ...
        > Liste des images Drone et des images IR
    """
    planVol = readFlightPlan(dirPlanVol)

    dirNameIRdrone = planVol['images'][
        'repertoireViR  (save)']  # folder for save photography  VIR, RedEdge,NIR, NDVI etc output
    dirNameDrone = planVol['images']['repertoireDrone']  # Drone photography folder   (input)
    dirNameIR = planVol['images']['repertoireIR']  # IR photography folder (input)
    dateEtude = planVol['etude']['date']  # date of flight > format DD MM et YYYY
    typeDrone = planVol['drone']['type']  # type of drone (see in the Exif tag of the image of the drone)
    extDrone = planVol['images']['extDrone']  # file format Vi
    typeIR = planVol['cameraIR']['type']  # type of camera in use (see in the Exif tag of the image of the IR camera)
    timeLapseDrone = float(planVol['drone']['timelapse'])  # Time Lapse of Drone camera
    timeLapseIR = float(planVol['cameraIR']['timelapse'])  # Time Lapse of IR camera
    extIR = planVol['images']['extIR']  # file format  IR
    deltaTimeDrone = planVol['drone']['deltatime']  # decalage horloge caméra du drone / horloge de référence
    deltaTimeIR = planVol['cameraIR']['deltatime']  # decalage horloge caméra infrarouge /horloge de référence

    #
    #    Liste des images de l'étude.
    #    Une liste pour les images du drone et une liste pour les images de la caméra infrarouge
    #    Chaque élément de la liste est un triplet (file name image , path name image, date image)

    imgListDrone = creatListImg(dirNameDrone, dateEtude, typeDrone, '*', extDrone)
    imgListIR = creatListImg(dirNameIR, dateEtude, typeIR, '*', extIR)

    if not mute:
        print('liste des fichiers Drone:')
        print([imgListDrone[i][0] for i in range(len(imgListDrone))])
        print('liste des fichiers IR:')
        print([imgListIR[i][0] for i in range(len(imgListIR))])

    return imgListDrone, deltaTimeDrone, timeLapseDrone, imgListIR, deltaTimeIR, timeLapseIR, dateEtude, dirNameIRdrone


def modelCamera(exifTag, mute=True):
    cameraMaker = exifTag.get('Make')  # Fabricant de la caméra ('MAKER NAME' pour SJcam! et 'DJI' pour le drone Dji
    cameraModel = exifTag.get('Model')  # Modèle de la caméra
    cameraSensor = exifTag.get('BodySerialNumber')  # Modèle de la caméra
    if not mute: print(' Fabricant : %s \n Drone : %s \n Capteur : %s' % (cameraMaker, cameraModel, cameraSensor))
    return cameraModel, cameraMaker, cameraSensor


def creatListImg(dirName, dateEtude, cameraModel, camera, typImg, debug=False):
    """

    :param dirName:
    :param dateEtude:
    :param cameraModel:
    :param camera:
    :param typImg:
    :param debug:
    :return:  imgList   [(file name image , path name image, date image),... ]
    """

    # Création de la liste des photos (extension donnée  ex: JPG, DNG) et contenues dans le répertoire dirName
    # Si camera="*' alors la liste comporte toutes les images avec l'extension typeImg
    # Si camera ='DJI'  la liste filtre les images prises par le drone
    # Pas possible de filtrer les images SJcam avec le non du fichier  (il faut regarder les données Exif de l'image)
    imlist = sorted(ut.imagepath(imgname="*%s*.%s" % (camera, typImg), dirname=dirName))

    imgList = []
    j = 0
    for i in range(len(imlist)):
        try:
            """
            extraction des données Exif de l'image
            si pas de données Exif image ignorée.
            """
            exifTag = extractExif(imlist[i])  # extraction des données Exif de l'image
            cameraModelImg, makerCameraImg, sensorlImg = modelCamera(exifTag)
            if cameraModelImg == cameraModel:  # images prises par d'autres caméras. images ignorées
                dateImg = dateExif2Py(exifTag)  # extraction et conversion de la date de prise de vue
                if (dateImg.year, dateImg.month, dateImg.day) == (dateEtude.year, dateEtude.month, dateEtude.day):
                    j += 1
                    nameImg = imlist[i].split('\\')[len(imlist[i].split('\\')) - 1]
                    imgList.append((nameImg, imlist[i], dateImg))  # ajout à la liste des images
                else:
                    if debug: print('%s a été prise le %i %i %i. Cette date est différente de celle de l\'étude %i %i %i'
                        % (imlist[i], dateImg.day, dateImg.month, dateImg.year, dateEtude.day, dateEtude.month,
                           dateEtude.year))
            else:
                if debug: print('%s a été prise par un autre  appareil (Model %s) ' % (imlist[i], cameraModelImg))
        except:
            if debug: print("No Exif tags in %s" % imlist[i])

    imgList = sorted(imgList, key=itemgetter(2), reverse=False)  # tri par la date de prise de vue

    return imgList


def matchImagesFlightPath(imgListDrone, deltaTimeDrone, timeLapseDrone, imgListIR, deltaTimeIR, timeLapseIR, mute=False):
    """
    :param imgListDrone:  [...,(file name, path name, date), ...]
    :param deltaTimeDrone:
    :param timeLapseDrone
    :param imgListIR:     [...,(file name, path name, date), ...]
    :param deltaTimeIR:
    :param timeLapseIR:
    :param mute:
    :return:  listImgMatch   [..., (imgListDrone[i][1], imgListIR[k][1]), ...]
    """
    n = 0
    nRejet = 0
    listImgMatch = []
    for i in range(len(imgListDrone)):
        indexDT = []
        DT = []
        for k in range(len(imgListIR)):
            dateDrone = imgListDrone[i][2]
            dateIR = imgListIR[k][2]
            deltaTime = datetime.timedelta.total_seconds(dateIR - dateDrone) - deltaTimeIR + deltaTimeDrone
            if abs(deltaTime) < timeLapseIR:
                #  Potentiellement cette image IR  peut s'apparier avec l'image visible
                indexDT.append(k)
                DT.append(abs(deltaTime))
            else:
                pass
        if len(DT) == 0:
            # Aucune image IR ne s'apparie avec l'image visible
            nRejet += 1
            if not mute: print(Style.YELLOW + 'INFO:  l\'image visible ', imgListDrone[i][0],
                               'ne s\'apparie avec aucune image IR.' + Style.RESET)
            pass
        else:
            #    Une ou plusieurs iamges IR peuvent s'apparier avec l'image visible.
            #    On choisi celle qui est la plus proche et on l'ajoute à la liste des images appariées
            kIRmatch = indexDT[DT.index(min(DT))]
            # On test si cette image n'a pas déjà été utilisée (lié à un pb de précision des horloges!)
            if n > 1 and listImgMatch[n - 1][1] == imgListIR[kIRmatch][0]:
                if not mute:
                    print(Style.RED + 'Attention pb avec les deux images \n'
                                      'Visible :', imgListDrone[i][0],
                          '\nInfrarouge :', imgListIR[kIRmatch][0], '\n'
                          'L\'image infrarouge  est déjà appariée avec l\'image ',
                          listImgMatch[n - 1][1] + Style.RESET)
                nRejet += 1
            else:
                n += 1
                listImgMatch.append((imgListDrone[i][1], imgListIR[kIRmatch][1]))
                if not mute: print(Style.GREEN + 'N°', n, ' DT ', min(DT), '  ', imgListDrone[i][0], ' | ',
                                   imgListIR[kIRmatch][0] + Style.RESET)

    print(Style.GREEN + '%i couples d\'images Visible-InfraRouge ont été détectés pour le vol du %s' % (
    len(listImgMatch), dateEtude), '\n',
          ' Images visibles éliminées : %i' % nRejet + Style.RESET)
    return listImgMatch


def summaryFlight(listImg, seaLevel=False, mute=True):
    """
    :param listImg:      list of image pairs VI/IR matched at the same point
    :param mute:
    :return: summary  [... (imgNameVi,imgNameIRoriginal,imgNameIR,imgNameViR, dateVi,
                          imgLat, imgLon, imgAltGround , imgAltSeaLevel ,imgAltRef,
                          distToLastPt,capToLastPt   ) ...]                   list
        imgNameVi---------------------- file name of visible RGB color image at point P
        imgNameIRoriginal-------------- file name of original infrared image  at point P
        imgNameIR---------------------- file name of infrared image  (match with visible image at point P)
        imgNameViR -------------------- file name of multispectral image at point P
        dateVi ------------------------ date the visible image was shot             [YY-MM-DD hh:mm:ss]
        imgLat, imgLon, imgAltGround    GPS coordinates of point P                  [°],[°],[m]
        imgAltSeaLevel ---------------  altitude of the drone in relation to the sealevel [m]
        imgAltRef---------------------- altitude of the drone relative to take-off point  (DJI ref)  [m]
        distToLastPt------------------- distance to next point                                       [m]
        capToLastPt-------------------- direction to the next point relative to the geographic north [°]

        lists the essential elements of flight after treatment by IRdrone.
        This data will be saved in the Excel file of the flight plan  (sheet "Summary"

    """
    summary = []
    nPt = len(listImg)

    listPtGPS = []
    distP0P1 = []
    capP0P1 = []
    for i in range(nPt):
        lat, lon, altRef = getGPS_Exif(listImg[i][0], mute=mute)
        listPtGPS.append((round(lat[4], 6), round(lon[4], 6),round(altRef, 3)))

    if seaLevel :
        #  Pour tout les points de la liste on détermine les altitudes du sol par rapport au niveau de la mer.
        #  ATTENTION :  utilise une API qui peut échouer car réponse du serveur lente ?
        #  On fait trois tentatives. Si échec l'altitude du sol par rapport au niveau de la mer est fixée à zéro.

        altGeo=getAltitudeEncode(listPtGPS)  # liste des altitudes du sol par raport au niveau de la mer
        altGeoPref = altGeo[0]  # altitude du point de décollage par rapport au niveau de la mer
    else:
        altGeo = [0.0]*len(listPtGPS)
        altGeoPref = 0.

    #  Calcule la distance et le cap entre la point P0 et le point suivant P1
    for i in range(nPt - 1):
        imgDist, imgCap = segmentUTM(listPtGPS[i][0], listPtGPS[i][1], listPtGPS[i + 1][0], listPtGPS[i + 1][1])
        distP0P1.append(imgDist)
        capP0P1.append(imgCap)
    distP0P1.append(0)  # pas de distance ni de cap au point suivant pour point d'atterrissage !
    capP0P1.append(0)  # on fixe la valeur arbitrairement à 0


    altDrone=[]
    altGround=[]
    distFlight=[]
    for i in range(nPt):
        exifVi = extractExif(listImg[i][1])
        imgNameVi = listImg[i][0].split('\\')[len(listImg[i][0].split('\\')) - 1]
        numeroRef = imgNameVi.split('_')[len(imgNameVi.split('_')) - 1]
        dateVi = dateExif2Excel(exifVi.get('DateTimeOriginal'))
        altDrone.append(listPtGPS[i][2] + altGeoPref)   # altitude du drone par rapport au niveau de la mer
        altGround.append(altGeo[i])                     # altitude du sol par rapport au niveau de la mer
        distToLastPt = round(distP0P1[i], 2)            # distance au point suivant   (précision au cm !)
        if i==0:
            distFlight.append(0.)
        else:
            distFlight.append(distFlight[i - 1] + distToLastPt)

        capToLastPt = round(capP0P1[i], 3)
        #    images infra-rouges
        exifIR = extractExif(listImg[i][1])
        imgNameIRoriginal = listImg[i][1].split('\\')[len(listImg[i][1].split('\\')) - 1]
        imgNameIR = 'IRdroneIR_%s' % numeroRef
        imgNameViR = 'IRdroneViR_%s' % numeroRef
        summary.append(
            (imgNameVi, imgNameIRoriginal, imgNameIR, imgNameViR,
             dateVi,
             round(listPtGPS[i][0],5), round(listPtGPS[i][1],5),
             round(altDrone[i],3), round(altGeo[i], 3),round(listPtGPS[i][2],3),
             distToLastPt, capToLastPt,round(distFlight[i],2)
             )
        )

    #showFlightProfil(distFlight,altDrone,altGround, mute=False)


    if not mute:
        txtSummary = 'Liste des images du vol :'
        for i in range(nPt):
            txtSummary = txtSummary + '\n' + str(summary[i])
        print(Style.GREEN + txtSummary + Style.RESET)


    return summary


def writeSummaryFlight(flightPlanSynthesis, pathName):
    """
    :param flightPlanSynthesis:
    :param pathName:
    :return:

            Write the Flight Plan Summary in Excel file
    ToDo créer la feuille Summary si elle n'existe pas. Normalement à ce stade le fichier Excel plan de vol existe
        puisqu'il a été lu dès le début
    """
    workbook = openpyxl.load_workbook(pathName)
    listSheets = workbook.sheetnames
    if len(listSheets) < 2:
        print(Style.RED + ' le fichier Excel du plan de vol ne contient pas de feuille Summary' + Style.RESET)
        return
    elif listSheets[1] != 'Summary':
        print(Style.RED + ' le fichier Excel du plan de vol ne contient pas de feuille Summary' + Style.RESET)
        return
    else:
        pass

    sheet = workbook['Summary']
    for i in range(len(flightPlanSynthesis)):
        sheet.cell(i + 2, 1).value = i
        for k in range(len(flightPlanSynthesis[i])):
            sheet.cell(i + 2, k + 2).value = flightPlanSynthesis[i][k]
    sheet = workbook['Plan_de_Vol']
    sheet.cell(2, 2).value = None
    workbook.save(pathName)
    workbook.close()
    print(Style.GREEN + 'Ecriture du résumé du vol dans %s  terminée.' % pathName)


# ---------     Visualisation  Graphiques & Images  --------------------------

def colorMapIRlayer(IRband):
    #   myColorMap    >  'autumn', 'bone', 'cool', 'copper', 'flag', 'gray', 'hot'
    #                 'hsv', 'jet', 'pink', 'prism','spring', 'summer', 'winter'

    if IRband == 1:
        myColorMap = 'hot_r'
        myMinColorBar = 0.
        myMaxColorBar = 1.
    elif IRband == 2:
        myColorMap = 'copper'
        myMinColorBar = 0.
        myMaxColorBar = 1.
    elif IRband == 3:
        myColorMap = 'bone'
        myMinColorBar = 0.
        myMaxColorBar = 1.
    elif IRband == 4:
        myColorMap = 'gray'
        myMinColorBar = 0.
        myMaxColorBar = 1.
    elif IRband == 5:
        myColorMap = colorMapNDVI()
        myMinColorBar = -1
        myMaxColorBar = 1

    elif IRband == 6:
        myColorMap = colorMapNDVI()
        myMinColorBar = -1
        myMaxColorBar = 1
    else:
        myColorMap = 'gray'
        myMinColorBar = 0
        myMaxColorBar = 1

    return myColorMap, myMinColorBar, myMaxColorBar


def colorMapNDVI():
    #  définition de la palette des couleurs pour l'indice NDVI à partir de couleurs prédéfinies
    #  voir les couleurs par exemple ici   http://xymaths.free.fr/Informatique-Programmation/Couleurs/Liste.php
    colors = ["black",
              "dimgray",
              "lightgray",
              "burlywood",
              "lawngreen",
              "lightseagreen",
              "forestgreen",
              "lightgray"
              ]
    #   répartition des plages de couleurs  (entre 0 et 1)
    nodes = [0.0,
             35. / 100,
             45. / 100,
             51. / 100,
             60. / 100,
             65. / 100,
             75. / 100,
             1.0
             ]
    myColorMap = LinearSegmentedColormap.from_list("mapNDVI", list(zip(nodes, colors)))

    # Autre possibilité: egale répartition entre les couleurs
    # myColorMap = LinearSegmentedColormap.from_list("mapNDVI", colors)  #

    return myColorMap


def showDualImages(listImgMatch, modulo=1, seeDualImages=False):
    if seeDualImages:
        for numImg in range(len(listImgMatch) // modulo):
            listPairesImg = [
                [Image(listImgMatch[modulo * numImg + j][0],
                       name="Wpt N°%i" % j).data,
                 Image(listImgMatch[modulo * numImg + j][1],
                       name="Wpt N°%i" % j).data]
                for j in range(modulo)
            ]

            show(listPairesImg, title=None, compare=True, block=True, suptitle=None,
                 figsize=(15, 10))  # figsize=(15,13)


def flightProfil(listMatch, seaLevel=False, dirSaveFig=r'C:\Air-Mission\Drone-Flight-profil.png',mute=True):
    d_list=[]
    elev_Drone=[]
    elev_Ground=[]

    for i in range(len(listMatch)):
        d_list.append(listMatch[i][12])
        elev_Drone.append(listMatch[i][7])
        elev_Ground.append(listMatch[i][8])

    showFlightProfil(d_list, elev_Drone, elev_Ground,dirSaveFig, mute=mute)


def showFlightProfil(d_list,elev_Drone,elev_Ground,dirSaveFig,mute=True):
    # BASIC STAT INFORMATION
    mean_elev = round((sum(elev_Drone) / len(elev_Drone)), 3)
    min_elev = min(elev_Ground)
    max_elev = max(elev_Drone)
    distance = d_list[-1]

    # PLOT ELEVATION PROFILE
    base_reg = min_elev
    plt.figure(figsize=(10, 4))
    plt.plot(d_list, elev_Drone, '.r', label='Drone: ')
    plt.plot(d_list, elev_Ground,'g', label='Ground ')
    plt.fill_between(d_list, elev_Ground, base_reg, alpha=0.1)
    plt.text(d_list[0], elev_Drone[0], "Take off")
    plt.text(d_list[-1], elev_Drone[-1], "*")
    plt.xlabel("Distance (m)")
    plt.ylabel("Altitude (m)")
    plt.grid()
    plt.legend(fontsize='small')
    filepath=dirSaveFig+'\\Flight profil IRdrone'
    plt.savefig(filepath, dpi=75,facecolor='w', edgecolor ='w', orientation ='portrait',
                    papertype = None, format = None, transparent = False,
                    bbox_inches ='tight', pad_inches = 0.1, frameon = None, metadata = None)
    print(Style.GREEN+'Save flight profil in %s'%filepath+Style.RESET)
    if not mute:
        print(Style.YELLOW + 'Look your Drone Flight profil >>>>' + Style.RESET)
        plt.show()


# --------------------------    Utilitaires  ---------------------------------

def explorExifTags(imgName):
    try:
        img = PIL.Image.open(imgName)
        exifTag = {
            PIL.ExifTags.TAGS[k]: v
            for k, v in img._getexif().items()
            if k in PIL.ExifTags.TAGS
        }
        for x, y in exifTag.items():
            print(x, y)
            pass
    except:
        print(Style.RED + "No Exif tags in %s" % imgName + Style.RESET)
    return


def mytest_segmentUTM():
    latP_1 = 45.169366328045726
    lonP_1 = 3.393778698518872
    latP_2 = 45.168724358081818
    lonP_2 = 3.395170848816633
    distP1P2, capP1P2 = segmentUTM(latP_1, lonP_1, latP_2, lonP_2)
    checksum = (distP1P2 - 130.59392177383668) + (capP1P2 - 122.82028841055205)
    if checksum == 0:
        print(Style.YELLOW + 'def segmentUTM is ', Style.GREEN + 'OK' + Style.RESET)
    else:
        print(Style.YELLOW + 'def segmentUTM is ', Style.RED + 'No OK' + Style.RESET)


def mytest_getAltitude():
    latP_1 = 45.169366328045726
    lonP_1 = 3.393778698518872
    latP_2 = 45.168724358081818
    lonP_2 = 3.395170848816633
    latP_3 = 45.1677
    lonP_3 = 3.3961
    altP_1 = 621.
    altP_2 = 607.
    altP_3 = 557.

    listCoordGPS = [(latP_1, lonP_1), (latP_2, lonP_2), (latP_3, lonP_3)]
    listAltitudeOpenElevation = [altP_1, altP_2, altP_3]
    listAltitude = getAltitudeEncode(listCoordGPS)
    checksum = np.sum([listAltitude[i] - listAltitudeOpenElevation[i] for i in range(len(listAltitude))])
    if checksum == 0:
        print(Style.YELLOW + 'def getAltitudeEncode is ', Style.GREEN + 'OK' + Style.RESET)
    else:
        print(Style.YELLOW + 'def getAltitudeEncode is ', Style.RED + 'No OK' + Style.RESET)


def printGPS(gpsLatitude, gpsLongitude, gpsAltitude):
    stringgpsLong = "%s %d° %d\' %.5f\" " % (gpsLongitude[0], gpsLongitude[1], gpsLongitude[2], gpsLongitude[3])
    stringgpsLat = "%s %d° %d\' %.5f\" " % (gpsLatitude[0], gpsLatitude[1], gpsLatitude[2], gpsLatitude[3])
    stringgpsAlt = "%.1f" % gpsAltitude
    print("  Longitude :", stringgpsLong, " |  Latitude :", stringgpsLat, " | Altitude : ", stringgpsAlt, " m")
    return


def printPlanVol(planVol):
    print(' Vol à : %s  (%s)  du %s   '
          '\n Client: %s'
          '\n Pilote: %s sur drone %s'
          '\n Caméra IR: %s'
          '\n Filtre infrarouge  IR %i nm'
          '\n Météo : %s  Vent %.1f m/s  Température %.1f °C' %
          (planVol['etude']['lieu'], (planVol['etude']['GPS'] + '  ' + str(planVol['etude']['altitude']) + ' m'),
           (planVol['etude']['date']),
           planVol['etude']['client'],
           planVol['etude']['pilote'],
           (planVol['drone']['marque'] + ' ' + planVol['drone']['type']),
           (planVol['cameraIR']['marque'] + '  ' + planVol['cameraIR']['type']),
           (planVol['images']['filtreIR']),
           planVol['meteo']['ensoleillement'], planVol['meteo']['vent'], planVol['meteo']['temperature']
           )
          )
    return



#
# ------------------   traitement Alain  version en chantier  !  -------------
#
#   partie à totalement reprendre   l'idée est là mais la manière ... moins :-D
#

from application import registrationCached,warp

def  imgMultiSpectral(imList, multispectralPath,bandIR=1,typVisu=False, debug=False):
    ircal=ut.cameracalibration(camera='sjcam')

    for idx in range(len(imList)):
        imgMultiSpectralName = multispectralPath+'%i-IRdrone-Multi.jpg' % (idx)
        imgIRName = multispectralPath+'%i-IRdrone-IR.jpg' % (idx)

        viImg=Image(imList[idx][0],'Vi')
        irImg=Image(imList[idx][1],'Ir')


        # ONLY PERFORM IMAGE REGISTRATION WHEN NOT CACHED
        homog, visimgSize, _aligned = registrationCached(
            viImg,
            irImg,
            ircalib=ircal,
            debug=debug
        )
        aligned = pr.Image(
            warp(irImg, ircal, homog, visimgSize),
            name="REGISTERED IMAGE"
        )
        aligned.save(imgIRName)


        #____________________________________________________________
        #
        #                 Traitement  multi spectral :  MultiSpectralBand
        #
        #   > Commencer par définir la Class du traitement.
        #       C'est là qu'on fait les manipulations des couches
        #   > Instancier la Class juste en dessous pour pouvoir l'appeler
        #   > Appeler le traitement dans le ipipe
        #
        #_____________________________________________________________
        class MultiSpectralBand(ipipe.ProcessBlock):
            def apply(self,Vi,iR, val01,val02,**kwargs):
                imMultiSpectral = np.zeros_like(Vi)    # float
                IRlayer = np.zeros_like(Vi[:,:,0])

                #  Calcul de la bande IR
                #
                # Attention en float les valeurs des couches de l'image doient être entre 0 et 1

                BandIR = bandIR

                if bandIR == 1 :
                    #  La bande spectrale RedEdge classique est définie entre 715nm - 745nm
                    #  Ici elle est simulée à partir de l'image IRsolaire  720nm -1100nm
                    #  L'opération est réalisée en deux phases :
                    #    1)   soustraction des canaux IRr et IRg  (normalisés sur le max)
                    #    2)   étalement entre 11 et 245   pour du 8bits entre 0 et 255
                    #    attention en float  l'image est définie entre 0 et 1 (et pas entre 0 et 255!)
                    IRlayer=iR[:, :, 0]/np.max(iR[:,:,0])-iR[:, :, 1]/np.max(iR[:,:,1])
                    minIRredEdge=5./255.
                    maxIRredEdge=250./255.
                    aRedEdge=(minIRredEdge-maxIRredEdge)/(np.min(IRlayer)-np.max(IRlayer))
                    bRedEdge=maxIRredEdge-aRedEdge*np.max(IRlayer)
                    IRlayer =  ( aRedEdge * IRlayer + bRedEdge )

                elif BandIR == 2 :
                    #  La bande spectrale IR est représentée par la couche IRred 720nm - 1100nm
                    #  si  le filtre fixé sur l'objectif de la caméra IR est un filtre IR720nm
                    IRlayer = iR[:, :, 0] / np.max(iR[:, :, 0])

                elif BandIR == 3 :
                    #  Bande NIR (Near Infra Red)
                    #  On utilise la couche IRblue  qui par combinaison du filtre matriciel de Bayer
                    #  et du filtre d'objectif IR720nm représente asser bien la bande NIR.
                    #  On obtient une bande NIR 'large'   820nm 1100nm
                    IRlayer =  iR[:, :, 2]/np.max(iR[:,:,2])

                elif BandIR == 4 :
                    #  La bande spectrale IRrgb est la somme des trois couches IRr,IRg et IRb   entre 720nm et 1100nm
                    #  On utilise les trois couches de l'image de la caméra IR munie du filtre IR720nm
                    IRlayer =  (iR[:, :, 0] + iR[:, :, 1] + iR[:, :, 2] )/3.


                # Choix de visualiser la couche IR seule   ou bien le traitement multispectral  (trois couches)
                visuCouche = False

                if typVisu  :
                    return IRlayer
                else:
                    # Transformation multi spectrale avec shift du spectre
                    #      spectral band   VIgreen > layer blue
                    #      spectral band   VIred   > layer green
                    #      spectral band   IR      > layer red
                    imMultiSpectral[:, :, 0] = val01*IRlayer      # spectral band   IR       > layer red
                    imMultiSpectral[:, :, 1] = Vi[:, :, 0]  # spectral band   VIred    > layer green
                    imMultiSpectral[:, :, 2] = Vi[:, :, 1]  # spectral band   VIgreen  > layer blue
                    return imMultiSpectral

        multiSpectralBand=MultiSpectralBand('dif',
                              slidersName=['luminosité','  none'],
                              vrange=[(0.7,1.,1.),(0.,1.,0.5)],
                              inputs=[1,2])

        ##_________________________________________________________________________________
        #   Commentaire sur ipipeq
        #  Attention floatpipe=False INDISPENSABLE  pour valeurs autres que entre 0 et 255
        #
        #  Dans le ipipe le nom sliders est ambigüe. Il désigne la suite des traitements qui seront appliqués
        #  Attention certains traitements ne s'appliquent qu'à une image trois canaux (type RGB)
        #  Par exemple le "slider"     ipipe.WB  ne s'applique qu'à une image tr-icanaux
        ##_________________________________________________________________________________
        ipipe.ImagePipe(
            [
                1.*viImg[0]/255.,
                1.*aligned[0]/255.
            ],
            sliders=[multiSpectralBand],
            floatpipe=False ,
            # winname= "IMAGE %d : VISIBLE versus registered IR image - Use S to save"%(imageRange[idx]),
            winname="%d -- " % idx + "VISIBLE:q  %s" % viImg + "---   FUSED WITH   --- IR : %s" % irImg
        ).gui()  # save(name=imgMultiSpectralName)

        #_________________________           fin de mon traitement     _____________________________________


        #___________________________________________________________________________________________________
        #
        #               Traitement ViR    avec la couche IRrgb  totalement règlable  G/R/IR > B/G/R
        #
        #___________________________________________________________________________________________________
        if False:
            BnWIR = ipipe.BnW("MonochromeIR", inputs=[2], outputs=[2], slidersName=[])
            brIR = ipipe.Brightness("BrightnessIR", inputs=[2], outputs=[2])
            gamIR = ipipe.Gamma("GammaIR", inputs=[2], outputs=[2])

            forcedParams = {
                "BrightnessIR": [-0.420000],
                "GammaIR": [0.160000],
                "Mix IR and Visible": [1.096000, 0.856000, 0.946000],
                "ALPHA": [0.870000],
            }

            """
            ipipe.ImagePipe(
                [
                    viImg[0],
                    aligned[0]
                ],
                sliders=[BnWIR, brIR, gamIR, mix, ipipe.ALPHA, preconvert],
                # winname= "IMAGE %d : VISIBLE versus registered IR image - Use S to save"%(imageRange[idx]),
                winname="%d -- " % idx + "VISIBLE:  %s" % viImg + "---   FUSED WITH   --- IR : %s" % irImg,
                **forcedParams,
            ).save(name=imgMultiSpectralName)   #gui()
            """
            ipipe.ImagePipe(
                [
                    viImg[0],
                    aligned[0]
                ],
                sliders=[BnWIR, brIR, gamIR, mix, ipipe.ALPHA],
                # winname= "IMAGE %d : VISIBLE versus registered IR image - Use S to save"%(imageRange[idx]),
                winname="%d -- " % idx + "VISIBLE:q  %s" % viImg + "---   FUSED WITH   --- IR : %s" % irImg,
                **forcedParams,
            ).save(name=imgMultiSpectralName)  #gui()




class MixIR(ipipe.ProcessBlock):
    def apply(self, vis, ir, coeffr, coeffg, coeffb, **kwargs):
        out = vis.copy()
        out[:, :, 0] = coeffr * ir[:, :, 0]
        out[:, :, 1] = coeffg * vis[:, :, 0]
        out[:, :, 2] = coeffb * vis[:, :, 0]
        return out


mix = MixIR("Mix IR and Visible", inputs=[0, 2], slidersName=["r", "g", "b"], vrange=(0.7, 1.3, 1.))

# -----------------------------------------------------------------------------
#                          Programme Principal 
# -----------------------------------------------------------------------------

if __name__ == "__main__":

    dirPlanVol = r'C:\Air-Mission\FlightPath.xlsx'

    versionIRdrone = '1.01'

    # ----------------------------------------------------
    #        Début du programme
    timeDebut = datetime.datetime.now()
    startTime=time.clock()
    print(Style.CYAN + 'Start IRdrone-v%s  at  %s ' % (versionIRdrone, timeDebut.time()) + Style.RESET)

    # ----------------------------------------------------
    # 1 > Extraction des données du vol
    #     Date, heure, dossier d'images, synchro des horloges, type du drone et de la caméra IR ...
    #     Construction de la liste des images prises lors du vol (Drone et IR)

    imgListDrone, deltaTimeDrone, timeLapseDrone, imgListIR, deltaTimeIR, timeLapseIR, dateEtude, dirNameIRdrone \
        = extractFlightPlan(dirPlanVol, mute=True)

    # ----------------------------------------------------
    # 2 > Appariement des images des deux caméras
    #     On cherche les paires d'images Vi et IR prises au même instant.
    #     On peut éventuellement visualiser les paires d'images  IR et Vi.
    #     Ces images sont  sauvegardées dans le dossier  dirNameIRdrone
    #     qui a été spécifié dans le plan de vol  (.xlxs)
    #
    #   - Construction de  la trace GPS qui relie les points où ont été prises les paires d'images. (file format .gpx)
    #         (l'altitude est celle du drone / au pt de décollage)
    #   - Ecriture dans la feuille Summary du fichier Excel Plan de Vol
    #   - Tracé du profil de vol du drone dans une figure (file format .png)
    listImgMatch = matchImagesFlightPath(imgListDrone, deltaTimeDrone, timeLapseDrone,
                                         imgListIR, deltaTimeIR, timeLapseIR, mute=True)
    seeDualImages = False
    if seeDualImages:
        if len(listImgMatch) < 2:
            modulo = 1
        else:
            modulo = 3
        showDualImages(listImgMatch, modulo=modulo, seeDualImages=seeDualImages)  #visu des paires trouvées


    writeGPX(listImgMatch, dirNameIRdrone, dateEtude, mute=True)  # ecriture  du tracé GPS au format gpx Garmin

    flightPlanSynthesis = summaryFlight(listImgMatch, seaLevel=True, mute=True)

    flightProfil(flightPlanSynthesis, seaLevel=True, dirSaveFig=dirNameIRdrone,mute=True)


    # ----------------------------------------------------
    # 3 > Traitement des paires d'images

    # ToDo  Placer ici le process Image pipe de Balthou ...
    # ToDo  Pouvoir spécifier le chemin de sauvegarde des images IR redressssées , VIR , NDVI
    # ToDo  Pouvoir choisir la liste des trairtements à faire ?

    #   bandiR  1> RedEdge  2> IRred    3> NIR    4>  IRrgb
    #   typVisu True > IR layer          False>  ViR

    # Demande le type de traitement à appliquer à l'image
    bandiR = input(Style.MAGENTA+"Type de bande infrarouge : 1> RedEdge  2> IRred    3> NIR    4>  IRrgb "+Style.RESET)
    # Demande le type de visualisation à l'écran
    typeVisualisation = input(Style.MAGENTA+"Type de visualisation : 1> IR layer   2> ViR "+Style.RESET)
    typeVisu=False
    if typeVisualisation==1: typeVisu=True

    folderPath = dirNameIRdrone+'\\test_traitement\\'
    imgMultiSpectral(listImgMatch, folderPath, bandIR=bandiR, typVisu=typeVisu, debug=False)


    # -----------------------------------------------------
    # 4 > Résumé du traitement

    writeSummaryFlight(flightPlanSynthesis, dirPlanVol)

    # ----------------------------------------------------
    #        Fin du programme
    timeFin = datetime.datetime.now()
    stopTime = time.clock()
    tempsExe = stopTime-startTime

    mytest_getAltitude()

    print(Style.CYAN + 'End IRdrone-v%s  at %s   CPU : %3.f s' % (versionIRdrone, timeFin.time(), tempsExe) + Style.RESET)
