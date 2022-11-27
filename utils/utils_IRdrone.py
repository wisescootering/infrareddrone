# -*- coding: utf-8 -*-
# pylint: disable=C0103, C0301, W0703

"""
Created on 2021-10-05 19:17:16
version 1.05 2021-12-01
version 1.06 2021-12-31 16:41:05.   theoretical Yaw,Pitch,Roll for NIR images
version 1.07 2022-02-15 18:50:00.   Class ShootPoint
version 1.3  2022-09-27 19:37:00
@authors: balthazar/alain
"""




import logging
import os
import os.path as osp
import pandas as pd
import sys
import pickle
import json
from copy import copy, deepcopy
from pathlib import Path
import subprocess
sep = '\\' if os.name == "nt" else "/"
try:
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename, askdirectory
except:
    pass
from operator import itemgetter
import datetime
import openpyxl
from openpyxl import Workbook
import numpy as np
from datetime import timedelta

sys.path.append(osp.join(osp.dirname(__file__), ".."))
import irdrone.process as pr
import irdrone.utils as ut
from irdrone.utils import Style
import utils.utils_GPS as uGPS
import utils.utils_odm as odm
import utils.utils_IRdrone_Class as IRcl
import utils.angles_analyzis as analys
import config as cf



# -----   Convertisseurs de dates   Exif<->Python  Excel->Python    ------
def dateExcelString2Py(dateTimeOriginal):
    """
    :param dateTimeOriginal: Date of shooting of an image. String format
    :return: datePhotoPython     (datetime python)
    """
    imgYear = int(dateTimeOriginal[0:4])
    imgMonth = int(dateTimeOriginal[5:7])
    imgDay = int(dateTimeOriginal[8:11])
    imgHour = int(dateTimeOriginal[11:13])
    imgMin = int(dateTimeOriginal[14:16])
    second = dateTimeOriginal[17:].__str__()
    if len(second) <= 2:
        imgSecond = int(second)
        imgMicroSecond = 0
    else:
        second, microsecond = second.split('.')
        imgSecond = int(second)
        imgMicroSecond = int(microsecond)

    datePy = datetime.datetime(imgYear, imgMonth, imgDay, imgHour, imgMin, imgSecond, imgMicroSecond)

    return datePy


def dateExif2Py(exifTag):
    """
    :param exifTag: données Exif d'une image
    :return: datePhotoPython     (datetime python)

        Extraction of the date of the photo taking from its Exif data.
        The date is provided by the camera clock. (accuracy: 1s)
        Format Exif  YYYY:MM:DD  hh:mm:ss  type string
        We follow the python datetime conversion by analysing the Exif character string.
        We get Year Month Day Hour Min and Second string.
        You must convert the string to integer (via the function int() ).

        Please note that the Exif key to use is 'DateTimeOriginal'.
        The Exif 'DateTime' key indicates the date the file was created on the computer!
        The value of the 'DateTimeOriginal' key is a string (string) 'YYYY:MM:DD hh:mm:ss' .
        You must extract the numeric values (integer) DD,MM,YYYY,hh,mm,ss from this string and then build a "python"
        datetime.datetime(YYYY,MM,DD,hh,mm,ss,micros).
        In this form we can manipulate the dates. In particular we can know the duration between two dates (in second).
        It is also possible to know the name of the day and the month.

    """

    dateTimeOriginal = exifTag.get('DateTimeOriginal')  # date exacte de la prise de vue
    datePy = dateExcelString2Py(dateTimeOriginal)

    return datePy


def datePy2Exif(datePy):
    """
    :param datePy:    (Year, Month, Day, Hour, Minute, Second, Microsecond, timezone)
    :return: dateExif  str  format  YYYY:MM:DD hh:mm:ss
    """
    dateExif = str(datePy.year) + ':' + str(datePy.month) + ':' + str(datePy.day) + ' ' \
               + str(datePy.hour) + ':' + str(datePy.minute) + ':' + str(datePy.second)
    return dateExif


def dateExif2Excel(dateExif):
    """
    :param dateExif:   str format  YYYY:MM:DD hh:mm:ss
    :return: dateExcel str format  YYYY-MM-DD hh:mm:ss
    """
    date = dateExif.split(' ')[0].split(':')
    date.extend(dateExif.split(' ')[1].split(':'))
    dateExcel = str(date[0]) + '-' + str(date[1]) + '-' + str(date[2]) + ' ' \
                + str(date[3]) + ':' + str(date[4]) + ':' + str(date[5])

    return dateExcel


def dateExcel2Py(dateExcel):
    """
    :param dateExcel:    str    format  YYYY-MM-DD hh:mm:ss
    :return: datePy     (Year, Month, Day, Hour, Minute, Second, Microsecond, timezone)
    """
    datePy = dateExcelString2Py(dateExcel)
    return datePy


def extract_synchro_from_dict(di):
    # Allow manual override
    dateMission, deltaTimeIR, coord_GPS_take_off = None, None, None
    if di is not None:
        if "synchro_date" in di.keys():
            dateMission = datetime.datetime.strptime(di["synchro_date"], r'%d/%m/%Y %H:%M:%S')
        for inpkey in ["synchro_deltatime", "synchro deltatime", "deltatime"]:
            if inpkey in di.keys():
                deltaTimeIR = di[inpkey]
        for inpkey in ["coord_GPS_take_off", "coord GPS take off", "coord GPS Take Off"]:
            if inpkey in di.keys():
                coord_GPS_take_off = di[inpkey]
    return dateMission, deltaTimeIR, coord_GPS_take_off


def extractDateNir(nameImg):
    imgYear = int(nameImg[0:4])
    imgMonth = int(nameImg[5:7])
    imgDay = int(nameImg[7:9])
    imgHour = int(nameImg[10:12])
    imgMin = int(nameImg[12:14])
    imgSecond = int(nameImg[14:16])
    imgMicroSecond = 0
    dateImg = datetime.datetime(imgYear, imgMonth, imgDay, imgHour, imgMin, imgSecond, imgMicroSecond)
    return dateImg


# ------------------     Plan de Vol      ------------------------------

def extractFlightPlan(dirPlanVol, mute=True):
    """
    Extract the datas of Fligth Plan
    > Chemin des images Visible et IR, type de drone et de caméra, synchroniation horloges ...
    > Images list of Drone and IR cameras

    :param dirPlanVol:  path of Fligth Plan in  a Excel
                 mute: affiche des informatios si True  (utile en phase debug)
    :return: (planVol,imgListDrone,deltaTimeDrone,timeLapseDrone,imgListIR,deltaTimeIR,timeLapseIR,
    dirNameIRdrone,coordGPS_TakeOff,altiTakeOff)   )
    """
    timeLapseDrone, deltaTimeDrone, timeLapseIR, deltaTimeIR = 2.0, 0., 3.0, None
    offsetTheoretical = [0., 0., 0.]
    # typeDrone, typeIR = "FC3170", "M20"
    typeDrone = None
    if osp.basename(dirPlanVol).lower().endswith(".xlsx"):
        planVol = readFlightPlan(dirPlanVol, mute=mute)

        dateMission, deltaTimeIR, coord_GPS_take_off = extract_synchro_from_dict(None)
        if planVol["synchro"] is not None:
            synchro_file = osp.join(osp.dirname(dirPlanVol), planVol["synchro"])
            assert osp.isfile(synchro_file), "no synchro file {}".format(synchro_file)
            di_synchro = np.load(synchro_file, allow_pickle=True).item()
            dateMission, deltaTimeIR, coord_GPS_take_off = extract_synchro_from_dict(di_synchro)
            if coord_GPS_take_off is not None:
                planVol['mission']['coord GPS Take Off'] = coord_GPS_take_off
            if dateMission is not None:
                planVol['mission']['date'] = dateMission
            if deltaTimeIR is not None:
                planVol['mission']['deltatime'] = deltaTimeIR
        dirNameIRdrone = planVol['images']['repertoireViR  (save)']  # folder for save photography  VIR,  NDVI (output)
        dirNameDrone = planVol['images']['repertoireDrone']  # Drone photography folder   (input)
        dirNameIR = planVol['images']['repertoireIR']  # IR photography folder (input)
        if planVol['mission']['date'] is not None:
            dateMission = planVol['mission']['date']  # date of flight > format DD MM et YYYY
        typeDrone = planVol['drone']['type']  # type of drone (see in the Exif tag of the image of the drone)
        extDrone = planVol['images']['extDrone']  # file format Vi
        # typeIR = planVol['cameraIR']['type']  # type of camera in use (see in the Exif tag of the image of the IR camera)
        timeLapseDrone = float(planVol['drone']['timelapse'])  # Time Lapse of Drone camera
        timeLapseIR = float(planVol['cameraIR']['timelapse'])  # Time Lapse of IR camera
        extIR = planVol['images']['extIR']  # file format  IR
        deltaTimeDrone = float(planVol['drone']['deltatime'])  # decalage horloge caméra du drone / horloge de référence
        if planVol['cameraIR']['deltatime'] is not None:
            deltaTimeIR = float(
                planVol['cameraIR']['deltatime'])  # decalage horloge caméra infrarouge /horloge de référence
        regex_nir, regex_drone = "*.%s" % extIR, '*.%s' % extDrone
    elif osp.basename(dirPlanVol).lower().endswith(".json"):
        with open(dirPlanVol, 'r') as openfile:
            di = json.load(openfile)
        for inpkey in ["input", "rootdir", "folder", "main"]:
            if inpkey in di.keys():
                assert osp.isdir(di[inpkey]), "Please provide a correct folder for key {} = {}".format(inpkey,
                                                                                                       di[inpkey])
                dirPlanVol = di[inpkey]
        if osp.isfile(dirPlanVol):
            dirPlanVol = os.path.dirname(dirPlanVol)
        dirNameIRdrone = di["output"]

        # dirNameDrone = osp.dirname(di["visible"])
        # extDrone = osp.basename(di["visible"]).split('.')[-1]
        # dirNameIR = osp.dirname(di["nir"])
        # extIR = osp.basename(di["nir"]).split('.')[-1]
        dirNameIR, dirNameDrone = None, None
        regex_nir, regex_drone = di["nir"], di["visible"]
        dateMission, deltaTimeIR, coord_GPS_take_off = extract_synchro_from_dict(None)
        if "synchro" in di.keys():
            if osp.isfile(di["synchro"]):
                synchro_file = di["synchro"]  # provided an absolute path
            else:
                synchro_file = osp.join(dirPlanVol, di["synchro"])
            assert osp.isfile(synchro_file), "Synchro file shall be provided {}".format(synchro_file)
            di_synchro = None
            if osp.basename(synchro_file).lower().endswith(".json"):
                with open(synchro_file, 'r') as openfile:
                    di_synchro = json.load(openfile)
            elif osp.basename(synchro_file).lower().endswith(".pkl") or osp.basename(synchro_file).lower().endswith(
                    ".npy") or osp.basename(synchro_file).lower().endswith(".synchro"):
                di_synchro = np.load(synchro_file, allow_pickle=True).item()
            dateMission, deltaTimeIR, coord_GPS_take_off = extract_synchro_from_dict(di_synchro)
        for inpkey in ["timelapse_nir", "nir_timelapse", "timelapse nir", "nir timelapse", "interval nir",
                       "interval_nir"]:
            if inpkey in di.keys():
                timeLapseIR = di[inpkey]
        for inpkey in ["timelapse_visible", "visible_timelapse", "timelapse visible", "visible timelapse",
                       "interval visible", "interval_visible"]:
            if inpkey in di.keys():
                timeLapseDrone = di[inpkey]
        for inpkey in ["offsetTheoretical", "offset_Theoretical", "offset", "offset angles",
                       "offset_angles"]:
            if inpkey in di.keys():
                offsetTheoretical = di[inpkey]
                assert isinstance(offsetTheoretical, list) and len(offsetTheoretical) ==3
                logging.info("Found offset angles")
        # Allow manual override
        dateMission_over, deltaTimeIR_over, coord_GPS_take_off_over = extract_synchro_from_dict(di)
        if dateMission_over is not None:
            logging.warning("Manually overriden synchro_date")
            dateMission = dateMission_over
        if deltaTimeIR_over is not None:
            logging.warning("Manually overriden synchro_deltatime")
            deltaTimeIR = deltaTimeIR_over
        if coord_GPS_take_off_over is not None:
            logging.warning("Manually overriden GPS take off coordinate")
            coord_GPS_take_off = coord_GPS_take_off_over

        planVol = dict(mission={}, drone={})
        planVol['mission']['date'] = dateMission
        planVol['mission']['coord GPS Take Off'] = coord_GPS_take_off
        planVol['drone']['timelapse'] = timeLapseDrone
        planVol["offset_angles"] = offsetTheoretical

    else:
        raise NameError("File not supported")
    assert deltaTimeDrone is not None, "Need to provide decent synchronization"
    #    Liste des images de l'étude.
    #    Une liste pour les images du drone et une liste pour les images de la caméra infrarouge
    #    Chaque élément de la liste est un triplet (file name image , path name image, date image)
    if osp.isfile(dirPlanVol):
        dirPlanVol = os.path.dirname(dirPlanVol)
    dirNameDrone = reformatDirectory(dirNameDrone, rootdir=dirPlanVol)
    dirNameIR = reformatDirectory(dirNameIR, rootdir=dirPlanVol, makeOutdir=True)
    dirNameIRdrone = reformatDirectory(dirNameIRdrone, rootdir=dirPlanVol, makeOutdir=True)
    imgListDrone = creatListImgVIS(dirNameDrone, dateMission, regex_drone, timeLapseDrone, deltaTimeDrone,
                                   cameraModel=typeDrone)

    imgListIR = creatListImgNIR(dirNameIR, regex_nir)

    if not mute:
        if len(imgListDrone) == 0:
            print('No visible images detected for this flight.')
            sys.exit(2)
        else:
            print('%i visible images detected for this flight.' % len(imgListDrone))
            print([(imgListDrone[i][0], imgListDrone[i][2]) for i in range(len(imgListDrone))])

        if len(imgListIR) == 0:
            print('No near infrared images detected for this flight.')
            sys.exit(2)
        else:
            print('%i near infrared images detected for this flight.' % len(imgListIR))
            print([(imgListIR[i][0], imgListIR[i][2]) for i in range(len(imgListIR))])

        if len(imgListDrone) == 0 and len(imgListIR) == 0:
            print('%i visible images (Vi) and %i near infrared images (NiR) for this flight.' %
                  (len(imgListDrone), len(imgListIR)))
    planVol["timelapse_vis_interval"] = timeLapseDrone
    planVol["delta_time_vis"] = deltaTimeDrone
    planVol["timelapse_ir_interval"] = timeLapseIR
    planVol["delta_time_ir"] = deltaTimeIR
    planVol["images_path_list_vis"] = imgListDrone
    planVol["images_path_list_ir"] = imgListIR
    planVol["out_images_folder"] = dirNameIRdrone
    return planVol


def offsetAnglesCheck(planVol, mute=False):
    offsetTheoretical = [0., 0., 0.]
    try:
        assert planVol['images']['offset_angles'] is not None
        offset_angles = planVol['images']['offset_angles'].split(',')
        assert isinstance(offset_angles, list) and len(offset_angles) == 3
        offsetTheoretical = []
        for a in offset_angles:
            angle = float(a)
            if not(-20 <= angle <= 20):
                angle = 0.
            offsetTheoretical.append(angle)
        logging.info("Found offset angles")
    except AssertionError as e:
        print(
            Style.YELLOW + '[warning] Offset angles have not been found or are incorrect in xlsx file. Using [0., 0., 0.]' + Style.RESET)
    except ValueError as e:
        print(
            Style.YELLOW + '[warning] Offset angles have not been found or are incorrect in xlsx file. Using [0., 0., 0.]' + Style.RESET)
    if not mute: print(offsetTheoretical)

    return offsetTheoretical


def readFlightPlan(pathPlanVolExcel, mute=None):
    """
        Read the Flight Plan  in Excel file.

    :param pathPlanVolExcel: chemin du fichier Excel qui contient le plan de vol  (string)
           mute : affiche des informatios si True  (utile en phase debug)
    :return: planVol  Dictionnaire contenant les données du plan de vol  (dic)
    """
    workbook = openpyxl.load_workbook(pathPlanVolExcel, read_only=True, data_only=True)

    sheet = workbook.worksheets[0]

    nuetude = 2  # 2      numéro première ligne de données du fichier Excel
    nudrone = (nuetude + 8) + 1  # 2+8+1 =11
    nucameraIR = (nudrone + 7) + 1  # 11+7+1=19
    nuimages = (nucameraIR + 5) + 1  # 19+5+1=25
    numeteo = (nuimages + 11) + 1  # 25+11+1=37
    planVol = {'mission':
                   {'client': sheet.cell(nuetude + 1, 2).value,
                    'lieu': sheet.cell(nuetude + 2, 2).value,
                    'coord GPS Take Off': sheet.cell(nuetude + 3, 2).value,
                    'altitude Take 0ff': sheet.cell(nuetude + 4, 2).value,
                    'date': sheet.cell(nuetude + 5, 2).value,  # ' DD-MM-YYYY  hh:mm:ss
                    'heure_solaire': sheet.cell(nuetude + 6, 2).value,
                    'numero_du_vol': sheet.cell(nuetude + 7, 2).value,
                    'pilote': sheet.cell(nuetude + 8, 2).value
                    },
               'drone':
                   {'marque': sheet.cell(nudrone + 1, 2).value,
                    'type': sheet.cell(nudrone + 2, 2).value,
                    'timelapse': sheet.cell(nudrone + 3, 2).value,
                    'deltatime': sheet.cell(nudrone + 4, 2).value,
                    'imatriculation': sheet.cell(nudrone + 5, 2).value,
                    'altitude  de vol': sheet.cell(nudrone + 6, 2).value,
                    'synchro horloges': sheet.cell(nudrone + 7, 2).value  # libre
                    },
               'cameraIR':
                   {'marque': sheet.cell(nucameraIR + 1, 2).value,
                    'type': sheet.cell(nucameraIR + 2, 2).value,
                    'timelapse': sheet.cell(nucameraIR + 3, 2).value,
                    'deltatime': sheet.cell(nucameraIR + 4, 2).value,
                    'synchro horloges': sheet.cell(nucameraIR + 5, 2).value  # libre
                    },
               'images':
                   {'repertoireViR  (save)': sheet.cell(nuimages + 1, 2).value,
                    'repertoireDrone': sheet.cell(nuimages + 2, 2).value,
                    'extDrone': sheet.cell(nuimages + 3, 2).value,
                    'filtreDrone': sheet.cell(nuimages + 4, 2).value,
                    'libre_1': sheet.cell(nuimages + 5, 2).value,  # libre
                    'libre_2': sheet.cell(nuimages + 6, 2).value,  # libre
                    'repertoireIR': sheet.cell(nuimages + 7, 2).value,
                    'extIR': sheet.cell(nuimages + 8, 2).value,
                    'filtreIR': sheet.cell(nuimages + 9, 2).value,
                    'offset_angles': sheet.cell(nuimages + 10, 2).value,  #
                    'libre_4': sheet.cell(nuimages + 11, 2).value  # libre
                    },
               'meteo':
                   {'ensoleillement': sheet.cell(numeteo + 1, 2).value,
                    'vent': sheet.cell(numeteo + 2, 2).value,
                    'temperature': sheet.cell(numeteo + 3, 2).value,
                    'humidite': sheet.cell(numeteo + 4, 2).value
                    },
               'synchro': sheet.cell(numeteo + 5, 2).value
               }
    workbook.close()
    planVol['offset_angles'] = offsetAnglesCheck(planVol, mute=mute)

    if not mute:
        printPlanVol(planVol)

    return planVol


def creatListImgNIR(dirName, rege):
    """
    :return:  imgList   [(),...,(file name image , path name image, date image),...,()]
    Creation of the list of IR photos (raw given extension) and contained in the dirName directory
    If camera="*' then the list includes all images with the extension typeImg (here RAW)
    For an SJcam M20 we extract the date of the file name, because the Exif data in
    the SJCam RAW file is unreadable by exifread or PIL software. ExifTags
    """
    print(Style.CYAN + 'INFO : ------ Creating the list of near-infrared spectrum images' + Style.RESET)
    imlist = sorted(ut.imagepath(imgname=rege, dirname=dirName))
    imgList = []
    for i in range(len(imlist)):
        nameImg = imlist[i].split(sep)[len(imlist[i].split(sep)) - 1]
        dateImg = extractDateNir(nameImg)
        imgList.append((nameImg, imlist[i], dateImg))  # added to image list
    return imgList


def creatListImgVIS(dirName, dateMission, rege, timelapse, deltatime, cameraModel=None, debug=False):
    """
    :param dirName:
    :param dateMission:
    :param debug:
    :param cameraModel: optional to filter out wrong cameras
    :return:  imgList   [(), ...,(file name image , path name image, date image), ..., ()]
    """

    # Creation of the list of photos (extension given ex: JPG, DNG) and contained in the dirName directory.
    # If camera="*' then the list contains all the images with the extension typeImg.
    # If camera ='DJI'  the list only includes images taken by the drone.
    # It is not possible to filter the SJcam images with the file name (you have to look at the Exif data of the image).

    print(Style.CYAN + 'INFO : ------ Creating the list of visible spectrum images' + Style.RESET)
    imlist = sorted(ut.imagepath(imgname=rege, dirname=dirName))

    imgList = []
    j = 0
    if dateMission is None:
        dateMission = pr.Image(imlist[0]).date
    for i in range(len(imlist)):
        img = pr.Image(imlist[i])
        img.camera["timelapse"] = float(timelapse)
        img.camera["deltatime"] = float(deltatime)

        try:
            # Extract Exif data from the image. If no Exif data, image is ignored.
            debug = True
            cameraMakerImg = img.camera['maker']
            cameraModelImg = img.camera['model']
            serial_number = img.camera['serial number']

            if cameraModel is None or cameraModelImg == cameraModel or ( cameraMakerImg == 'irdrone' and cameraModelImg == 'multispectral'):
                dateImg = img.date
                if (dateImg.year, dateImg.month, dateImg.day) == (dateMission.year, dateMission.month, dateMission.day):
                    j += 1
                    nameImg = imlist[i].split(sep)[len(imlist[i].split(sep)) - 1]
                    imgList.append((nameImg, imlist[i], dateImg))  # Add to image list.
                else:
                    # Image was taken by another camera. This image is ignored.
                    if debug: print(Style.YELLOW,
                                    '%s was taken on  %i %i %i. This date is different from the mission date %i %i %i'
                                    % (imlist[i], dateImg.day, dateImg.month, dateImg.year, dateMission.day,
                                       dateMission.month,
                                       dateMission.year), Style.RESET)
            else:
                if debug: print(Style.YELLOW,
                                '%s was taken by another camera (Model %s) ' % (imlist[i], cameraModelImg),
                                Style.RESET)
        except:
            if debug: print("No Exif tags in %s" % imlist[i])

    if float(timelapse) > 0:
        # Dates are only corrected if the images have been taken in hyperlapse.
        # For single shoot images the rectification would not make sense.
        imgList = timelapseRectification(imgList, dateMission)

    return imgList


def timelapseRectification(imgList, date_mission):
    """
    :param imgList:         [(), ...,(file name image , path name image, original date image), ..., ()]
    :return:  new_imgList   [(), ...,(file name image , path name image, rectified date image ), ..., ()]

    Correction of the timelapse dates.
    The recording date is rounded to the nearest second in the Exif data and the timlapse pitch value is not exactly
    an integer value from where sometimes a brutal 'jump' of a second for "ratrapting time".

    This correction only makes sense if the images are saved in timelapse mode!
    """

    imgList = sorted(imgList, key=itemgetter(2), reverse=False)  # Sort according to the date of shooting.

    stopTime = imgList[-1][2]
    nbImage = int(imgList[-1][0].split('_')[-1].split('.')[0])
    _, av_Timelapse_timedelta = average_Timelapse(date_mission, stopTime, nbImage, mute=False)

    new_imgList = []
    for i in range(len(imgList)):
        # Substitution by the corrected date.
        newDate = imgList[-1][2] - (len(imgList) - 1 - i) * av_Timelapse_timedelta
        new_imgList.append((imgList[i][0], imgList[i][1], newDate, imgList[i][2]))

    return new_imgList


# -----------------------------------     Appairement des images      -------------------------------------------------


def matchImagesFlightPath(imgListDrone,
                          deltaTimeDrone,
                          timeLapseDrone,
                          imgListIR,
                          deltaTimeIR,
                          timeLapseIR,
                          dateMission=None,
                          mute=False):
    """
    :param imgListDrone:  [...,(file name, path name, date), ...]
    :param deltaTimeDrone:
    :param timeLapseDrone
    :param imgListIR:     [...,(file name, path name, date), ...]
    :param deltaTimeIR:
    :param timeLapseIR:
    :param dateMission:   date of first DNG used for synchronization
    :param mute:
    :return:  listImgMatch   [..., (imgListDrone[i][1], imgListIR[k][1]), ...]
    """

    listImgMatch, DtImgMatch, listdateMatch, listPts = [], [], [], []

    repA, imgListA, deltaTimeA, repB, imgListB, deltaTimeB, timeDeviationMax = \
        matchImagesAorB(timeLapseDrone, imgListDrone, deltaTimeDrone, timeLapseIR, imgListIR, deltaTimeIR)

    dateA = [imgListA[i][2] for i in range(len(imgListA))]
    dateB = [imgListB[k][2] for k in range(len(imgListB))]
    # Shooting of image B is after shooting of image A if deltaTime < 0
    # three equivalent methods ...
    # datetime.timedelta.total_seconds(dateA[i] - dateB[k])
    # timedelta.total_seconds(dateA[i] - dateB[k])
    # (dateA[i] - dateB[k]).total_seconds()

    deltaTime = [[(dateA[i] - dateB[k]).total_seconds() + deltaTimeB - deltaTimeA
                  for k in range(len(imgListB))] for i in range(len(imgListA))]

    DTime = [[abs(deltaTime[i][k]) for k in range(len(imgListB))] for i in range(len(imgListA))]
    n = 0
    nRejet = 0
    if dateMission is None:
        logging.warning("Please provide the date of the first synchronization DNG image")
        logging.warning("you may get bad image pairing therefore parallax / bad alignment results")
    originDate = copy(imgListA[0][2]) if dateMission is None else copy(dateMission)
    for i in range(len(imgListA)):
        k = np.argmin(DTime[i][:])
        if -timeDeviationMax <= deltaTime[i][k] <= timeDeviationMax:
            # Construction of the image pair IR & Vi  (with rectified DateTime of images)
            # Warning : assume  timeLapseDrone =< timeLapseIR:
            if n == 0:  # first image. Defined time line origin
                originDate = copy(imgListA[i][2])
            kBmatch = k
            n += 1
            timeline = (imgListA[i][2] - originDate).total_seconds()
            DtImgMatch.append(deltaTime[i][k])
            listImgMatch.append((imgListA[i][1], imgListB[kBmatch][1]))
            listdateMatch.append((imgListA[i][2], imgListB[kBmatch][2]))
            listPts.append(IRcl.ShootPoint(numero=n, nameVis=imgListA[i][1], nameNir=imgListB[kBmatch][1],
                                           visDate=str(imgListA[i][2]), nirDate=str(imgListB[kBmatch][2]),
                                           timeDeviation=deltaTime[i][k], timeLine=timeline))

            if not mute:
                print(Style.GREEN + 'N°', n, ' DT ', round(deltaTime[i][kBmatch], 3), 's   ',
                      imgListA[i][0], ' | ', imgListB[kBmatch][0] + Style.RESET)

        else:
            # No match for type B image
            nRejet += 1
            if not mute: print(Style.YELLOW + 'INFO:  Image ', repA, ' ', imgListA[i][0],
                               ' does not match any image ', repB,
                               '.  Minimum time deviation = ',
                               round(min(DTime[i][:]), 3), 's' + Style.RESET)

    if len(listImgMatch) == 0:
        print(Style.RED, 'No pair of Visible-Infrared images were detected for this flight.', Style.RESET)
        sys.exit(2)
    else:
        print(Style.GREEN + ' %i pairs of Visible-Infrared images were detected for the flight on  %s' % (
            len(listImgMatch), dateMission), '\n',
              '%i  images Vi (%s) have been eliminated  ' % (nRejet, repA) + Style.RESET)

    return listImgMatch, DtImgMatch, listdateMatch, listPts


def matchImagesAorB(timeLapseDrone, imgListDrone, deltaTimeDrone, timeLapseIR, imgListIR, deltaTimeIR):
    if timeLapseDrone <= 0 or timeLapseDrone <= timeLapseIR:
        # If timeLapseDrone <= 0 the photos are taken in manual mode.
        # If timeLapseDrone > 0 and  =< timeLapseIR    the photos are taken in time lapse mode.
        # This is the best configuration  (2s DJI | 3s SJCam M20]
        # Warning in some cases (skipping a NIR image) we can have :
        #                             timeDeviationMax = timeLapseIR
        # instead of timeLapseIR / 2
        repB = 'IR'
        repA = 'drone'
        imgListB = imgListIR
        imgListA = imgListDrone
        deltaTimeB = deltaTimeIR
        deltaTimeA = deltaTimeDrone
        timeDeviationMax = timeLapseIR

    else:
        # unwise !
        repA = 'drone'
        repB = 'IR'
        imgListA = imgListDrone
        imgListB = imgListIR
        deltaTimeA = deltaTimeDrone
        deltaTimeB = deltaTimeIR
        timeDeviationMax = timeLapseDrone / 2.

    return repA, imgListA, deltaTimeA, repB, imgListB, deltaTimeB, timeDeviationMax


# -------------------------     Synthèse de informations sur la mission      ------------------------------------------

def spatialAttitude(listPts, listImg):
    flightAngle, gimbalAngle = [], []
    for i in range(len(listImg)):
        img = pr.Image(listImg[i][0])
        finfo = img.flight_info
        flightAngle.append((finfo["Flight Yaw"], finfo["Flight Pitch"], finfo["Flight Roll"]))
        gimbalAngle.append((finfo["Gimbal Yaw"], finfo["Gimbal Pitch"], finfo["Gimbal Roll"]))
        listPts[i].yawDrone = finfo["Flight Yaw"]
        listPts[i].pitchDrone = finfo["Flight Pitch"]
        listPts[i].rollDrone = finfo["Flight Roll"]
        listPts[i].yawGimbal = finfo["Gimbal Yaw"]
        listPts[i].pitchGimbal = finfo["Gimbal Pitch"]
        listPts[i].rollGimbal = finfo["Gimbal Roll"]
    return listPts, flightAngle, gimbalAngle


def spatialCoordinates(listPts, listImg, coordGPSTakeOff, seaLevel, altitude_api_disabled=False):
    listPts, listPtGPS, listCoordGPS = gpsCoordinate(listPts, listImg, coordGPSTakeOff)

    listPts, altGeo, altiTakeOff = altiIGN(listPts, listImg, listPtGPS, coordGPSTakeOff, seaLevel, bypass=altitude_api_disabled)

    listPts, altDroneSol, altGround, altDroneSealLevel = altiDrone(listPts, listImg, listCoordGPS, altGeo, altiTakeOff)

    return listPts, altGeo, altDroneSol, altGround, altDroneSealLevel


def gpsCoordinate(listPts, listImg, coordGPSTakeOff):
    listPtGPS, listCoordGPS = [], []

    for i in range(len(listImg)):
        img = pr.Image(listImg[i][0])
        try:
            listPts[i].gpsSN = img.gps['latitude'][0]
            listPts[i].gpsLat = img.gps['latitude'][4]
            listPts[i].gpsEW = img.gps['longitude'][0]
            listPts[i].gpsLon = img.gps['longitude'][4]
            listPts[i].altTakeOff = img.gps['altitude']
            listPtGPS.append((img.gps['latitude'][4], img.gps['longitude'][4]))
            listCoordGPS.append((img.gps['latitude'][4], img.gps['longitude'][4], img.gps['altitude']))
        except:
            listPts[i].gpsSN = 'N'
            listPts[i].gpsLat = float(coordGPSTakeOff[0])
            listPts[i].gpsWE = 'E'
            listPts[i].gpsLon = float(coordGPSTakeOff[1])
            listPts[i].gpsaltTakeOff = 2.
            listPtGPS.append((float(coordGPSTakeOff[0]), float(coordGPSTakeOff[1])))
            listCoordGPS.append((float(coordGPSTakeOff[0]), float(coordGPSTakeOff[1]), 2.))

    return listPts, listPtGPS, listCoordGPS


def altiDrone(listPts, listImg, listCoordGPS, altGeo, altiTakeOff):
    altDroneSol, altGround, altDroneSealLevel = [], [], []

    for i in range(len(listImg)):
        # drone elevation relative to ground level
        listPts[i].altGround = round(listCoordGPS[i][2] - (altGeo[i] - altiTakeOff), 5)

        altDroneSol.append(round(listCoordGPS[i][2] - (altGeo[i] - altiTakeOff), 5))
        altGround.append(altGeo[i])
        altDroneSealLevel.append(altGeo[i] + altDroneSol[i])

    return listPts, altDroneSol, altGround, altDroneSealLevel


def altiIGN(listPts, listImg, listPtGPS, coordGPSTakeOff, seaLevel, bypass=False):
    """
    For all points of the list, the altitudes of the ground relative to the sea level are determined..
    For France the code uses the API of the IGN (national geographical institute)
    Warning : this API limits the number of points that can be sent in the request
    An attempt is made. If the ground elevation relative to sea level fails, it is set at zero.
    https://geoservices.ign.fr/documentation/services/api-et-services-ogc/calcul-altimetrique-rest
    For each request the IGN API accepts in theory 5000 pts. In practice 189 works but 199 fails!

    """
    print(Style.GREEN + 'Request to  https://wxs.ign.fr/essentiels/alti/rest/elevation' + Style.RESET)
    altGeo = []

    pas = 99
    if len(listImg) < pas:
        pas = len(listImg) - 1

    if seaLevel:
        altiTakeOff = coordGPSTakeOff[2]
        if coordGPSTakeOff[2] < 0:
            altiTakeOff = 0
        #  Splitting the list of points into multiple segments because the number of points in the IGN API is limited.
        for k in range(0, int(round(len(listImg), 0) / pas) + 1):
            listPairPtGps = listPtGPS[k * pas:(k + 1) * pas]
            altGeo += uGPS.altitude_IGN(listPairPtGps, bypass=bypass)
            print(Style.GREEN, '%i / %.0f ' % (k, len(listImg) / pas), Style.RESET)
    else:
        altGeo = [0.0] * len(listPtGPS)
        altiTakeOff = 0.
    for i in range(len(listPts)):
        listPts[i].altGeo = altGeo[i]

    return listPts, altGeo, altiTakeOff


def dist(listPts):
    """
    Calculates : > the distance and cape between point P0 and the next point P1.
                 > the total distance traveled by the drone.

    """
    distP0P1, capP0P1, distFlight = [], [], []

    for i in range(len(listPts)):
        if i == len(listPts) - 1:
            listPts[i].gpsDist, listPts[i].gpsCap = [0, 0]  # last point
        else:
            listPts[i].gpsDist, listPts[i].gpsCap = \
                uGPS.segmentUTM(listPts[i].gpsLat, listPts[i].gpsLon, listPts[i + 1].gpsLat, listPts[i + 1].gpsLon)

        if listPts[i].gpsDist < 0.01 or listPts[i].altGround < 3.:  # lack of precision or drone too low
            listPts[i].gpsDist = 0.00

        if i == 0:
            listPts[0].gpsDistTot = 0.  # first point
        else:
            listPts[i].gpsDistTot = listPts[i - 1].gpsDistTot + listPts[i - 1].gpsDist

        distP0P1.append(listPts[i].gpsDist)
        capP0P1.append(listPts[i].gpsCap)
        distFlight.append(listPts[i].gpsDistTot)

    return listPts, distP0P1, capP0P1, distFlight


def nameImageVisSummary(nameImageComplet):
    imgNameOriginal = nameImageComplet.split(sep)[len(nameImageComplet.split(sep)) - 1]
    str = imgNameOriginal.split('_')[1]
    numeroRef = str.split('.')[0]
    imgExt = str.split('.')[1]
    recordType = imgNameOriginal.split('_')[0]  # DJI (single shoot), TIMELAPSE, HYPERLASE, PANORAMIC  etc
    imgName = '%s_%s.%s' % (recordType, numeroRef, imgExt)

    return imgName


def nameImageNIRSummary(nameImageComplet):
    imgName = nameImageComplet.split(sep)[len(nameImageComplet.split(sep)) - 1]

    return imgName


def summaryFlight(listPts, listImg, planVol, dirPlanVol, offsetTheoreticalAngle=None,
                  seaLevel=False, dirSavePlanVol=None, saveGpsTrack=False,
                  saveExcel=False, savePickle=True, mute=True, altitude_api_disabled=False,
                  optionAlignment=None, ratioSynchro=0.25, muteGraph=False):
    """
    :param listImg:      list of image pairs VI/IR matched at the same point
    :param mute:
    :param saveExcel: save data in Excel file if True
    :param savePickle: save data in Binary file if True
    :param altitude_api_disabled: True if you don't want or can't access IGN API to retrieve altitudes
    :return: summary     listPts
        imgNameVi---------------------- file name of visible RGB color image at point P
        imgNameIRoriginal-------------- file name of original infrared image  at point P
        dateVis ----------------------- date of shooting of the visible image at point P.  Corrected time
        dateNir ----------------------- date of shooting of the near infrared image at point P.
        timeline ---------------------- delay between visible image at point P and the first visible image.
        dateVi ------------------------ date the visible image was shot             [YY-MM-DD hh:mm:ss]
        Lat, Lon, AltGround ----------- GPS coordinates of point P                  [°],[°],[m]
        altGeo ------------------------ altitude of ground / sealevel               [m]
        altTakeOff -------------------- altitude of the drone relative to take-off point  (DJI ref)  [m]
        distToNextPt------------------- distance to next point                                       [m]
        capToNextPt-------------------- direction to the next point relative to the geographic north [°]
        UTM---------------------------- UTM coordinates of point P  xUTM (axe west>east) yUTM (axe south>north) [m],[m]
        flightAngle-------------------- yaw,pitch,roll of drone        [°]
        gimbalAngle-------------------- yaw, pitch, roll of gimbal     [°]
        motionDrone-------------------- x_1,x_2,x_3 axis of the drone; orthogonal to the axis of the drone; z axis [m]
        theoreticalAngleIr2Vi --------- yaw,pitch,roll   [°]


        lists the essential elements of flight after treatment by IRdrone.
        This data will be saved in the Excel file of the flight directory  (sheet "Summary")
        and in binary file
    """
    timeLapseDrone = float(planVol['drone']['timelapse'])  # > 0 if hyperlapse mode  =< O if single shoot mode
    # ---  GPS coordinates, trajectory, take-off, altitude relative to sea level ...
    if planVol['mission']['coord GPS Take Off'] is None:
        logging.warning("Ignoring flight anaytics (altitude of the flight etc...)")
        raise Exception("No GPS take off coordinate in excel - Ignoring flight analytics")
    coordGPSTakeOff, altiTakeOff = uGPS.TakeOff(planVol['mission']['coord GPS Take Off'], bypass=altitude_api_disabled)

    print(Style.CYAN +
          'INFO : ------ Calculation of drone attitude, trajectory, flight profile and theoretical Yaw-Pitch-Roll'
          + Style.RESET)

    # ----------- Drone attitude in space. (angles drone and gimbal)
    listPts, flightAngle, gimbalAngle = spatialAttitude(listPts, listImg)  #

    # ---  GPS coordinates, trajectory, take-off, altitude relative to sea level ...  ---------------------

    listPts, altGeo, altDroneSol, altGround, altDroneSealLevel = \
        spatialCoordinates(listPts, listImg, coordGPSTakeOff, seaLevel, altitude_api_disabled=altitude_api_disabled)

    listPts, distP0P1, capP0P1, distFlight = dist(listPts)

    for i in range(len(listImg)):
        listPts[i].gpsUTM_X, listPts[i].gpsUTM_Y, listPts[i].gpsZone = \
            uGPS.geo2UTM(listPts[i].gpsLat, listPts[i].gpsLon)

    if timeLapseDrone > 0:
        av_timelapse_Vis, _ = average_Timelapse(planVol['mission']['date'], listPts[-1].dateVis,
                                                int(nameImageVisSummary(listPts[-1].Vis).split('_')[-1].split('.')[0]),
                                                mute=True)
        motion_in_DroneAxis(listPts, mute=True)
    else:
        av_timelapse_Vis = timeLapseDrone

    # -- Theoretical angles (yaw, Pitch, Roll) to overlay the infrared image on the visible image.(coarse process)
    listPts, theoreticalPitch, theoreticalYaw, theoreticalRoll = \
        theoreticalIrToVi(listPts, av_timelapse_Vis, offset=offsetTheoreticalAngle)

    # ---------  Plot the flight profile --------------------------------------------------------------------
    dirSaveFig = None
    if not muteGraph:
        dirSaveFig = osp.join(dirSavePlanVol, "Flight Analytics", "Topo")
        if not osp.isdir(dirSaveFig):
            Path(dirSaveFig).mkdir(parents=True, exist_ok=True)

        title = os.path.dirname(dirPlanVol).split("/")[-1]
        analys.flightProfil_plot(distFlight, altDroneSealLevel, altGround, title=title, dirSaveFig=dirSaveFig, mute=False)

    # ---------  Save GPS Track in Garmin format (.gpx) -----------------------------------------------------
        if saveGpsTrack:
            uGPS.writeGPX(listPts, dirSaveFig, planVol['mission']['date'], mute=True)

    # ---- Selecting the image pairs for the alignment among the available images according to the alignment option value :
    #  optionAlignment == 'all' or None   Selecting all available image pairs in the AerialPhotography folder.
    #  optionAlignment == 'best-synchro'         Selecting the best synchronized image pairs for alignment.
    #  optionAlignment == 'best-mapping'         Selection of image pairs for mapping among aligned images.
    #  optionAlignment == 'best-offset'          Selection of image pairs with a timing deviation of less than 0.05s. Used to calculate offset angles.
    ImgMatchForAlignment, PtsForAlignment = [], []
    if optionAlignment == None or optionAlignment == 'all' or timeLapseDrone <= 0:
        ImgMatchForAlignment, PtsForAlignment = selectAllImages(listImg, listPts, planVol, ratioSynchro=ratioSynchro)
    elif optionAlignment == 'best-synchro':
        ImgMatchForAlignment, PtsForAlignment = selectBestSynchro(listImg, listPts, planVol, ratioSynchro=ratioSynchro)
    elif optionAlignment == 'best-mapping':
        ImgMatchForAlignment, PtsForAlignment = selectBestMapping(listImg, listPts, planVol, overlap_x=cf.OVERLAP_X, overlap_y=cf.OVERLAP_Y, ratioSynchro=ratioSynchro)
    elif optionAlignment == 'best-offset':
        timeOffsetMin = 0.039
        print(Style.CYAN + 'INFO : ------ Selects pairs of images with a timing deviation of less than ~ %.4f s. Used to calculate offset angles. '%timeOffsetMin + Style.RESET)
        ImgMatchForAlignment, PtsForAlignment = selectBestOffset(listImg, listPts, planVol, timeOffset=timeOffsetMin,  mute=True)
    else:
        print(Style.RED + '[error] ')
        pass
    # --------- Selection of image pairs for mapping among aligned images. ------------------------------------------
    visualize_matches_on_map(PtsForAlignment, listPts, folder_save=dirSaveFig, name=optionAlignment)
    # ----------  Save summary in Excel format -----------------------------------------------------------
    SaveSummaryInExcelFormat(dirSavePlanVol, saveExcel, listPts, listImg, mute=True)
    # ----------  Save summary in Pickle format -----------------------------------------------------------
    SaveSummaryInNpyFormat(dirSavePlanVol, savePickle, planVol, listPts)

    return ImgMatchForAlignment, PtsForAlignment

def legacy_best_mapping_selection(listPts, folder_save=None):
    return odm.legacy_buildMappingList(listPts, overlap_x=cf.OVERLAP_X, overlap_y=cf.OVERLAP_Y, dirSaveFig=folder_save)

def visualize_matches_on_map(selected_points, all_points_list, folder_save=None, name=None):
    _camera_make, _camera_type, lCapt_x, lCapt_y, _focal_factor, focalPix = lectureCameraIrdrone()
    odm.visu_mapping(selected_points, all_points_list, focal_DJI=focalPix, lCapt_x=lCapt_x, lCapt_y=lCapt_y, dirSaveFig=folder_save, name=name)

def SaveSummaryInExcelFormat(dirMission, saveExcel, listPts, listImg, mute=True):
    dirAnalytics = Path(dirMission) / "Flight Analytics"
    if saveExcel:
        if not osp.isdir(dirAnalytics):
            Path(dirAnalytics).mkdir(parents=True, exist_ok=True)
        summaryExcel = buildSummaryExcel(listPts, listImg)
        writeSummaryFlightExcel(summaryExcel, dirAnalytics)
        if not mute:
            txtSummary = 'List of images of the flight:'
            listSummary = list(summaryExcel)
            for i in range(len(listImg)):
                txtSummary = txtSummary + '\n' + listSummary[i].__str__()
            print(Style.GREEN + txtSummary + Style.RESET)
    return


def SaveSummaryInNpyFormat(dirMission, savePickle, planVol, listPts):
    summaryPickl = buildMissionAndPtsDicPickl(planVol, listPts)
    dirAnalytics = Path(dirMission) / "Flight Analytics"
    if savePickle:
        if not osp.isdir(dirAnalytics):
            Path(dirAnalytics).mkdir(parents=True, exist_ok=True)
        fileNamePickl = 'MissionSummary.npy'
        saveMissionAndPtsPickl(summaryPickl, fileNamePickl, dirAnalytics)
    return


def avAltitude(listPts):
    Hav = sum(listPts[i].altGround for i in range(len(listPts))) / len(listPts)
    return Hav


def buildSummaryExcel(listPts, listImg):
    summaryExcel = []
    for i in range(len(listPts)):
        summaryExcel.append(
            (listPts[i].num,  # 0  N° shooting point
             nameImageVisSummary(listPts[i].Vis),  # 1  original image VIS name
             nameImageNIRSummary(listImg[i][1]),  # 2  original image NIR name
             round(listPts[i].timeDeviation, 3),  # 3  delay between dateTime VI and dateTime IR.
             listPts[i].dateVis.__str__(),  # 4  date of shooting of the visible image. Corrected time
             listPts[i].dateNir.__str__(),  # 5  date of shooting (image NIR)
             round(listPts[i].timeLine, 3),  # 6  Time Line
             listPts[i].dateVis.__str__(),  # 7  date of shooting (image VIS)
             round(listPts[i].gpsLat, 6),  # 8  latitude  +/- dd.ddddd   (N if > 0)
             round(listPts[i].gpsLon, 6),  # 9  longitude +/- dd.ddddd   (E if > 0)
             listPts[i].altGround,  # 10 altitude drone / ground
             round(listPts[i].altGeo, 3),  # 11 altitude ground / sealevel
             round(listPts[i].altTakeOff, 3),  # 12 altitude drone / takeoff
             round(listPts[i].gpsDist, 4),  # 13
             round(listPts[i].gpsCap, 4),  # 14
             round(listPts[i].gpsDistTot, 2),  # 15
             round(listPts[i].gpsUTM_X, 6),  # 16 xUTM
             round(listPts[i].gpsUTM_Y, 6),  # 17 yUTM
             listPts[i].gpsZone,  # 18 zoneUTM
             round(listPts[i].yawDrone, 5),  # 19 yaw drone
             round(listPts[i].pitchDrone, 5),  # 20 pitch drone
             round(listPts[i].rollDrone, 5),  # 21 roll drone
             round(listPts[i].yawGimbal, 5),  # 22 yaw gimbal
             round(listPts[i].pitchGimbal, 5),  # 23 pitch gimbal
             round(listPts[i].rollGimbal, 5),  # 24 roll gimbal
             round(listPts[i].x_1, 3),  # 25 motion in principal axis drone
             round(listPts[i].x_2, 3),  # 26 motion in orthogonal axis drone
             round(listPts[i].x_3, 3),  # 27 motion in Z axis
             round(listPts[i].yawIR2VI, 5),  # 28 theoretical yaw for superimpose imgNIR on imgVIS
             round(listPts[i].pitchIR2VI, 5),  # 29 theoretical pitch for superimpose imgNIR on imgVIS
             round(listPts[i].rollIR2VI, 5),  # 30 theoretical roll for superimpose imgNIR on imgVIS
             listPts[i].bestSynchro,   # 31 value = 1 if best synchro
             listPts[i].bestMapping,   # 32 value = 1 if image pairs selected for mapping among aligned images.
             listPts[i].alignment,     # 33 value = 1 if image pairs selected for process alignment
             round(listPts[i].yawCoarseAlign, 5),  # 34 coarse yaw for superimpose imgNIR on imgVIS
             round(listPts[i].pitchCoarseAlign, 5),  # 35 coarse pitch for superimpose imgNIR on imgVIS
             round(listPts[i].rollCoarseAlign, 5)  # 36 coarse roll for superimpose imgNIR on imgVIS
             )
        )

    return summaryExcel


def buildMissionAndPtsDicPickl(dicMission, listDicPts):
    listDic = [dicMission]
    for n in range(len(listDicPts)):
        listDic.append(listDicPts[n].loadPoint2DicPoint())
    return listDic


def saveMissionAndPtsPickl(listDic, fileName, pathName):
    pathName = pathName/fileName
    fh = open(pathName, 'wb')  # In binary format
    pickler = pickle.Pickler(fh, pickle.HIGHEST_PROTOCOL)
    for n in range(len(listDic)):
        pickler.dump(listDic[n])
    fh.close()
    print(Style.CYAN + "INFO : ------ Write Mission and List of mission points. Saved successfully in %s"%fileName + Style.RESET)


def readMissionAndPtsPickl(fileName):
    endFile = False
    fh = open(fileName, 'rb')
    unpickler = pickle.Unpickler(fh)
    # mission dictionary
    dicplanVol = unpickler.load()
    # list of points and point dictionary
    n = 0
    listDicPts, listPtPy = [], []
    while endFile is not True:
        try:
            dictPt = unpickler.load()
            listDicPts.append(dictPt)
            listPtPy.append(IRcl.ShootPoint())
            IRcl.ShootPoint.loadDicPoint2Point(listPtPy[n], dictPt)
            n = n + 1
        except:
            endFile = True
    fh.close()
    txt = '------ Read flight plan and listPts.   Read successfully from ' + fileName
    print(Style.CYAN + txt + Style.RESET)
    printPlanVol(dicplanVol)

    return dicplanVol, listDicPts, listPtPy


def writeSummaryFlightExcel(flightPlanSynthesis, pathName):
    """
     Write the Flight Plan Summary in Excel file
    If the file not exist the file FlightSummary.xlsx  is create withe one sheet (sheetname = Summary)

    :param flightPlanSynthesis:
    :param pathName:
    :return: .
    """
    pathName = pathName/'FlightSummary.xlsx'

    if pathName.is_file():
        print(Style.CYAN + 'INFO : ------ Write file FlightSummary.xlsx' + Style.RESET)
        pass
    else:
        print(Style.CYAN + 'INFO : ------ Create file FlightSummary.xlsx' + Style.RESET)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        wb.save(pathName)

    workbook = openpyxl.load_workbook(pathName)
    listSheets = workbook.sheetnames
    if listSheets[0] != 'Summary':
        print(Style.YELLOW + 'INFO :  Create sheet Summary' + Style.RESET)
        ws_sum = workbook.create_sheet("Summary", 0)
        ws_sum.title = "Summary"
        ws_sum.protection.sheet = False
    else:
        pass

    sheet = workbook['Summary']
    sheet.protection.sheet = False
    #sheet.delete_rows(2, 1000)
    for i in range(1000):
        for j in range(37):
            sheet.cell(row=i+2, column=j+1, value="")
    listHeadCol = headColumnSummaryFlight()
    for k in range(len(listHeadCol)):
        sheet.cell(1, k + 1).value = listHeadCol[k]
    for i in range(len(flightPlanSynthesis)):
        for k in range(len(flightPlanSynthesis[i])):
            sheet.cell(i + 2, k + 1).value = flightPlanSynthesis[i][k]
    sheet.protection.sheet = False
    workbook.save(pathName)
    workbook.close()
    print(Style.GREEN + ' %s  successfully saved.' % pathName + Style.RESET)


def readFlightSummary(dir_mission, mute=None, no_pandas=False):
    """
        Read the FlightSummary  in Excel file.

    :param dir_mission: path of the folder that contains the mission data  type= string
    :param mute: if True displays information.                                type= bool
    :return: listSummaryFlight  data list for each pair of VIS-NIR images (timeDeviation, altitude, etc.)
    """
    summaryPath = osp.join(dir_mission, "FlightSummary.xlsx")

    txt = 'Read data from  ' + summaryPath
    print(Style.CYAN + txt + Style.RESET)
    if not no_pandas:
        df = pd.read_excel(summaryPath)
        return df.values.tolist()
    workbook = openpyxl.load_workbook(summaryPath, read_only=True, data_only=True)
    sheet = workbook['Summary']
    listSummaryFlight = []
    nulg = 2  # first line of data
    while sheet.cell(nulg, 1).value != None:
        print("new line")
        num_pt = int(sheet.cell(nulg, 1).value)  # N° point
        imgVisName = str(sheet.cell(nulg, 2).value)  # name of image VIS
        imgNirName = str(sheet.cell(nulg, 3).value)  # name  of image NIR
        timeDeviation = float(sheet.cell(nulg, 4).value)  # timeDeviation between  image NIR et VIS (en s)
        dateVisString = str(sheet.cell(nulg, 5).value)  # date image VIS (rectifiee si time-lapse)  format string
        dateNirString = str(sheet.cell(nulg, 6).value)  # date image NIR (rectifiee si time-lapse)  format string
        timeLine = float(sheet.cell(nulg, 7).value)  # date image VIS mesurée en secondes  sur la time line
        date_string = str(sheet.cell(nulg, 8).value)  # date image VIS (rectifiee si time-lapse)  format string
        gpsLat = float(sheet.cell(nulg, 9).value)  # latitude  dd.dddddd°
        gpsLong = float(sheet.cell(nulg, 10).value)  # longitude dd.dddddd°
        altiDrone2Sol = float(sheet.cell(nulg, 11).value)  # altitude du drone par rapport au sol (en m)
        altGeo = float(sheet.cell(nulg, 12).value)  # altitude du sol (en m)
        altiDrone2TakeOff = float(sheet.cell(nulg, 13).value)  # altitude du drone par rapport au Take-Off (en m)
        distNextPt = float(sheet.cell(nulg, 14).value)  # distance to last point (en m)
        capNextPt = float(sheet.cell(nulg, 15).value)  # cap to last point (en °)
        dist2StartPoint = float(sheet.cell(nulg, 16).value)  # distance cumulée
        xUTM = float(sheet.cell(nulg, 17).value)  # coordonnee UTM x  (East)
        yUTM = float(sheet.cell(nulg, 18).value)  # coordonnee UTM y  (North)
        zoneUTM = int(sheet.cell(nulg, 19).value)  # UTM zone
        yawDrone = float(sheet.cell(nulg, 20).value)  # Yaw drone     (roll NIR camera)
        pitchDrone = float(sheet.cell(nulg, 21).value)  # Pitch drone   (pitch NIR camera)
        rollDrone = float(sheet.cell(nulg, 22).value)  # Roll drone    (yaw NIR camera)
        yawGimbal = float(sheet.cell(nulg, 23).value)  # Yaw gimbal    (roll VIS camera)
        pitchGimbal = float(sheet.cell(nulg, 24).value)  # Pitch gimbal  (pitch VIS camera)
        rollGimbal = float(sheet.cell(nulg, 25).value)  # Roll gimbal   (yaw VIS camera)
        x_1 = float(sheet.cell(nulg, 26).value)  # drone motion   drone axis principal
        x_2 = float(sheet.cell(nulg, 27).value)  # drone motion
        x_3 = float(sheet.cell(nulg, 28).value)  # drone motion  # drone motion   Z axis
        yawIr2Vi = float(sheet.cell(nulg, 29).value)  # theoretical yaw coarse process
        pitchIr2Vi = float(sheet.cell(nulg, 30).value)  # theoretical pitch coarse process
        rollIr2Vi = float(sheet.cell(nulg, 31).value)  # theoretical roll coarse process
        bestSynchro = int(sheet.cell(nulg, 32).value)  # 1 if dt-synchro < timelapseDJI/4 (25%)
        bestMapping = int(sheet.cell(nulg, 33).value)  # 1 if image selected for mapping
        bestAlignment = int(sheet.cell(nulg, 34).value)  # 1 if image selected for alignment
        yawCoarseAlign = float(sheet.cell(nulg, 35).value)  # theoretical yaw coarse process
        pitchCoarseAlign = float(sheet.cell(nulg, 36).value)  # theoretical pitch coarse process
        rollCoarseAlign = float(sheet.cell(nulg, 37).value)  # theoretical roll coarse process

        data = num_pt, imgVisName, imgNirName, timeDeviation, dateVisString, dateNirString, timeLine, date_string, \
               gpsLat, gpsLong, altiDrone2Sol, altGeo, altiDrone2TakeOff, distNextPt, capNextPt, dist2StartPoint, \
               xUTM, yUTM, zoneUTM, yawDrone, pitchDrone, rollDrone, yawGimbal, pitchGimbal, rollGimbal, \
               x_1, x_2, x_3, yawIr2Vi, pitchIr2Vi, rollIr2Vi, bestSynchro, bestMapping, bestAlignment, \
               yawCoarseAlign, pitchCoarseAlign, rollCoarseAlign

        listSummaryFlight.append(data)
        nulg = nulg + 1
    workbook.close()

    if not mute: print(listSummaryFlight)

    return listSummaryFlight


def headColumnSummaryFlight():
    listHeadCol = ['Point',
                   'Image                       Visible                 (original)',
                   'Image                       Infrared                (original)',
                   'Time deviation                       [s]',
                   'Date of shooting. Vis                [s]',
                   'Date of shooting. Nir                [s]',
                   'Time  Line                           [s]',
                   'Date of shooting                     [s]',
                   'Latitude                dd°ddddd',
                   'Longitude               dd°ddddd',
                   'Drone Altitude     / ground          [m]',
                   'Ground Elevation   / sea level       [m]',
                   'Drone Altitude     / Take off        [m]',
                   'Distance to next point               [m]',
                   'Cape to next point                   [°]',
                   'Cumulative distance from the starting point.   [m]',
                   'x UTM                                [m]',
                   'y UTM                                [m]',
                   'Zone UTM',
                   'Yaw drone                            [°]',
                   'Pitch drone                          [°]',
                   'Roll drone                           [°]',
                   'Yaw Gimbal                           [°]',
                   'Pitch Gimbal                         [°]',
                   'Roll Gimbal                          [°]',
                   'x_1                                  [m]',
                   'x_2                                  [m]',
                   'x_3                                  [m]',
                   'Yaw Theoretical alignment        IR to VI                     [°]',
                   'Pitch Theoretical alignment      IR to VI                     [°]',
                   'Roll Theoretical alignment       IR to VI                     [°]',
                   'Best synchro',
                   'Mapping ODM',
                   'Alignment',
                   'Yaw Coarse alignment             IR to VI                     [°]',
                   'Pitch Coarse alignment           IR to VI                     [°]',
                   'Roll Coarse alignment            IR to VI                     [°]'
                   ]
    return listHeadCol


# -------------------------           affichage écran       -----------------------------------------------------------

def printPlanVol(planVol):
    print(' Flight at : %s  (%s)  du %s   '
          '\n Client : %s'
          '\n Pilot : %s sur drone %s'
          '\n Camera Vi : %s'
          '\n Timelapse  : %s  s    DeltaTime %s  s'
          '\n Camera IR: %s'
          '\n Timelapse  : %s  s    DeltaTime %s  s'
          '\n Infrared filter  IR %i nm'
          '\n Whether : %s  Vent %.1f m/s  Temperature %.1f °C' %
          (planVol['mission']['lieu'],
           (planVol['mission']['coord GPS Take Off'] + '  ' + str(planVol['mission']['altitude Take 0ff']) + ' m'),
           planVol['mission']['date'],
           planVol['mission']['client'],
           planVol['mission']['pilote'],
           (planVol['drone']['marque'] + ' ' + planVol['drone']['type']),
           planVol['drone']['marque'],
           planVol['drone']['timelapse'], planVol['drone']['deltatime'],
           (planVol['cameraIR']['marque'] + '  ' + planVol['cameraIR']['type']),
           planVol['cameraIR']['timelapse'], planVol['cameraIR']['deltatime'],
           planVol['images']['filtreIR'],
           planVol['meteo']['ensoleillement'], float(planVol['meteo']['vent']), float(planVol['meteo']['temperature'])
           )
          )
    return


# ------------------     Gestion des fichiers      ---------------------

def reformatDirectory(di, rootdir=None, makeOutdir=False):
    if di is None:
        return rootdir
    if os.path.exists(di):
        return di
    else:
        if rootdir is not None:
            newdi = os.path.join(rootdir, di)
            return reformatDirectory(newdi, rootdir=None, makeOutdir=makeOutdir)
        if makeOutdir:
            os.mkdir(di)
            print("creating output dir: %s" % di)
            return di

    raise NameError("Cannot find directory %s" % di)


def loadFileGUI(mute=True, **options):
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename(**options)  # show an "Open" dialog box and return the path to the selected file
    if not mute: print(filename)
    return filename


def loadFolderGUI(mute=True, **options):
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
    filename = askdirectory(**options)
    if not mute: print(filename)
    return filename


def answerYesNo(txt) -> object:
    ans = input(txt)
    tryAgain = True
    countTry = 0
    while tryAgain:
        try:
            if ans.lower() in ["y", "yes", "1"]:
                return True
            elif ans.lower() in ["n", "no", "0"]:
                return False
            raise ValueError

        except ValueError:
            try:
                if ans == "":
                    return
                if countTry < 2:
                    print('You need to type 0 or 1')
                    ans = input(txt)
                else:
                    return
            except:
                return ans
            else:
                countTry += 1
                tryAgain = True


def check_numeric(x):
    if not isinstance(x, (int, float, complex)):
        raise ValueError('{0} is not numeric'.format(x))


# ----------------------   A priori calculation of the pitch, yaw & roll "coarse".  -----------------------------------

def motion_in_DroneAxis(listPts, mute=True):
    """
    Geographical axes:
        e_EW vector  West > East          |e_EW|=1
        e_SN vector  South > North        |e_SN|=1     e_EW . e_SN = 0  ; warning Clockwise  (e_z downwards!)
    Axe orientation  e_EW <=> West > East ,   e_SN <=> South > North
                    nord > 0°  ;  east > 90°  ;  south > 180°  ;  west > 270°

    Axes of the drone:
        e_1  vector normal to the axis of the drone       |e_1|=1
        e_2  vector of the axis of the drone (forward)    |e_2|=1  ;  e_1 . e_2 =0
        e_3 = e_1 x e_2 .        Counterclockwise ( e_3 upwards ).

    x_1  distance travelled along the e_1 axis
    x_2  distance travelled along the drone axis
    x_3  distance travelled along the vertical axis.
    """
    x_WE, y_SN = motionDrone_in_GeographicAxis(listPts, mute=mute)  # motion in geographic axis
    for i in range(len(listPts)):
        sin_Yaw = np.sin(np.deg2rad(listPts[i].yawDrone))
        cos_Yaw = np.cos(np.deg2rad(listPts[i].yawDrone))
        listPts[i].x_1 = -x_WE[i] * cos_Yaw + y_SN[i] * sin_Yaw
        listPts[i].x_2 = x_WE[i] * sin_Yaw + y_SN[i] * cos_Yaw

    motionDroneZaxis(listPts)

    return


def motionDrone_in_GeographicAxis(listPt, mute=True):
    """
        vector   D = x_EW e_EW + y_SN e_SN       |e_EW|=1, |e_SN|=1, e_EW.e_SN=0
        Axe orientation  e_EW <=> West > East ,   e_SN <=> South > North

            N  e_SN
              |
        W ----E ----> E  e_WE
              |
              S
        """
    x_WE, y_SN = [], []
    for i in range(len(listPt)):
        if i >= len(listPt) - 1:
            distWE = float(listPt[-1].gpsUTM_X) - float(listPt[-2].gpsUTM_X)
            distSN = float(listPt[-1].gpsUTM_Y) - float(listPt[-2].gpsUTM_Y)
        else:
            distWE = float(listPt[i + 1].gpsUTM_X) - float(listPt[i].gpsUTM_X)
            distSN = float(listPt[i + 1].gpsUTM_Y) - float(listPt[i].gpsUTM_Y)

        x_WE.append(distWE)
        y_SN.append(distSN)
        if not mute:
            print('point N° ', i, '   x_WE = ', distWE, ' m    y_SN = ', distSN, ' m    distance ',
                  (distWE ** 2 + distSN ** 2) ** 0.5)

    return x_WE, y_SN


def motionDroneZaxis(listPts):
    listPts[-1].x_3 = 0.
    for i in range(0, len(listPts) - 1):
        listPts[i].x_3 = (listPts[i + 1].altGround - listPts[i].altGround) + (listPts[i + 1].altGeo - listPts[i].altGeo)


def theoreticalIrToVi(listPts, timelapse_Vis, offset=None):
    #   theoretical  Yaw
    angle = [listPts[n].rollDrone for n in range(len(listPts))]
    x = [listPts[n].x_1 for n in range(len(listPts))]
    theoreticalYaw = theoreticalAngleDeviation(listPts, angle, x, timelapse_Vis, axe=1)
    #   theoretical  Pitch
    angle = [listPts[n].pitchDrone for n in range(len(listPts))]
    x = [listPts[n].x_2 for n in range(len(listPts))]
    theoreticalPitch = theoreticalAngleDeviation(listPts, angle, x, timelapse_Vis, axe=2)
    #   theoretical  Roll   It is assumed that the "gimbal-lock" mode is used.  In other words gimbal_Yaw=drone_Yaw
    # The prediction of the roll image with the gimbal yaw is totally false when the "gimbal-lock" mode is used ... force 0
    #theoreticalRoll = [0. for _ in range(len(listPts))]
    theoreticalRoll = rollDeviation(listPts, timelapse_Vis)


    for i in range(len(listPts)):
        try:
            theoreticalYaw[i] = theoreticalYaw[i] + offset[0]
            theoreticalPitch[i] = theoreticalPitch[i] + offset[1]
            theoreticalRoll[i] = theoreticalRoll[i] + offset[2]
        except:
            pass
        listPts[i].yawIR2VI = theoreticalYaw[i]
        listPts[i].pitchIR2VI = theoreticalPitch[i]
        listPts[i].rollIR2VI = theoreticalRoll[i]

    return listPts, theoreticalPitch, theoreticalYaw, theoreticalRoll


def theoreticalAngleDeviation(listPts, angle, x, timelapse_Vis, axe=0):
    """
    u  composante du déplacement    x = dist . e_idx
    idx=1    Yaw    (le roll du drone correspond au yaw de la caméra NIR!)
    idx=2    Pitch  (attention offset de 90° pour le DJI)
    e_2  vecteur de l'axe du drone (vers l'avant)    |e_2|=1
    e_1  vecteur normal à l'axe du drone    |e_1|=1  ;  e_1 . e_2 =0 ; repère direct

    The distance between the lenses of two cameras (DJI Mavic Air 2 and SJCam M20) is CNIRCVIS_0 = 46 mm.
    This distance is not negligible if the drone is at very low altitude.
    For example during the synchronization step the drone is 2 m above the ground.
    This distance must be added to the projection of the base line on the axis of the drone (Axis 2) for the
    calculation of the pitch.
    """
    theoreticalAngle = []
    for i in range(len(listPts)):
        alpha, dt = interpolParabolicAngle(listPts, angle, i, timelapse_Vis)
        Cvi_t_Cvi_tk = interpolationCameraCenterVis(x, i, dt, timelapse_Vis)  # Algebraic !!!
        H = listPts[i].altGround
        if axe == 1:
            thetaVis = listPts[i].rollGimbal  # Roll Gimbal <=>  Yaw Camera VIS
            baseline = Cvi_t_Cvi_tk
        else:
            thetaVis = listPts[i].pitchGimbal + 90.  # Pitch Gimbal <=> Pitch Camera VIS
            baseline = Cvi_t_Cvi_tk + cf.CNIRCVIS_0

        anglePhi = np.rad2deg(np.arctan(baseline / H + np.tan(np.deg2rad(thetaVis))))
        anglePsi = anglePhi - alpha
        theoreticalAngle.append(anglePsi)

    return theoreticalAngle


def rollDeviation(listPts, timelapse_Vis):
    """
    yaw drone  <=> roll NIR camera
    yaw gimbal <=> roll VIS camera     is totally wrong if DJI "gimbal-lock" mode is used!

    In "Gimbal lock" mode the information provided by the DJI drone about the gimbal’s yaw angle is not reliable.
    We make the hypothesis (reasonable) that the gimbal is aligned on the axis of the drone.
    The roll angle of the near infrared image (NIR) is obtained by interpolation of the yaw angle of the drone.
    The "roll image" between the visible (VIS) and near infrared (NIR) images is obtained by difference between
    the yaw angle of the drone (at the moment of the VIS shooting) and the roll angle of the NIR image.
    Added to this is the roll offset caused by the misalignment of the infrared camera in relation to the drone’s longitudinal axis.

    """
    yaw_tVis = [listPts[n].yawDrone for n in range(len(listPts))]
    theoreticalRoll = []
    for i in range(len(listPts)):
        yaw_tNir, dt = interpolParabolicAngle(listPts, yaw_tVis, i, timelapse_Vis)
        rollNir2Vis = yaw_tVis[i] - yaw_tNir
        theoreticalRoll.append(rollNir2Vis)

    return theoreticalRoll


def interpolParabolicAngle(listPts, angle, i, timelapse_Vis):
    """
        Parabolic interpolation of the drone angle at the moment (t) the NIR image was taken.
        dt = t0-t with t0 : date camera VIS and  t : date camera NIR.
        If k is index of t0 = tk then t_1 = tk-1 ; t1 = tk+1 etc.
        t_1 <= t <= t1
        "Forward" interpolation if dt<0.
    """
    dt = listPts[i].timeDeviation
    try:
        t_2 = listPts[i - 2].timeLine
        t_1 = listPts[i - 1].timeLine
        t0 = listPts[i].timeLine
        t1 = listPts[i + 1].timeLine
        t2 = listPts[i + 2].timeLine
        t = t0 - dt
        if t_1 <= t <= t1:
            alpha = Parabolic(listPts, angle, i, t)
        elif t_2 <= t <= t_1:
            alpha = Parabolic(listPts, angle, i-1, t)
        elif t1 <= t <= t2:
            alpha = Parabolic(listPts, angle, i+1, t)
        else:
            alpha, dt = interpolLinearAngle(listPts, angle, i, timelapse_Vis)

    except:
        alpha, dt = interpolLinearAngle(listPts, angle, i, timelapse_Vis)

    return alpha, dt


def Parabolic(listPts, angle, k, t):
    t_1 = listPts[k - 1].timeLine
    t0 = listPts[k].timeLine
    t1 = listPts[k + 1].timeLine
    d_1 = (angle[k - 1] - angle[k]) / (t_1 - t0)
    d1 = (angle[k + 1] - angle[k]) / (t1 - t0)
    a = (d_1 - d1) / (t_1 - t1)
    b = d_1 - a * (t_1 + t0)
    c = angle[k] - a * t0**2 - b * t0
    alpha = a * t**2 + b * t + c
    return alpha


def interpolLinearAngle(listPts, angle, i, timelapse_Vis):
    """
    Linear interpolation of the drone angle at the moment the NIR image was taken.
    dt = tk-t with tk : date camera VIS and  t : date camera NIR.
    "Forward" interpolation if dt<0.
    """
    dt = listPts[i].timeDeviation
    if i == 1 or i == len(listPts) - 1:
        alpha = angle[i]
    elif dt < 0:
        alpha = (angle[i] * (dt / timelapse_Vis + 1) - angle[i + 1] * dt / timelapse_Vis)
    else:
        alpha = (angle[i - 1] * dt / timelapse_Vis - angle[i] * (dt / timelapse_Vis - 1))
    return alpha, dt


def interpolationCameraCenterVis(x, k, dt, timelapse_Vis):
    """
    DeltaCvis =(Cvis_t Cvis_tk) is the distance ( algebraic) between the center Cvis_t of the Vis  camera at  time t
    and the center Cvis_tk of the Vis camera at time tk. Linear interpolation of the base line is used..
    A simple rule of three is used since x is directly the distance travelled during the period of the VIS timelapse.
    TimeDeviation :   dt = (tk - t)  .
                  > tk Date of the image  VIS
                  > t  Date of nearest NIR image.
    The dates are measured on the time-line (synchronized clocks).
    x_k is the distance (Cvis(k+1)  Cvis(k)). This  distance is > 0  if drone flight forward and < 0 if backward.

    "Forward" interpolation if dt<0.

    Taking into account the possibility of a missing NIR image in the time lapse series.
    (The phenomenon is related to a recording fault on the SD card of the SJCam M20 camera.)

    """
    try:
        if dt < 0:
            if abs(dt) > timelapse_Vis:
                DeltaCvis = (x[k + 1] - x[k]) + x[k + 1] * (dt / timelapse_Vis)
            else:
                DeltaCvis = x[k] * (dt / timelapse_Vis)
        else:
            if dt > timelapse_Vis:
                DeltaCvis = (x[k - 1] - x[k - 2]) + x[k - 2] * (dt / timelapse_Vis)
            else:
                DeltaCvis = x[k - 1] * (dt / timelapse_Vis)
    except:
        DeltaCvis = 0

    return DeltaCvis


def average_Timelapse(tStart, tEnd, nbImageMission, mute=True):
    """
    Real period of the timelapse of the VIS camera.
    # Note: for the DJI mavic Air 2 with a theoretical timelapse of 2s it takes more than 55 frames to notice a jump
        of one second in the recording dates (real timelapse ~ 2,018s)
    The calculation is carried out over the total duration of the mission.
    The start date of the mission is defined in the flight plan (date of the first image in the visible spectrum).
    Note: listPts[0].dateVis is the date of the first image of the flight AFTER the synchronization phase.
    """
    total_FlightTime = (tEnd - tStart).total_seconds()
    av_Timelapse = round(total_FlightTime / (nbImageMission - 1), 6)
    microsec = float(str(av_Timelapse)) - int(str(av_Timelapse).split('.')[0])
    av_Timelapse_timedelta = timedelta(seconds=int(str(av_Timelapse).split('.')[0]),
                                       microseconds=int(round(microsec, 6) * 10 ** 6))
    if not mute:
        print(Style.GREEN + 'Visible image acquisition interval : %s  s '%(str(np.round(av_Timelapse, 6))) + Style.RESET)
    return av_Timelapse, av_Timelapse_timedelta


def selectAllImages(listImgMatch, listPts, planVol, ratioSynchro=0.25):
    """
        Selects all pairs of images for alignment process.
        Tag alignment and best synchro (if ok) in summary-flight file
    """
    print(Style.CYAN + 'INFO : ------ Selecting all available image pairs in the AerialPhotography folder.' + Style.GREEN)
    ImgMatch, Pts = [], []
    DTmin = planVol['drone']['timelapse'] * ratioSynchro
    for pointImage in listPts:
        listPts[pointImage.num - 1].alignment = 1
        Pts.append(listPts[pointImage.num - 1])
        ImgMatch.append(listImgMatch[pointImage.num - 1])
        if -DTmin <= pointImage.timeDeviation <= DTmin:
            listPts[pointImage.num - 1].bestSynchro = 1
    return ImgMatch, Pts


def selectBestSynchro(listImgMatch, listPts, planVol, ratioSynchro=0.25):
    """
    Selects pairs of images with a timing difference of less than :
        DTmin = drone timelapse * ratioSynchro.
    'ratioSynchro' is a percentage of drone timelapse.  DTmin in [0.25 , 1.0]
    """
    print(Style.CYAN + '------ Selecting the best synchronized image pairs for alignment.  ' + Style.GREEN)
    ImgMatchSync, PtsSync = [], []
    DTmin = planVol['drone']['timelapse'] * ratioSynchro
    for pointImage in listPts:
        if -DTmin <= pointImage.timeDeviation <= DTmin:
            listPts[pointImage.num-1].bestSynchro = 1
            listPts[pointImage.num - 1].alignment = 1
            pointImage.alignment = 1
            ImgMatchSync.append(listImgMatch[pointImage.num-1])
            PtsSync.append(pointImage)
        else:
            listPts[pointImage.num - 1].bestSynchro = 0
            listPts[pointImage.num - 1].alignment = 0

    return ImgMatchSync, PtsSync


def selectBestOffset(listImg, listPts, planVol, timeOffset=0.025, mute=True):
    """
        Selects pairs pairs of images with a timing deviation of less than 0.05s.
        Used to calculate offset angles.
            DTmin = drone timelapse * ratioOffset.
        'ratioOffset' is a percentage of drone timelapse.  DTmin in [0.05 , 0.25]
    """
    ImgMatch, Pts, numPts= [], [], []
    for pointImage in listPts:
        if abs(pointImage.timeDeviation) <= timeOffset:
            listPts[pointImage.num-1].bestOffset = 1
            listPts[pointImage.num - 1].bestSynchro = 1
            listPts[pointImage.num - 1].alignment = 1
            pointImage.alignment = 1
            ImgMatch.append(listImg[pointImage.num-1])
            Pts.append(pointImage)
            numPts.append(pointImage.num - 1)
        else:
            listPts[pointImage.num - 1].bestOffset = 0
            listPts[pointImage.num - 1].bestSynchro = 0
            listPts[pointImage.num - 1].alignment = 0
    if len(Pts) < 4:
        #print('len(Pts)  ', len(Pts), '   timeOffset  ', timeOffset, '>>>>' , 1.05*timeOffset)
        timeOffset = 1.05 * timeOffset
        ImgMatch, Pts = selectBestOffset(listImg, listPts, planVol, timeOffset= timeOffset)
    else:
        print(Style.GREEN + '%s images were found for offsets calculation. Timing deviation < %.3f s' % (len(Pts), timeOffset) + Style.RESET)
        if not mute:
            for n in range(len(Pts)):
                print(listPts[numPts[n]])

    return ImgMatch, Pts

def estimOffset(ptsOffset):
    """
    Utilisation du critère basique basé sur la moyenne.
    Attention faible taille de l'échantillon et écart type parfois important
    :param ptsOffset:
    :return offsetYaw, offsetPitch, offsetRoll:
    """
    sumYaw, sumPitch = 0, 0
    for ptOffset in ptsOffset:
        sumYaw += ptOffset.rollDrone + ptOffset.yawCoarseAlign
        sumPitch += ptOffset.pitchDrone + ptOffset.pitchCoarseAlign
    offsetYaw = sumYaw / len(ptsOffset)
    offsetPitch = sumPitch / len(ptsOffset)
    offsetRoll = 0. # @TODO:  USE ESTIMATED COARSE ROLL WHEN AVAILABLE!
    return offsetYaw, offsetPitch, offsetRoll


def selectBestMapping(listImgMatch, listPts, planVol,  overlap_x=0.30, overlap_y=0.75, ratioSynchro=0.25):
    """
    overlap_x = 0.30  # percentage of overlap between two images  axe e_1
    overlap_y = 0.75  # percentage of overlap between two images  axe e_2   [50% , 75%]
    lCapt_x = width_capteur     image size VIS  axe e_1                pixels
    lCapt_y = height_capteur    image size VIS  axe e_2                pixels
    f = focalPix                focal length camera VIS                pixels
    """
    print(Style.CYAN + 'INFO : ------ Selection of mapping image pairs for the alignment process.  ' + Style.GREEN)
    ImgMatch, Pts = [], []
    camera_make, camera_type, width_capteur, height_capteur, focal_factor, focalPix = lectureCameraIrdrone()
    # TODO  prendre en compte la trajectoire exacte et pas seulement le mouvement  suivant e_2
    lCapt_x = width_capteur
    lCapt_y = height_capteur
    f = focalPix
    DTmin = planVol['drone']['timelapse'] * ratioSynchro
    # ------------------------------------------------------------------------------------------
    d = 0  # raz odometre
    firstImg = False
    for pointImage in listPts:
        d_y = pointImage.altGround * (1 - overlap_y) * lCapt_y / f
        if -DTmin <= pointImage.timeDeviation <= DTmin:
            pointImage.bestSynchro = 1
            listPts[pointImage.num - 1].bestSynchro = 1
            if d == 0 and firstImg is False:  # The first point for mapping has been found.
                listPts[pointImage.num - 1].alignment = 1
                listPts[pointImage.num - 1].bestMapping = 1
                Pts.append(listPts[pointImage.num - 1])
                ImgMatch.append(listImgMatch[pointImage.num - 1])
                firstImg = True
            elif d >= d_y:  # The next point for mapping has been found.
                listPts[pointImage.num - 1].alignment = 1
                listPts[pointImage.num - 1].bestMapping = 1
                Pts.append(listPts[pointImage.num - 1])
                ImgMatch.append(listImgMatch[pointImage.num - 1])
                d = 0  # Reset the odometer.
        if firstImg:  # Increment of the odometer.
            d = d + (pointImage.x_1 ** 2 + pointImage.x_2 ** 2) ** 0.5

    return ImgMatch, Pts


def save_motion_camera(dirMission, shootingPts, planVol, listImgMatch, mute=True):
    timeLine, yawIR2VI, pitchIR2VI, rollIR2VI = [], [], [], []
    timeLinePostProcess, yawCoarse, pitchCoarse, rollCoarse = [], [], [], []

    # List of motion models of IRdrone images.  Ext .py
    result_dir = Path(dirMission) / "ImgOffset"
    ImgPostProcess = sorted(result_dir.glob("*.npy"))

    # -------------------------------------------------------
    k = 0
    for shootPt in shootingPts:
        timeLine.append(shootPt.timeLine)
        yawIR2VI.append(shootPt.yawIR2VI)
        pitchIR2VI.append(shootPt.pitchIR2VI)
        rollIR2VI.append(shootPt.rollIR2VI)
        if shootPt.alignment == 1:
            timeLinePostProcess.append(shootPt.timeLine)
            try:
                assert ImgPostProcess[k].is_file()
                #  yaw & pitch .  Coarse  registration angle of near infrared images.
                mouvement = np.load(ImgPostProcess[k], allow_pickle=True).item()
                yawCoarse.append(mouvement["yaw"])
                pitchCoarse.append(mouvement["pitch"])
                rollCoarse.append(mouvement["roll"])
                shootingPts[shootPt.num - 1].yawCoarseAlign = mouvement["yaw"]
                shootingPts[shootPt.num - 1].pitchCoarseAlign = mouvement["pitch"]
                shootingPts[shootPt.num - 1].rollCoarseAlign = mouvement["roll"]
                k += 1
            except Exception as exc:
                print(Style.RED + "erreur lors de la lectures des angles process \nError = {}"%exc + Style.RESET)
    SaveSummaryInExcelFormat(dirMission, True, shootingPts, listImgMatch, mute=mute)
    SaveSummaryInNpyFormat(dirMission, False, planVol, shootingPts)
    return

def lectureCameraIrdrone():
    odm_camera_conf = Path(__file__).parent / ".." / "thirdparty" / "odm_data" / "irdrone_multispectral.json"
    # print('le fichier de calibration de la caméra du dji est dans : ', odm_camera_conf)
    file = open(odm_camera_conf, 'r')
    dicCameraIRdrone = json.load(file)
    for inpkey in dicCameraIRdrone.keys():
        list = inpkey.split()
        camera_make = list[0]
        camera_type = list[1]
        width_capteur = int(list[2])
        height_capteur = int(list[3])
        #projection = list[4]
        #focal_factor =float(list[5])
    odm_camera_calibration = Path(__file__).parent / ".." / "calibration" / "DJI_RAW" / "calibration.json"
    file = open(odm_camera_calibration, 'r')
    dicCameraIRdroneCalibration = json.load(file)
    focalPix = round((dicCameraIRdroneCalibration['mtx'][0][0]+ dicCameraIRdroneCalibration['mtx'][1][1])/2 ,1)
    focal_factor = focalPix / width_capteur
    #focal2436 = float(focal_factor) * 36.


    return camera_make, camera_type, width_capteur, height_capteur, focal_factor, focalPix


def logoIRDrone(num=None):
    try:
        if num == None: pass
        elif num == 1: logoIRDrone_1()
        elif num == 2: logoIRDrone_2()
        elif num == 3: logoIRDrone_3()
        elif num == 4: logoIRDrone_4()
        else: pass
    except:
        pass
    return


def logoIRDrone_1():
    print('\n ')
    print(Style.CYAN +    '  III    IIIIIIIIII      IIIIIIIIIII')
    print(Style.BLUE +    '  III    III      III    III      IIII')
    print(Style.MAGENTA + '  III    III        II   III        III')
    print(Style.RED +     '  III    II        II    III         III   II IIIII      IIIIII    IIIIIII     IIIIII')
    print(Style.YELLOW +  '  III    IIIIIIIIIII     III         III   III     II   II    II   II     II  II     II')
    print(Style.GREEN +   '  III    III    III      III         III   III         II      II  II     II  IIIIIII')
    print(Style.WHITE +   '  III    III     III     III        III    III          II    II   II     II  II        ')
    print(Style.RESET +   '  III    III      III    IIIIIIIIIIII      III           IIIIII    II     II   IIIIIII ')
    print('  ')
    return


def logoIRDrone_2():
    print('\n ')
    print(Style.CYAN +    '  III    RRRRRRRRRR      DDDDDDDDDD ')
    print(Style.BLUE +    '  III    RRR      RRR    DDD      DDDD')
    print(Style.MAGENTA + '  III    RRR        RR   DDD        DDD')
    print(Style.RED +     '  III    RRR      RR     DDD          DD   RR RRRRR      OOOOOO     NNNNNNN     EEEEEE')
    print(Style.YELLOW +  '  III    RRRRRRRRRR      DDD          DD   RRR     RR   OO    OO   NN     NN  EE      E')
    print(Style.GREEN +   '  III    RRR    RRR      DDD         DDD   RRR         OO      OO  NN     NN  EEEEEEEE')
    print(Style.WHITE +   '  III    RRR     RRR     DDD       DDDD    RRR          OO    OO   NN     NN  EE        ')
    print(Style.RESET +   '  III    RRR      RRR    DDDDDDDDDDD       RRR           OOOOOO    NN     NN   EEEEEEE ')
    print('  ')
    return


def logoIRDrone_3():
    print(Style.CYAN + '\n ')
    print('  III    IIIIIIIIII      IIIIIIIIIII')
    print('  III    III      III    III      IIII')
    print('  III    III        II   III        III')
    print('  III    III       II    III         III   II IIIII      IIIIII    IIIIIII     IIIIII')
    print('  III    IIIIIIIIIII     III         III   III     II   II    II   II     II  II     II')
    print('  III    III    III      III         III   III         II      II  II     II  IIIIIII')
    print('  III    III     III     III        III    III          II    II   II     II  II        ')
    print('  III    III      III    IIIIIIIIIIII      III           IIIIII    II     II   IIIIIII ')
    print('  ')
    return


def logoIRDrone_4():
    print(Style.CYAN + '\n ')
    print(' _______   __________      ___________    ')
    print('   |I|     |R|      \R\     |D|      \D\  ' )
    print('   |I|     |R|       |R|    |D|       \D\ ' )
    print('   |I|     |R|      /R/     |D|        |D|   --- _----_    _------_      _------_       _-------_    ')
    print('   |I|     |R|_____/R/      |D|        |D|   |R|/      \  /O/     \O\   |N|     \I\    /E/     \E\   ')
    print('   |I|     |R|    \R\       |D|        |D|   |R|         |0|       |O|  |N|      |N|  |E|       |E|  ')
    print('   |I|     |R|     \R\      |D|        |D|   |R|         |0|       |O|  |N|      |N|  |E|______/E/   ')
    print('   |I|     |R|      \R\     |D|       /D/    |R|         |O|       |O|  |N|      |N|  |E|            ')
    print(' __|I|__  _|R|_     _\R\_  _|D|______/D/    _|R|_         \O\_____/O/  _|N|_     |N|_  \E\______/E/  ')
    print('')
    print('=====================================================================================================')
    return

if __name__ == "__main__":
    logoIRDrone(num=1)
    logoIRDrone(num=2)
    logoIRDrone(num=3)
    logoIRDrone(num=4)
    logoIRDrone(num=5)
    logoIRDrone(num=-1)
    logoIRDrone('color')
    logoIRDrone(0)
    logoIRDrone(num=0)

