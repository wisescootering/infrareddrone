# -*- coding: utf-8 -*-
# pylint: disable=C0103, C0301, W0703

"""
Created on 2021-10-05 19:17:16
version 1.05 2021-12-01
version 1.06 2021-12-31 16:41:05.   Theoritical Yaw,Pitch,Roll for NIR images

@authors: balthazar/alain
"""

import irdrone.process as pr
import irdrone.utils as ut
from irdrone.utils import Style
import utils_GPS as uGPS
import utils_IRdrone_Plot as IRdplt
import os
import os.path as osp
import sys
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from operator import itemgetter
import datetime
import openpyxl
from openpyxl import Workbook
import numpy as np
from datetime import timedelta


# -----   Convertisseurs de dates   Exif<->Python  Excel->Python    ------
def dateExcelString2Py(dateTimeOriginal):
    """
    :param dateTimeOriginal: date  d'une image en format string
    :return: datePhotoPython     (datetime python)
    """
    imgYear = int(dateTimeOriginal[0:4])
    imgMonth = int(dateTimeOriginal[5:7])
    imgDay = int(dateTimeOriginal[8:11])
    imgHour = int(dateTimeOriginal[11:13])
    imgMin = int(dateTimeOriginal[14:16])
    second = dateTimeOriginal[17:].__str__()
    if len(second) <= 2:
        imgSecond = int(second)
        imgMicroSecond = 0
    else:
        second, microsecond = second.split('.')
        imgSecond = int(second)
        imgMicroSecond = int(microsecond)
    # construction de la date python
    #  Rem: Pour accéder aux éléments individuels de la date python
    #       datePhotoPython.year(.month | .day | .hour | .minute | .second | .microseconde)
    #
    datePy = datetime.datetime(imgYear, imgMonth, imgDay, imgHour, imgMin, imgSecond, imgMicroSecond)

    return datePy


def dateExif2Py(exifTag):
    """
    :param exifTag: données Exif d'une image
    :return: datePhotoPython     (datetime python)

       Extraction de la date  de la prise de photo à partir de ses données Exif
       La date est fournie par  l'horloge de la caméra.
       Format Exif  YYYY:MM:DD  hh:mm:ss  type string
       Puis conversion en datetime python en analysant la chaine de caractère Exif
       On récupère Year Month Day Hour Min et Second  de type string
       Il faut convertir les string en integer (via la fonction int()  )

       Attention la clé Exif à utiliser est 'DateTimeOriginal'
       La clé Exif 'DateTime' indique la date de création du fichier sur l'ordinateur!
       La valeur de la clé 'DateTimeOriginal' est une chaine de caractères (string) 'YYYY:MM:DD  hh:mm:ss'
       Il faut extraire de cette chaine les valeurs numériques (integer) DD,MM,YYYY,hh,mm,ss
       puis construire une date "python" datetime.datetime(YYYY,MM,DD,hh,mm,ss,micross). Sous cette
       forme on peut manipuler les dates. En particulier on peut connaître  la durée entre deux
       dates  (en seconde). Il est aussi possible de connaître le nom du jour et du mois.

    """

    dateTimeOriginal = exifTag.get('DateTimeOriginal')  # date exacte de la prise de vue
    datePy = dateExcelString2Py(dateTimeOriginal)

    return datePy


def datePy2Exif(datePy):
    """
    :param datePy:    (Year, Month, Day, Hour, Minute, Second, Microsecond, timezone)
    :return: dateExif  str  format  YYYY:MM:DD hh:mm:ss
    """
    dateExif = str(datePy.year) + ':' + str(datePy.month) + ':' + str(datePy.day) + ' ' + str(datePy.hour) + ':' + \
               str(datePy.minute) + ':' + str(datePy.second)
    return dateExif


def dateExif2Excel(dateExif):
    """
    :param dateExif:   str format  YYYY:MM:DD hh:mm:ss
    :return: dateExcel str format  YYYY-MM-DD hh:mm:ss
    """
    date = dateExif.split(' ')[0].split(':')
    date.extend(dateExif.split(' ')[1].split(':'))
    dateExcel = str(date[0]) + '-' + str(date[1]) + '-' + str(date[2]) + ' ' + \
                str(date[3]) + ':' + str(date[4]) + ':' + str(date[5])

    return dateExcel


def dateExcel2Py(dateExcel):
    """
    :param dateExcel:    str    format  YYYY-MM-DD hh:mm:ss
    :return: datePy     (Year, Month, Day, Hour, Minute, Second, Microsecond, timezone)
    """
    datePy = dateExcelString2Py(dateExcel)
    return datePy


# ------------------     Plan de Vol      ------------------------------


def readFlightPlan(pathPlanVolExcel, mute=None):
    """
        Read the Flight Plan  in Excel file.

    :param pathPlanVolExcel: chemin du fichier Excel qui contient le plan de vol  (string)
           mute : affiche des informatios si True  (utile en phase debug)
    :return: planVol  Dictionnaire contenant les données du plan de vol  (dic)
    """
    workbook = openpyxl.load_workbook(pathPlanVolExcel, read_only=True, data_only=True)

    sheet = workbook['Plan_de_Vol']

    nuetude = 2  # 2      numéro première ligen de données du fichier Excel
    nudrone = (nuetude + 8) + 1  # 2+8+1 =11
    nucameraIR = (nudrone + 7) + 1  # 11+7+1=19
    nuimages = (nucameraIR + 5) + 1  # 19+5+1=25
    numeteo = (nuimages + 11) + 1  # 25+11+1=37
    planVol = {'mission':
                   {'client': sheet.cell(nuetude + 1, 2).value,
                    'lieu': sheet.cell(nuetude + 2, 2).value,
                    'coord GPS Take Off': sheet.cell(nuetude + 3, 2).value,
                    'altitude Take 0ff': sheet.cell(nuetude + 4, 2).value,
                    'date': sheet.cell(nuetude + 5, 2).value,  # ' DD-MM-YYYY  hh:mm:ss
                    'heure_solaire': sheet.cell(nuetude + 6, 2).value,
                    'numero_du_vol': sheet.cell(nuetude + 7, 2).value,
                    'pilote': sheet.cell(nuetude + 8, 2).value
                    },
               'drone':
                   {'marque': sheet.cell(nudrone + 1, 2).value,
                    'type': sheet.cell(nudrone + 2, 2).value,
                    'timelapse': sheet.cell(nudrone + 3, 2).value,
                    'deltatime': sheet.cell(nudrone + 4, 2).value,
                    'imatriculation': sheet.cell(nudrone + 5, 2).value,
                    'altitude  de vol': sheet.cell(nudrone + 6, 2).value,
                    'synchro horloges': sheet.cell(nudrone + 7, 2).value  # libre
                    },
               'cameraIR':
                   {'marque': sheet.cell(nucameraIR + 1, 2).value,
                    'type': sheet.cell(nucameraIR + 2, 2).value,
                    'timelapse': sheet.cell(nucameraIR + 3, 2).value,
                    'deltatime': sheet.cell(nucameraIR + 4, 2).value,
                    'synchro horloges': sheet.cell(nucameraIR + 5, 2).value  # libre
                    },
               'images':
                   {'repertoireViR  (save)': sheet.cell(nuimages + 1, 2).value,
                    'repertoireDrone': sheet.cell(nuimages + 2, 2).value,
                    'extDrone': sheet.cell(nuimages + 3, 2).value,
                    'filtreDrone': sheet.cell(nuimages + 4, 2).value,
                    'libre_1': sheet.cell(nuimages + 5, 2).value,  # libre
                    'libre_2': sheet.cell(nuimages + 6, 2).value,  # libre
                    'repertoireIR': sheet.cell(nuimages + 7, 2).value,
                    'extIR': sheet.cell(nuimages + 8, 2).value,
                    'filtreIR': sheet.cell(nuimages + 9, 2).value,
                    'libre_3': sheet.cell(nuimages + 10, 2).value,  # libre
                    'libre_4': sheet.cell(nuimages + 11, 2).value  # libre
                    },
               'meteo':
                   {'ensoleillement': sheet.cell(numeteo + 1, 2).value,
                    'vent': sheet.cell(numeteo + 2, 2).value,
                    'temperature': sheet.cell(numeteo + 3, 2).value,
                    'humidite': sheet.cell(numeteo + 4, 2).value
                    }
               }
    workbook.close()

    if not mute:
        printPlanVol(planVol)

    return planVol


def extractFlightPlan(dirPlanVol, mute=True):
    """
    Extract the datas of Fligth Plan
    > Chemin des images Visible et IR, type de drone et de caméra, synchroniation horloges ...
    > Images list of Drone and IR cameras

    :param dirPlanVol:  path of Fligth Plan in  a Excel
               mute : affiche des informatios si True  (utile en phase debug)
    :return: (planVol,imgListDrone,deltaTimeDrone,timeLapseDrone,imgListIR,deltaTimeIR,timeLapseIR,
    dirNameIRdrone,coordGPS_TakeOff,altiTakeOff)   )
    """
    planVol = readFlightPlan(dirPlanVol, mute=mute)
    dirNameIRdrone = planVol['images']['repertoireViR  (save)']  # folder for save photography  VIR,  NDVI (output)
    dirNameDrone = planVol['images']['repertoireDrone']  # Drone photography folder   (input)
    dirNameIR = planVol['images']['repertoireIR']  # IR photography folder (input)
    dateMission = planVol['mission']['date']  # date of flight > format DD MM et YYYY
    typeDrone = planVol['drone']['type']  # type of drone (see in the Exif tag of the image of the drone)
    extDrone = planVol['images']['extDrone']  # file format Vi
    typeIR = planVol['cameraIR']['type']  # type of camera in use (see in the Exif tag of the image of the IR camera)
    timeLapseDrone = float(planVol['drone']['timelapse'])  # Time Lapse of Drone camera
    timeLapseIR = float(planVol['cameraIR']['timelapse'])  # Time Lapse of IR camera
    extIR = planVol['images']['extIR']  # file format  IR
    deltaTimeDrone = float(planVol['drone']['deltatime'])  # decalage horloge caméra du drone / horloge de référence
    deltaTimeIR = float(planVol['cameraIR']['deltatime'])  # decalage horloge caméra infrarouge /horloge de référence

    #    Liste des images de l'étude.
    #    Une liste pour les images du drone et une liste pour les images de la caméra infrarouge
    #    Chaque élément de la liste est un triplet (file name image , path name image, date image)

    dirNameDrone = reformatDirectory(dirNameDrone, xlpath=dirPlanVol)
    dirNameIR = reformatDirectory(dirNameIR, xlpath=dirPlanVol, makeOutdir=True)
    dirNameIRdrone = reformatDirectory(dirNameIRdrone, xlpath=dirPlanVol, makeOutdir=True)

    imgListDrone = creatListImgVIS(dirNameDrone, dateMission, typeDrone, '*', extDrone, planVol)

    imgListIR = creatListImgNIR(dirNameIR, '*', extIR)

    if not mute:
        if len(imgListDrone) == 0:
            print('No visible images detected for this flight.')
            sys.exit(2)
        else:
            print('%i visible images detected for this flight.' % len(imgListDrone))
            print([(imgListDrone[i][0], imgListDrone[i][2]) for i in range(len(imgListDrone))])

        if len(imgListIR) == 0:
            print('No near infrared images detected for this flight.')
            sys.exit(2)
        else:
            print('%i near infrared images detected for this flight.' % len(imgListIR))
            print([(imgListIR[i][0], imgListIR[i][2]) for i in range(len(imgListIR))])

        if len(imgListDrone) == 0 and len(imgListIR) == 0:
            print('%i visible images (Vi) and %i near infrared images (NiR) for this flight.' %
                  (len(imgListDrone), len(imgListIR)))

    return planVol, imgListDrone, deltaTimeDrone, timeLapseDrone, imgListIR, deltaTimeIR, timeLapseIR, dirNameIRdrone


def creatListImgNIR(dirName, camera, typImg):
    """
    :return:  imgList   [(),...,(file name image , path name image, date image),...,()]
    """
    # Création de la liste des photos IR  (extension donnée  RAW) et contenues dans le répertoire dirName
    # Si camera="*' alors la liste comporte toutes les images avec l'extension typeImg (ici RAW)
    # Pour une  SJcam M20 on extrait la date du nom du fichier  ...
    # car les données Exif du fichier RAW de SJCam sont illisibles par exifread ou bien PIL.ExifTags ???
    #
    print(Style.CYAN + '------ Creating the list of near-infrared spectrum images' + Style.RESET)
    imlist = sorted(ut.imagepath(imgname="*%s*.%s" % (camera, typImg), dirname=dirName))
    imgList = []
    for i in range(len(imlist)):
        nameImg = imlist[i].split('\\')[len(imlist[i].split('\\')) - 1]
        imgYear = int(nameImg[0:4])
        imgMonth = int(nameImg[5:7])
        imgDay = int(nameImg[7:9])
        imgHour = int(nameImg[10:12])
        imgMin = int(nameImg[12:14])
        imgSecond = int(nameImg[14:16])
        imgMicroSecond = 0
        dateImg = datetime.datetime(imgYear, imgMonth, imgDay, imgHour, imgMin, imgSecond, imgMicroSecond)
        imgList.append((nameImg, imlist[i], dateImg))  # ajout à la liste des images

    # imgList = timelapseRectification(imgList)

    return imgList


def creatListImgVIS(dirName, dateMission, cameraModel, camera, typImg, planVol, debug=False):
    """
    :param dirName:
    :param dateMission:
    :param cameraModel:
    :param camera:
    :param typImg:
    :param debug:
    :return:  imgList   [(), ...,(file name image , path name image, date image), ..., ()]
    """

    # Création de la liste des photos (extension donnée  ex: JPG, DNG) et contenues dans le répertoire dirName
    # Si camera="*' alors la liste comporte toutes les images avec l'extension typeImg
    # Si camera ='DJI'  la liste filtre les images prises par le drone
    # Pas possible de filtrer les images SJcam avec le non du fichier  (il faut regarder les données Exif de l'image)

    print(Style.CYAN + '------ Creating the list of visible spectrum images' + Style.RESET)
    imlist = sorted(ut.imagepath(imgname="*%s*.%s" % (camera, typImg), dirname=dirName))

    imgList = []
    j = 0
    for i in range(len(imlist)):
        img = pr.Image(imlist[i])
        img.camera["timelapse"] = float(planVol['drone']['timelapse'])
        img.camera["deltatime"] = float(planVol['drone']['deltatime'])
        try:
            """
            extraction des données Exif de l'image
            si pas de données Exif image ignorée.
            """
            debug = True
            cameraModelImg = img.camera['model']

            if cameraModelImg == cameraModel:  # images prises par d'autres caméras. images ignorées
                dateImg = img.date
                if (dateImg.year, dateImg.month, dateImg.day) == (dateMission.year, dateMission.month, dateMission.day):
                    j += 1
                    nameImg = imlist[i].split('\\')[len(imlist[i].split('\\')) - 1]
                    imgList.append((nameImg, imlist[i], dateImg))  # ajout à la liste des images
                else:
                    if debug: print(Style.YELLOW,
                                    '%s a été prise le %i %i %i. Cette date est différente de celle de la mission %i %i %i'
                                    % (imlist[i], dateImg.day, dateImg.month, dateImg.year, dateMission.day,
                                       dateMission.month,
                                       dateMission.year), Style.RESET)
            else:
                if debug: print(Style.YELLOW,
                                '%s a été prise par un autre  appareil (Model %s) ' % (imlist[i], cameraModelImg),
                                Style.RESET)
        except:
            if debug: print("No Exif tags in %s" % imlist[i])

    if float(planVol['drone']['timelapse']) > 0:
        # Les dates ne sont rectifiées que si les images ont été prises en hyperlapse.
        # Pour des images en single shoot la rectification n'aurait pas de sens.
        imgList = timelapseRectification(imgList)

    return imgList


def timelapseRectification(imgList):
    """
    :param imgList:         [(), ...,(file name image , path name image, original date image), ..., ()]
    :return:  new_imgList   [(), ...,(file name image , path name image, rectified date image ), ..., ()]

    Rectification des dates du timelapse
    La date d'enregistrment est arrondie à la seconde près dans l'Exif et la valeur du pas du timlapse
    n'est pas exactement une valeur entière d'où parfois un 'saut' brutal d'unse seconde.
    """

    imgList = sorted(imgList, key=itemgetter(2), reverse=False)  # tri par la date de prise de vue

    startTime = imgList[0][2]
    stopTime = imgList[-1][2]
    timelapseStep = (stopTime - startTime) / (len(imgList) - 1)
    timelapseStep = timedelta(seconds=timelapseStep.seconds,
                              microseconds=(timelapseStep.total_seconds() - timelapseStep.seconds) * 10 ** 6)

    new_imgList = []
    for i in range(len(imgList)):
        newDate = imgList[0][2] + i * timelapseStep
        new_imgList.append((imgList[i][0], imgList[i][1], newDate, imgList[i][2]))  # substitution de la date corrigée

    return new_imgList


# -----------------------------------     Appairement des images      -------------------------------------------------


def matchImagesFlightPath(imgListDrone,
                          deltaTimeDrone,
                          timeLapseDrone,
                          imgListIR,
                          deltaTimeIR,
                          timeLapseIR,
                          dateMission,
                          timeDeviationFactor=0.5,
                          mute=False):
    """
    :param imgListDrone:  [...,(file name, path name, date), ...]
    :param deltaTimeDrone:
    :param timeLapseDrone
    :param imgListIR:     [...,(file name, path name, date), ...]
    :param deltaTimeIR:
    :param timeLapseIR:
    :param dateMission:   date of flight
    :param timeDeviationFactor : as a percentage of the lowest time lapse
     Value 1/2 is a good choice for timing deviation |   1/4  is very selective.
    :param mute:
    :return:  listImgMatch   [..., (imgListDrone[i][1], imgListIR[k][1]), ...]
    """
    n = 0
    nRejet = 0
    listImgMatch = []
    DtImgMatch = []
    listdateMatch = []

    repA, imgListA, deltaTimeA, repB, imgListB, deltaTimeB, timeDeviation = \
        matchImagesAorB(timeLapseDrone, imgListDrone, deltaTimeDrone,
                        timeLapseIR, imgListIR, deltaTimeIR,
                        timeDeviationFactor
                        )

    dateA = [imgListA[i][2] for i in range(len(imgListA))]
    dateB = [imgListB[k][2] for k in range(len(imgListB))]
    # Shooting of image B is after shooting image A if deltaTime < 0
    deltaTime = [[datetime.timedelta.total_seconds(dateA[i] - dateB[k]) + deltaTimeB - deltaTimeA
                  for k in range(len(imgListB))] for i in range(len(imgListA))]
    DTime = [[abs(deltaTime[i][k]) for k in range(len(imgListB))] for i in range(len(imgListA))]

    for i in range(len(imgListA)):
        k = np.argmin(DTime[i][:])
        if DTime[i][k] < timeDeviation:
            kBmatch = k
            n += 1
            DtImgMatch.append(deltaTime[i][k])
            #  construction of the image pair IR & Vi  (with rectified DateTime of images)
            if timeLapseDrone >= timeLapseIR:
                listImgMatch.append((imgListB[kBmatch][1], imgListA[i][1]))
                listdateMatch.append((imgListB[kBmatch][2], imgListA[i][2]))
            else:
                listImgMatch.append((imgListA[i][1], imgListB[kBmatch][1]))
                listdateMatch.append((imgListA[i][2], imgListB[kBmatch][2]))

            if not mute:
                print(Style.GREEN + 'N°', n, ' DT ', round(deltaTime[i][kBmatch], 3), 's   ',
                      imgListA[i][0], ' | ', imgListB[kBmatch][0] + Style.RESET)

        else:
            # No match for type B image
            nRejet += 1
            if not mute: print(Style.YELLOW + 'INFO:  Image ', repA, ' ', imgListA[i][0],
                               ' does not match any image ', repB,
                               '.  Minimum time deviation = ',
                               round(min(DTime[i][:]), 3), 's' + Style.RESET)

    if timeLapseDrone >= timeLapseIR:
        print(Style.GREEN + '%i pairs of Visible-Infrared images were detected for the flight on  %s' % (
            len(listImgMatch), dateMission), '\n',
              '%i images  %s  have been eliminated :' % (nRejet, repA) + Style.RESET)
    else:
        print(Style.GREEN + '%i pairs of Visible-Infrared images were detected for the flight on  %s' % (
            len(listImgMatch), dateMission), '\n',
              '%i  images Vi (%s) have been eliminated :' % (nRejet, repA) + Style.RESET)

    if len(listImgMatch) == 0:
        print(Style.RED, 'No pair of Visible-Infrared images were detected for this flight.', Style.RESET)
        sys.exit(2)

    return listImgMatch, DtImgMatch, listdateMatch


def matchImagesAorB(timeLapseDrone, imgListDrone, deltaTimeDrone, timeLapseIR, imgListIR, deltaTimeIR,
                    timeDeviationFactor):
    if timeLapseDrone <= 0:
        # if timeLapseDrone=0 or -1 the photos are taken in manual mode
        repB = 'IR'
        repA = 'drone'
        imgListB = imgListIR
        imgListA = imgListDrone
        deltaTimeB = deltaTimeIR
        deltaTimeA = deltaTimeDrone
        timeDeviation = timeDeviationFactor * 2.1  # 2s is the weakest time lapse of the drone DJI-Mavic-Air2 (jpeg+dng)

    elif timeLapseDrone > timeLapseIR:
        # unwise !
        repA = 'drone'
        repB = 'IR'
        imgListA = imgListDrone
        imgListB = imgListIR
        deltaTimeA = deltaTimeDrone
        deltaTimeB = deltaTimeIR
        timeDeviation = timeDeviationFactor * timeLapseIR

    else:
        # timeLapseDrone =< timeLapseIR    This is the best configuration  (2s DJI | 3s SJCam M20]
        repB = 'IR'
        repA = 'drone'
        imgListB = imgListIR
        imgListA = imgListDrone
        deltaTimeB = deltaTimeIR
        deltaTimeA = deltaTimeDrone
        timeDeviation = timeDeviationFactor * timeLapseDrone

    return repA, imgListA, deltaTimeA, repB, imgListB, deltaTimeB, timeDeviation


# -------------------------     Synthèse de informations sur la mission      ------------------------------------------


def summaryFlight(listImg, DtImgMatch, listdateMatch, planVol, seaLevel=False, dirSaveFig=None, mute=True):
    """
    :param listImg:      list of image pairs VI/IR matched at the same point
    :param mute:
    :return: summary  [... (imgNameVi,imgNameIRoriginal,imgNameIR,imgNameViR, dateVi,
                          imgLat, imgLon, imgAltGround , imgAltSeaLevel ,imgAltRef,
                          distToLastPt,capToLastPt   ) ...]                   list
        imgNameVi---------------------- file name of visible RGB color image at point P
        imgNameIRoriginal-------------- file name of original infrared image  at point P
        imgNameIR---------------------- file name of infrared image  (match with visible image at point P)
        imgNameViR -------------------- file name of multispectral image at point P
        dateVi ------------------------ date the visible image was shot             [YY-MM-DD hh:mm:ss]
        imgLat, imgLon, imgAltGround    GPS coordinates of point P                  [°],[°],[m]
        imgAltSeaLevel ---------------  altitude of the drone in relation to the sealevel [m]
        imgAltRef---------------------- altitude of the drone relative to take-off point  (DJI ref)  [m]
        distToLastPt------------------- distance to next point                                       [m]
        capToLastPt-------------------- direction to the next point relative to the geographic north [°]
        UTM---------------------------- UTM coordinates of point P  xUTM (axe west>east) yUTM (axe south>north) [m],[m]
        flightAngle-------------------- yaw,pitch,roll of drone        [°]
        gimbalAngle-------------------- yaw, pitch, roll of gimbal     [°]
        speedDrone--------------------- u[0] (axe drone), u[1]                     [m/s]
        theoriticalYaw-----------------   [°]
        theoriticalPitch---------------   [°]
        theoriticalRoll----------------   [°]

        lists the essential elements of flight after treatment by IRdrone.
        This data will be saved in the Excel file of the flight plan  (sheet "Summary")

    For all points of the list, the altitudes of the ground relative to the sea level are determined..
    For France the code uses the API of the IGN (national geographical institute)
    Warning : this API limits the number of points that can be sent in the request
    An attempt is made. If the ground elevation relative to sea level fails, it is set at zero.
    https://geoservices.ign.fr/documentation/services/api-et-services-ogc/calcul-altimetrique-rest
    For each request the IGN API accepts in theory 5000 pts. In practice 189 works but 199 fails!

    """
    print(Style.CYAN + 'Request to  https://wxs.ign.fr/essentiels/alti/rest/elevation')
    summary = []
    listPtGPS, listCoordGPS, distP0P1, capP0P1 = [],[], [],[]
    altGeo, altDroneSol, altDroneSealLevel, altGround, distFlight = [], [], [], [], []
    dateVi, UTM = [], []
    flightAngle, gimbalAngle, speedDrone_1, speedDrone_2 = [], [], [], []

    maxPas = 99
    if len(listImg) < maxPas:
        pas = len(listImg) - 1
    else:
        pas = maxPas

    for i in range(len(listImg)):
        img = pr.Image(listImg[i][0])
        finfo = img.flight_info
        flightAngle.append((finfo["Flight Yaw"], finfo["Flight Pitch"], finfo["Flight Roll"]))
        gimbalAngle.append((finfo["Gimbal Yaw"], finfo["Gimbal Pitch"], finfo["Gimbal Roll"]))
        listPtGPS.append((img.gps['latitude'][4], img.gps['longitude'][4]))
        listCoordGPS.append((img.gps['latitude'][4], img.gps['longitude'][4], img.gps['altitude']))
        dateVi.append(listdateMatch[i][0])

    if seaLevel:
        # take-off point altitude relative to sea level
        coordGPSTakeOff, altiTakeOff = uGPS.TakeOff(planVol['mission']['coord GPS Take Off'])
        if altiTakeOff < 0:
            altiTakeOff = 0
        #  Splitting the list of points into multiple segments because the number of points in the IGN API is limited.
        for k in range(0, int(round(len(listImg), 0) / pas) + 1):
            listPairPtGps = listPtGPS[k * pas:(k + 1) * pas]
            altGeo += uGPS.altitude_IGN(listPairPtGps)
            print(Style.CYAN, '%i / %.0f ' % (k, len(listImg) / pas), Style.RESET)

    else:
        altGeo = [0.0] * len(listPtGPS)
        altiTakeOff = 0.

    #  Calculates the distance and heading between point P0 and the next point P1
    for i in range(len(listImg) - 1):
        imgDist, imgCap = uGPS.segmentUTM(listPtGPS[i][0], listPtGPS[i][1], listPtGPS[i + 1][0], listPtGPS[i + 1][1])
        distP0P1.append(imgDist)
        capP0P1.append(imgCap)
    distP0P1.append(0)  # no distance or heading to the next point for landing point!
    capP0P1.append(0)  # arbitrarily set the value to 0

    for i in range(len(listImg)):
        xUTM, yUTM, zoneUTM = uGPS.geo2UTMSimple(listPtGPS[i][0], listPtGPS[i][1])
        UTM.append((xUTM, yUTM, zoneUTM))
        if not mute:
            print('altitudes : drone / takeOff %f m | sol/ sea level %f m |  drone /sol %f m'
                  % (listCoordGPS[i][2], altGeo[i], round(listCoordGPS[i][2] + altiTakeOff - altGeo[i], 5)))
        # drone altitude relative to ground and sea level
        altDroneSol.append(round(listCoordGPS[i][2] + altiTakeOff - altGeo[i], 5))
        altGround.append(altGeo[i])  # ground elevation relative to sea level
        altDroneSealLevel.append(altGeo[i] + altDroneSol[i])
        distToLastPt = round(distP0P1[i], 2)  # distance to next point   (accuracy to cm!)
        if i == 0:
            distFlight.append(0.)
        else:
            distFlight.append(distFlight[i - 1] + distP0P1[i - 1])

        capToLastPt = round(capP0P1[i], 3)
        # Vi images
        imgNameViOriginal = listImg[i][0].split('\\')[len(listImg[i][0].split('\\')) - 1]
        str = imgNameViOriginal.split('_')[1]
        numeroRef = str.split('.')[0]
        imgViExt = str.split('.')[1]
        recordType = imgNameViOriginal.split('_')[0]  # DJI (single shoot), TIMELAPSE, HYPERLASE, PANORAMIC  etc
        imgNameVIS = '%s_%s_VIS.jpg' % (recordType, numeroRef)
        # IR images
        imgNameIRoriginal = listImg[i][1].split('\\')[len(listImg[i][1].split('\\')) - 1]
        imgNameNIR = '%s_%s_NIR_local_.jpg' % (recordType, numeroRef)
        imgNameViR = '_VIR_local_%s_%s.jpg ' % (recordType, numeroRef)

        summary.append(
            (i + 1,                           # 0    N° shooting point
             imgNameViOriginal,               # 1    original image VIS name
             imgNameIRoriginal,               # 2    original image NIR name
             round(DtImgMatch[i], 3),         # 3    delay between dateTime VI and dateTime IR.
             imgNameVIS,                      # 4    process image VIS name
             imgNameNIR,                      # 5    process image NIR name
             imgNameViR,                      # 6    process image ViR name
             dateVi[i].__str__(),             # 7    date of shooting (image VIS)
             round(listCoordGPS[i][0], 5),    # 8    latitude  +/- dd.ddddd   (N if > 0)
             round(listCoordGPS[i][1], 5),    # 9    longitude +/- dd.ddddd   (E if > 0)
             altDroneSol[i],                  # 10
             round(altGeo[i], 3),             # 11
             round(listCoordGPS[i][2], 3),    # 12
             round(distToLastPt, 3),          # 13
             capToLastPt,                     # 14
             round(distFlight[i], 2),         # 15
             UTM[i][0],                       # 16   xUTM
             UTM[i][1],                       # 17   yUTM
             UTM[i][2],                       # 18   zoneUTM
             round(flightAngle[i][0], 5),     # 19   yaw drone
             round(flightAngle[i][1], 5),     # 20   pitch drone
             round(flightAngle[i][2], 5),     # 21   roll drone
             round(gimbalAngle[i][0], 5),     # 22   yaw gimbal
             round(gimbalAngle[i][1], 5),     # 23   pitch gimbal
             round(gimbalAngle[i][2], 5)      # 24   roll gimbal
             )
        )
    # -----------------------------------------------------------------------------------------------------------------
    # Calcul des angles théoriques Yaw, Pitch & Roll de l'image NIR pour la superposer sur l'image VIS (coarse process)
    # -----------------------------------------------------------------------------------------------------------------
    if float(planVol['drone']['timelapse']) > 0:
        timelapse_Vis = average_Timelapse(dateVi, mute=True)
        speedDrone_1,  speedDrone_2 = vitesseDansRepereDrone(timelapse_Vis, flightAngle, UTM, mute=True)

    else:
        # ToDo  traiter le cas des prises en one shoot. CODE A VERIFIER
        # Dans ce cas la correction timelapse_Vis n'a pas de sens !
        # Le calcul de la vitesse doit être fait avec prudence. Normalement les images doivent être prises au
        # point fixe (donc à vitesse nulle). Autrement dit la baseline est théoriquement nulle.
        # Par contre il y a toujours une désynchronisation entre les images VIS et NIR.
        # On se contentera d'interpoler linéairement  ?   A SUIVRE ....
        timelapse_Vis = 2
        speedDrone_1, speedDrone_2 = [0] * len(summary), [0] * len(summary)

    theoriticalYaw = angleDeviation(summary, flightAngle, gimbalAngle, speedDrone_1, timelapse_Vis, idx=2)
    theoriticalPitch = angleDeviation(summary, flightAngle, gimbalAngle, speedDrone_2, timelapse_Vis, idx=1)
    theoriticalRoll = rollDeviation(summary, flightAngle, gimbalAngle, timelapse_Vis, idx=0)

    for i in range(len(summary)):
        summary[i] = summary[i] + (round(speedDrone_1[i], 3),         # 25 speed drone (axe drone)
                                   round(speedDrone_2[i], 3),         # 26 speed drone
                                   round(theoriticalYaw[i][1], 5),    # 27 coarse yaw for superimpose imgNIR on imgVIS
                                   round(theoriticalPitch[i][1], 5),  # 28 coarse pitch for superimpose imgNIR on imgVIS
                                   round(theoriticalRoll[i][1], 5))   # 29 coarse roll for superimpose imgNIR on imgVIS
    # -----------------------------------------------------------------------------------------------------------------
    # Affiche le graphique du profil du vol.
    # -----------------------------------------------------------------------------------------------------------------
    IRdplt.flightProfil_plot(distFlight, altDroneSealLevel, altGround, dirSaveFig=dirSaveFig, mute=False)

    if not mute:
        txtSummary = 'List of images of the flight:'
        listSummary = list(summary)
        for i in range(len(listImg)):
            txtSummary = txtSummary + '\n' + listSummary[i].__str__()
        print(Style.GREEN + txtSummary + Style.RESET)

    return summary


def writeSummaryFlight(flightPlanSynthesis, pathName, saveExcel=False):
    """

     Write the Flight Plan Summary in Excel file
    If the file not exist the file FlightSummary.xlsx  is create withe one sheet (sheetname = Summary)

    :param flightPlanSynthesis:
    :param pathName:
    :param saveExcel: save data in Excel file if True
    :return:


    """
    pathName = os.path.join(os.path.join(os.path.dirname(pathName)), 'FlightSummary.xlsx')
    if os.path.isfile(pathName):
        print(Style.CYAN + '------  Write file FlightSummary.xlsx' + Style.RESET)
        pass
    else:
        print(Style.YELLOW + '------  Create file FlightSummary.xlsx' + Style.RESET)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        wb.save(pathName)

    if saveExcel:
        workbook = openpyxl.load_workbook(pathName)
        listSheets = workbook.sheetnames
        if listSheets[0] != 'Summary':
            print(Style.YELLOW + ' Create sheet Summary' + Style.RESET)
            ws_sum = workbook.create_sheet("Summary", 0)
            ws_sum.title = "Summary"
            ws_sum.protection.sheet = False
        else:
            pass

        sheet = workbook['Summary']
        sheet.protection.sheet = False
        sheet.delete_rows(2, 1000)

        listHeadCol = headColumnSummaryFlight()
        for k in range(len(listHeadCol)):
            sheet.cell(1, k + 1).value = listHeadCol[k]
        for i in range(len(flightPlanSynthesis)):
            for k in range(len(flightPlanSynthesis[i])):
                sheet.cell(i + 2, k + 1).value = flightPlanSynthesis[i][k]
        sheet.protection.sheet = False
        workbook.save(pathName)
        workbook.close()
        print(Style.GREEN + ' %s  successfully saved.' % pathName + Style.RESET)


def readFlightSummary(dir_mission, mute=None):
    """
        Read the FlightSummary  in Excel file.

    :param dir_mission: chemin du dossier qui contient les données de la mission  type= string
    :param mute: affiche des informations si True.                                type= bool
    :return: listSummaryFlight  liste de données pour chaque couple d'images VIS-NIR (timeDeviation, altitude, etc)
    """
    summaryPath = osp.join(dir_mission, "FlightSummary.xlsx")
    txt = 'lecture des données de  ' + summaryPath
    print(Style.CYAN + txt + Style.RESET)

    workbook = openpyxl.load_workbook(summaryPath, read_only=True, data_only=True)
    sheet = workbook['Summary']
    listSummaryFlight = []
    nulg = 2  # première ligne de données
    while sheet.cell(nulg, 1).value != None:
        num_pt = int(sheet.cell(nulg, 1).value)               # numéro du point
        imgVisName = str(sheet.cell(nulg, 2).value)           # nom de l'image VIS
        imgNirName = str(sheet.cell(nulg, 3).value)           # nom de l'image VIS
        timeDeviation = float(sheet.cell(nulg, 4).value)      # ecart temporel entre image NIR et VIS (en s)
        imgNameVIS = str(sheet.cell(nulg, 5).value)           # nom de l'image VIS traitée
        imgNameNIR = str(sheet.cell(nulg, 6).value)           # nom de l'image NIR traitée
        imgNameViR = str(sheet.cell(nulg, 7).value)           # nom de l'image ViR traitée
        date_string = str(sheet.cell(nulg, 8).value)          # date image VIS (rectifiee si time-lapse)  format string
        gpsLat = float(sheet.cell(nulg, 9).value)             # latitude  dd.dddddd°
        gpsLong = float(sheet.cell(nulg, 10).value)           # longitude dd.dddddd°
        altiDrone2Sol = float(sheet.cell(nulg, 11).value)     # altitude du drone par rapport au sol (en m)
        altGeo = float(sheet.cell(nulg, 12).value)            # altitude du sol (en m)
        altiDrone2TakeOff = float(sheet.cell(nulg, 13).value)  # altitude du drone par rapport au Take-Off (en m)
        distLastPt = float(sheet.cell(nulg, 14).value)        # distance to last point (en m)
        capLastPt = float(sheet.cell(nulg, 15).value)         # cap to last point (en °)
        dist2StartPoint = float(sheet.cell(nulg, 16).value)   # distance cumulée
        xUTM = float(sheet.cell(nulg, 17).value)              # coordonnee UTM x  (East)
        yUTM = float(sheet.cell(nulg, 18).value)              # coordonnee UTM y  (North)
        zoneUTM = int(sheet.cell(nulg, 19).value)             # UTM zone
        yawDrone = float(sheet.cell(nulg, 20).value)          # Yaw drone     (roll NIR camera)
        pitchDrone = float(sheet.cell(nulg, 21).value)        # Pitch drone   (pitch NIR camera)
        rollDrone = float(sheet.cell(nulg, 22).value)         # Roll drone    (yaw NIR camera)
        yawGimbal = float(sheet.cell(nulg, 23).value)         # Yaw gimbal    (roll VIS camera)
        pitchGimbal = float(sheet.cell(nulg, 24).value)       # Pitch gimbal  (pitch VIS camera)
        rollGimbal = float(sheet.cell(nulg, 25).value)        # Roll gimbal   (yaw VIS camera)
        speed_1 = float(sheet.cell(nulg, 26).value)           # speed drone  (axe of drone)
        speed_2 = float(sheet.cell(nulg, 27).value)           # speed drone
        yawIr2Vi = float(sheet.cell(nulg, 28).value)          # theoritical yaw coarse process
        pitchIr2Vi = float(sheet.cell(nulg, 29).value)        # theoritical pitch coarse process
        rollIr2Vi = float(sheet.cell(nulg, 30).value)         # theoritical roll coarse process

        data = num_pt, imgVisName, imgNirName, timeDeviation, imgNameVIS, imgNameNIR, imgNameViR, date_string, \
               gpsLat, gpsLong, altiDrone2Sol, altGeo, altiDrone2TakeOff, distLastPt, capLastPt, dist2StartPoint, \
               xUTM, yUTM, zoneUTM, yawDrone, pitchDrone, rollDrone, yawGimbal, pitchGimbal, rollGimbal, \
               speed_1, speed_2, yawIr2Vi, pitchIr2Vi, rollIr2Vi

        listSummaryFlight.append(data)
        nulg = nulg + 1
    workbook.close()

    if not mute:
        print(listSummaryFlight)

    return listSummaryFlight


def headColumnSummaryFlight():
    listHeadCol = ['Point',
                   'Image                       Visible                 (original)',
                   'Image                       Infrared                (original)',
                   'Time deviation                       [s]',
                   'Image                    Visible aligned',
                   'Image                    Infrared aligned',
                   'Image                    Multi-Spectral ViR',
                   'Date of shooting.',
                   'Latitude                dd°ddddd',
                   'Longitude               dd°ddddd',
                   'Drone Altitude     / ground          [m]',
                   'Ground Elevation   / sea level       [m]',
                   'Drone Altitude     / Take off        [m]',
                   'Distance point suivant         [m]',
                   'Cap point suivant              [°]',
                   'Distance / Start Pt            [m]',
                   'x UTM                                [m]',
                   'y UTM                                [m]',
                   'Zone UTM',
                   'Yaw drone                            [°]',
                   'Pitch drone                          [°]',
                   'Roll drone                           [°]',
                   'Yaw Gimbal                           [°]',
                   'Pitch Gimbal                         [°]',
                   'Roll Gimbal                          [°]',
                   'u_1                                  [m/s]',
                   'u_2                                  [m/S]',
                   'Yaw                 IR to VI                     [°]',
                   'Pitch               IR to VI                     [°]',
                   'Roll                IR to VI                     [°]'
                   ]
    return listHeadCol


# -------------------------           affichage écran       -----------------------------------------------------------


def printPlanVol(planVol):
    print(' Flight at : %s  (%s)  du %s   '
          '\n Client : %s'
          '\n Pilot : %s sur drone %s'
          '\n Camera Vi : %s'
          '\n Timelapse  : %s  s    DeltaTime %s  s'
          '\n Camera IR: %s'
          '\n Timelapse  : %s  s    DeltaTime %s  s'
          '\n Infrared filter  IR %i nm'
          '\n Whether : %s  Vent %.1f m/s  Temperature %.1f °C' %
          (planVol['mission']['lieu'],
           (planVol['mission']['coord GPS Take Off'] + '  ' + str(planVol['mission']['altitude Take 0ff']) + ' m'),
           planVol['mission']['date'],
           planVol['mission']['client'],
           planVol['mission']['pilote'],
           (planVol['drone']['marque'] + ' ' + planVol['drone']['type']),
           planVol['drone']['marque'],
           planVol['drone']['timelapse'], planVol['drone']['deltatime'],
           (planVol['cameraIR']['marque'] + '  ' + planVol['cameraIR']['type']),
           planVol['cameraIR']['timelapse'], planVol['cameraIR']['deltatime'],
           planVol['images']['filtreIR'],
           planVol['meteo']['ensoleillement'], float(planVol['meteo']['vent']), float(planVol['meteo']['temperature'])
           )
          )
    return


# ------------------     Gestion des fichiers      ---------------------


def reformatDirectory(di, xlpath=None, makeOutdir=False):
    if os.path.exists(di):
        return di
    else:
        if xlpath is not None:
            newdi = os.path.join(os.path.dirname(xlpath), di)
            return reformatDirectory(newdi, xlpath=None, makeOutdir=makeOutdir)
        if makeOutdir:
            os.mkdir(di)
            print("creating output dir: %s" % di)
            return di

    raise NameError("Cannot find directory %s" % di)


def loadFileGUI(mute=True):
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename()  # show an "Open" dialog box and return the path to the selected file
    if not mute: print(filename)
    return filename


def answerYesNo(txt):
    ans = input(txt)
    tryAgain = True
    countTry = 0
    while tryAgain:
        try:
            ans = int(ans)
            if ans != 1 and ans != 0:
                raise ValueError
            else:
                ans = bool(int(ans))

            return ans

        except ValueError:
            try:
                if ans == "":
                    return
                if countTry < 2:
                    print('You need to type 0 or 1')
                    ans = input(txt)
                else:
                    return
            except:
                return ans
            else:
                countTry += 1
                tryAgain = True

def check_numeric(x):
    if not isinstance(x, (int, float, complex)):
        raise ValueError('{0} is not numeric'.format(x))


# --------   calcul a priori du pitch, yaw  & roll "grossier"  ----------


def speedmeter(UTM, timeStep=2., mute=True):
    """
    vector   U = u_EW e_EW + u_SN e_SN       |e_EW|=1, |e_SN|=1, e_EW.e_SN=0
    Axe orientation  e_EW <=> West > East ,   e_SN <=> South > North

        N  e_SN
          |
    W ----E ----> E  e_WE
          |
          S
    """
    speed_WE, speed_SN, speed = [], [], []
    for i in range(len(UTM)):
        if i == 0:
            distEW = float(UTM[1][0]) - float(UTM[0][0])
            distSN = float(UTM[1][1]) - float(UTM[0][1])
            dt = timeStep / 2
        elif i >= len(UTM) - 1:
            distEW = float(UTM[-1][0]) - float(UTM[-2][0])
            distSN = float(UTM[-1][1]) - float(UTM[-2][1])
            dt = timeStep / 2
        else:
            distEW = float(UTM[i + 1][0]) - float(UTM[i - 1][0])
            distSN = float(UTM[i + 1][1]) - float(UTM[i - 1][1])
            dt = timeStep
        u_WE = distEW / (2 * dt)
        u_SN = distSN / (2 * dt)
        dist = (distEW ** 2 + distSN ** 2) ** 0.5
        U = dist / (2 * dt)
        speed_WE.append(u_WE)
        speed_SN.append(u_SN)
        speed.append(U)
        if not mute:
            print('u_WE = ', u_WE, ' m/s     u_SN = ', u_SN, ' m/s')
    return speed_WE, speed_SN, speed


def projectSpeed(u_WE, u_SN, flightAngle, mute=True):
    """
    Repère géographique:
    e_EW vecteur  West > East          |e_EW|=1
    e_SN vecteur  South > North        |e_SN|=1     e_EW . e_SN = 0  ; attention repère indirect (e_z vers le bas!)
    direction : nord > 0°  ;  est > 90°  ;  sud > 180°  ;  ouest > 270°
    Repère Drone:
    e_2  vecteur de l'axe du drone (vers l'avant)    |e_2|=1
    e_1  vecteur normal à l'axe du drone    |e_1|=1  ;  e_1 . e_2 =0 ; repère direct ( ez vers le haut )

    """
    speedDrone_1, speedDrone_2 = [], []
    for i in range(len(flightAngle)):
        yaw_drone = flightAngle[i][0]
        a_WE = np.sin(np.deg2rad(float(yaw_drone)))
        a_SN = np.cos(np.deg2rad(float(yaw_drone)))
        u_1 = u_WE[i] * a_SN - u_SN[i] * a_WE
        u_2 = u_WE[i] * a_WE + u_SN[i] * a_SN
        speedDrone_1.append(u_WE[i] * a_SN - u_SN[i] * a_WE)
        speedDrone_2.append(u_WE[i] * a_WE + u_SN[i] * a_SN)
        if not mute:
            print('u_1 = ', u_1, ' m/s     u_2 = ', u_2, ' m/s')
    return speedDrone_1, speedDrone_2


def average_Timelapse(dateVi, mute=True):
    # période réelle du timalapse de la caméra VIS
    # Note: pour le DJI mavic Air 2 avec un timelapse théorique de 2s il faut plus de 55 images pour constater
    # un saut d'une seconde dans les dates d'enregistrement  (Timelapse réel ~ 2,018s)
    tStart = dateVi[0]
    tEnd = dateVi[-1]
    total_FlightTime = (tEnd - tStart).total_seconds()
    av_Timelapse = total_FlightTime / (len(dateVi) - 1)
    if not mute:
        txt = ' ...  Pas d\'acquisition des images VIS  ' + str(np.round(av_Timelapse, 6)) + ' s'
        print(Style.YELLOW + txt + Style.RESET)
    return av_Timelapse


def average_Speed(listSummaryFlight, mute=True):
    tStart = dateExcelString2Py(listSummaryFlight[0][7])
    tEnd = dateExcelString2Py(listSummaryFlight[-1][7])
    total_FlightTime = (tEnd - tStart).total_seconds()
    distTotal = ((float(listSummaryFlight[-1][16]) - float(listSummaryFlight[0][16])) ** 2 +
                 (float(listSummaryFlight[-1][17]) - float(listSummaryFlight[0][17])) ** 2) ** 0.5
    averageSpeed = distTotal / total_FlightTime  # vitesse du drone   en m/s
    if not mute:
        txt = ' ...  Vitesse drone ' + str(averageSpeed) + ' m/s'
        print(Style.YELLOW + txt + Style.RESET)
    return averageSpeed


def vitesseDansRepereDrone(timelapse_Vis, flightAngle, UTM, mute=True):
    u_WE, u_SN, U = speedmeter(UTM, timeStep=timelapse_Vis, mute=mute)
    speedDrone_1, speedDrone_2 = projectSpeed(u_WE, u_SN, flightAngle, mute=mute)
    return speedDrone_1, speedDrone_2


def interpolationAngle(flightAngle, summary, i, timelapse_Vis, idx=1):
    """
    Interpolation linéaire de l'angle du drone à l'instant où l'image NIR a ete prise.
    dt = tk-t où tk heure camera VIS et t heure caméra NIR.
    Interpolation "vers l'avant" si dt<0.
    """
    dt = float(summary[i][3])
    if i == 1 or i == len(summary) - 1:
        alpha = flightAngle[i][idx]
    elif dt < 0:
        alpha = (flightAngle[i][idx] * (dt / timelapse_Vis + 1) -
                 flightAngle[i + 1][idx] * dt / timelapse_Vis)
    else:
        alpha = (flightAngle[i - 1][idx] * dt / timelapse_Vis -
                 flightAngle[i][idx] * (dt / timelapse_Vis - 1))
    return alpha, dt


def angleDeviation(summary, flightAngle, gimbalAngle, u, timelapse_Vis, idx=1):
    """
    u  composante de la vitesse    u = U.e_idx
    idx=1    Yaw    (le roll du drone correspond au yaw de la caméra NIR!)
    idx=2    Pitch  (attention offset de 90° pour le DJI)
    e_2  vecteur de l'axe du drone (vers l'avant)    |e_2|=1
    e_1  vecteur normal à l'axe du drone    |e_1|=1  ;  e_1 . e_2 =0 ; repère direct
    """
    angle_Theorique = []
    for i in range(len(summary)):
        alpha, dt = interpolationAngle(flightAngle, summary, i, timelapse_Vis, idx=idx)
        dateImgVis = dateExcelString2Py(summary[i][7])
        CnirCvis = u[i] * dt
        H = float(summary[i][10])

        if idx == 2:
            thetaVis = gimbalAngle[i][idx]            # Roll Gimbal <=>  Yaw Camera VIS
        else:
            thetaVis = gimbalAngle[i][idx] + 90.      # Pitch Gimbal <=> Pitch Camera VIS
        anglePhi = np.rad2deg(np.arctan(CnirCvis / H + np.tan(np.deg2rad(thetaVis))))

        anglePsi = anglePhi - alpha

        data = dateImgVis, anglePsi
        angle_Theorique.append(data)
        # print("CnirCvis  ", CnirCvis, " H   ", H, "anglePhi ", anglePhi , "anglePsi   ",anglePsi)

    return angle_Theorique


def rollDeviation(summary, flightAngle, gimbalAngle,  timelapse_Vis, idx=0):
    """
    yaw drone  <=> roll NIR camera
    yaw gimbal <=> roll VIS camera
    """
    theoriticalRoll = []
    for i in range(len(summary)):
        rollInterpol_NIR, dt = interpolationAngle(flightAngle, summary, i, timelapse_Vis, idx=idx)
        rollInterpol_VIS, dt = interpolationAngle(gimbalAngle, summary, i, timelapse_Vis, idx=idx)
        dateImgVis = dateExcelString2Py(summary[i][7])
        data = dateImgVis, rollInterpol_NIR - rollInterpol_VIS
        theoriticalRoll.append(data)
    return theoriticalRoll